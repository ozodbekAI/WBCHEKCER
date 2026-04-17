#!/usr/bin/env python3
from __future__ import annotations

import argparse
import asyncio
import json
import sys
from datetime import date, timedelta, datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence

import httpx
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from sqlalchemy import select
from app.core.config import settings
from app.core.database import AsyncSessionLocal
from app.models import Store
from app.services.sku_economics_service import _chunks, _safe_positive, _to_float, _to_int, sku_economics_service
from app.services.wb_advert_repository import WBAdvertRepository
from app.services.wb_token_access import get_store_feature_api_key

REPORTS_DIR = PROJECT_ROOT / "reports"

COMMON_HEADERS = ["source", "endpoint", "status", "message", "raw_json"]
CONTENT_HEADERS = [
    "source",
    "nm_id",
    "vendor_code",
    "subject_id",
    "title",
    "brand",
    "with_photo",
    "stock_count",
    "sales_priority",
    "raw_json",
]
PROMOTION_HEADERS = [
    "source",
    "period_start",
    "period_end",
    "advert_id",
    "campaign_title",
    "campaign_total_spend",
    "raw_json",
]
FULLSTATS_HEADERS = [
    "source",
    "period_start",
    "period_end",
    "metric_date",
    "advert_id",
    "campaign_title",
    "campaign_total_spend",
    "nm_id",
    "nm_title",
    "views",
    "clicks",
    "ad_orders",
    "ad_gmv",
    "ad_spend",
    "raw_json",
]
FUNNEL_HEADERS = [
    "source",
    "period_start",
    "period_end",
    "nm_id",
    "nm_title",
    "open_count",
    "cart_count",
    "order_count",
    "order_sum",
    "buyout_count",
    "buyout_sum",
    "raw_json",
]
FINANCE_HEADERS = [
    "source",
    "period_start",
    "period_end",
    "metric_date",
    "nm_id",
    "doc_type",
    "quantity",
    "revenue",
    "payout",
    "wb_extra_costs",
    "is_return",
    "raw_json",
]


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    return date.fromisoformat(str(value).strip())


def _default_period(days: int) -> tuple[date, date]:
    period_end = date.today()
    period_start = period_end - timedelta(days=max(int(days or 14), 1) - 1)
    return period_start, period_end


def _json_cell(value: Any) -> str:
    try:
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    except TypeError:
        return json.dumps(str(value), ensure_ascii=False)


def _coerce_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Да" if value else "Нет"
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if isinstance(value, (dict, list, tuple)):
        return _json_cell(value)
    return value


def _retry_delay_from_response(resp: httpx.Response, attempt: int) -> float:
    retry_after = str(resp.headers.get("Retry-After") or "").strip()
    if retry_after:
        try:
            return min(max(float(retry_after), 1.0), 20.0)
        except Exception:
            pass
    return min(2.0 * (attempt + 1), 12.0)


async def _request_with_retry(
    client: httpx.AsyncClient,
    method: str,
    url: str,
    *,
    headers: Dict[str, str],
    params: Optional[Dict[str, Any]] = None,
    json_body: Optional[Dict[str, Any]] = None,
    retries: int = 4,
) -> httpx.Response:
    last_response = None
    for attempt in range(max(int(retries), 1)):
        response = await client.request(
            method,
            url,
            headers=headers,
            params=params,
            json=json_body,
        )
        last_response = response
        if response.status_code not in {429, 502, 503, 504} or attempt >= retries - 1:
            return response
        await asyncio.sleep(_retry_delay_from_response(response, attempt))
    if last_response is None:
        raise RuntimeError(f"{url} request failed before response")
    return last_response


async def _load_store(store_id: int) -> Optional[Store]:
    async with AsyncSessionLocal() as db:
        result = await db.execute(
            select(Store).where(Store.id == int(store_id))
        )
        return result.scalar_one_or_none()


def _candidate_tokens(store: Optional[Store], *, prefer_advert: bool, explicit_api_key: str | None = None) -> List[str]:
    seen: set[str] = set()
    out: list[str] = []

    candidates: list[str | None] = [
        explicit_api_key,
    ]

    if store is not None:
        feature_token = str(get_store_feature_api_key(store, "ad_analysis") or "").strip()
        candidates.extend([
            feature_token,
            settings.WB_ADVERT_API_KEY if prefer_advert else None,
            str(store.api_key or "").strip(),
            settings.WB_API_KEY,
            settings.WB_ADVERT_API_KEY,
        ])
    else:
        candidates.extend([settings.WB_API_KEY, settings.WB_ADVERT_API_KEY])

    for candidate in candidates:
        token = str(candidate or "").strip()
        if token and token not in seen:
            out.append(token)
            seen.add(token)
    return out


async def _fetch_common_rows(token: str, store_id: int | None = None) -> list[dict[str, Any]]:
    headers = {
        "Authorization": token,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    rows: list[dict[str, Any]] = []
    async with httpx.AsyncClient(timeout=httpx.Timeout(20.0)) as client:
        for endpoint in [
            "/ping",
            "/api/v1/seller-info",
        ]:
            url = f"{settings.WB_COMMON_API_URL}{endpoint}"
            response = await _request_with_retry(client, "GET", url, headers=headers, retries=4)
            if response.status_code >= 400:
                rows.append({
                    "source": "common",
                    "endpoint": endpoint,
                    "status": f"error_{response.status_code}",
                    "message": response.text[:500],
                    "raw_json": response.text[:12000],
                })
                continue
            payload = response.json() if response.text else {}
            rows.append({
                "source": "common",
                "endpoint": endpoint,
                "status": "ok",
                "message": f"store_id={store_id}" if store_id else "no_store",
                "raw_json": payload,
            })
    return rows


def _parse_content_cards_row(card: dict[str, Any], *, period_start: date, period_end: date) -> dict[str, Any]:
    return {
        "source": "content",
        "nm_id": _to_int(card.get("nmID") or card.get("nm_id")),
        "vendor_code": str(card.get("vendorCode") or "").strip(),
        "subject_id": _to_int(card.get("subjectID") or card.get("subjectId")),
        "title": str(card.get("title") or card.get("name") or "").strip(),
        "brand": str(card.get("brand") or card.get("brandName") or "").strip(),
        "with_photo": bool(card.get("photos") or card.get("photoCount")),
        "stock_count": _to_int(card.get("stock") or card.get("stocks") or card.get("quantity")),
        "sales_priority": str(card.get("salesPriority") or ""),
        "raw_json": card,
    }


async def _fetch_content_rows(token: str, limit: int, max_pages: int) -> list[dict[str, Any]]:
    if not token or limit <= 0 or max_pages <= 0:
        return []

    rows: list[dict[str, Any]] = []
    headers = {
        "Authorization": token,
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    cursor: dict[str, Any] = {}
    async with httpx.AsyncClient(timeout=httpx.Timeout(45.0)) as client:
        for page in range(max_pages):
            body: dict[str, Any] = {
                "settings": {
                    "sort": {"ascending": False},
                    "filter": {"withPhoto": 1},
                    "cursor": {
                        "limit": int(limit),
                        **({"updatedAt": cursor["updatedAt"]} if cursor.get("updatedAt") else {}),
                        **({"nmID": cursor["nmID"]} if cursor.get("nmID") is not None else {}),
                    },
                }
            }
            url = f"{settings.WB_CONTENT_API_URL}/content/v2/get/cards/list"
            response = await _request_with_retry(
                client,
                "POST",
                url,
                headers=headers,
                json_body=body,
                retries=4,
            )
            if response.status_code >= 400:
                raise RuntimeError(f"content/cards/list {response.status_code}: {response.text[:300]}")

            data = response.json() if response.text else {}
            cards = data.get("cards") or data.get("data", {}).get("cards") or []
            if not isinstance(cards, list):
                cards = []
            for card in cards:
                if isinstance(card, dict):
                    rows.append(_parse_content_cards_row(card, period_start=date.today(), period_end=date.today()))

            cursor_block = data.get("cursor") or {}
            if not cursor_block:
                break
            cursor_updated = cursor_block.get("updatedAt")
            cursor_nm = cursor_block.get("nmID")
            if not cursor_updated and cursor_nm is None:
                break
            if not cards:
                break
            cursor = {"updatedAt": cursor_updated, "nmID": cursor_nm}
            if not cursor_updated and cursor_nm is None:
                break
    return rows


def _campaign_title(payload: dict[str, Any]) -> str:
    return str(
        payload.get("advertName")
        or payload.get("advert_name")
        or payload.get("name")
        or payload.get("advertTitle")
        or ""
    ).strip()


async def _fetch_promotion_count_rows(tokens: Sequence[str]) -> tuple[list[dict[str, Any]], str]:
    if not tokens:
        raise RuntimeError("No advert token available")

    last_error = "No advert token available"
    url = f"{settings.WB_ADVERT_API_URL}/adv/v1/promotion/count"
    for token in tokens:
        try:
            async with httpx.AsyncClient(timeout=httpx.Timeout(25.0)) as client:
                response = await client.get(
                    url,
                    headers={
                        "Authorization": token,
                        "Content-Type": "application/json",
                        "Accept": "application/json",
                    },
                )
            if response.status_code >= 400:
                raise RuntimeError(f"promotion/count {response.status_code}: {response.text[:300]}")
            payload = response.json() if response.text else []
            groups = payload if isinstance(payload, list) else (payload.get("adverts") or [])
            rows: list[dict[str, Any]] = []
            for group in groups:
                if not isinstance(group, dict):
                    continue
                advert_list = group.get("advert_list") or group.get("advertList") or []
                if not isinstance(advert_list, list):
                    continue
                for advert in advert_list:
                    if not isinstance(advert, dict):
                        continue
                    rows.append({
                        "source": "advert_count",
                        "period_start": date.today().isoformat(),
                        "period_end": date.today().isoformat(),
                        "advert_id": _to_int(advert.get("advertId") or advert.get("advert_id") or advert.get("id")),
                        "campaign_title": str(advert.get("name") or advert.get("advertName") or "").strip(),
                        "campaign_total_spend": _to_float(advert.get("sum")),
                        "raw_json": _json_cell(advert),
                    })
            return rows, token
        except Exception as exc:
            last_error = str(exc)
    raise RuntimeError(f"WB promotion/count export failed: {last_error}")


def _collect_fullstats_leaf_rows(
    node: dict[str, Any],
    *,
    campaign_item: dict[str, Any],
    period_start: date,
    period_end: date,
    out: list[dict[str, Any]],
    day_node: dict[str, Any] | None = None,
    app_node: dict[str, Any] | None = None,
) -> None:
    day_values = node.get("days")
    if isinstance(day_values, list) and day_values:
        for child in day_values:
            if isinstance(child, dict):
                _collect_fullstats_leaf_rows(
                    child,
                    campaign_item=campaign_item,
                    period_start=period_start,
                    period_end=period_end,
                    out=out,
                    day_node=child,
                    app_node=None,
                )
        return

    app_values = node.get("apps")
    if isinstance(app_values, list) and app_values:
        for child in app_values:
            if isinstance(child, dict):
                _collect_fullstats_leaf_rows(
                    child,
                    campaign_item=campaign_item,
                    period_start=period_start,
                    period_end=period_end,
                    out=out,
                    day_node=day_node,
                    app_node=child,
                )
        return

    nms = node.get("nms")
    if not isinstance(nms, list):
        return

    metric_date = (
        sku_economics_service._parse_generic_date(
            (day_node or {}).get("date")
            or (day_node or {}).get("day")
            or (day_node or {}).get("dt")
            or (day_node or {}).get("begin")
            or (day_node or {}).get("beginDate")
        )
    )

    advert_id = _to_int(campaign_item.get("advertId") or campaign_item.get("advert_id") or campaign_item.get("id"))
    campaign_name = _campaign_title(campaign_item)
    campaign_total_spend = _to_float(campaign_item.get("sum"))

    for nm_row in nms:
        if not isinstance(nm_row, dict):
            continue
        row = {
            "source": "fullstats",
            "period_start": period_start,
            "period_end": period_end,
            "metric_date": metric_date,
            "advert_id": advert_id or "",
            "campaign_title": campaign_name,
            "campaign_total_spend": round(campaign_total_spend, 2),
            "nm_id": _to_int(nm_row.get("nmId") or nm_row.get("nm_id")),
            "nm_title": str(nm_row.get("name") or "").strip(),
            "views": _to_int(nm_row.get("views")),
            "clicks": _to_int(nm_row.get("clicks")),
            "ad_orders": _to_int(nm_row.get("orders")),
            "ad_gmv": round(_to_float(nm_row.get("sum_price")), 2),
            "ad_spend": round(_to_float(nm_row.get("sum")), 2),
            "raw_json": _json_cell({
                "campaign": campaign_item,
                "day": day_node,
                "app": app_node,
                "nm": nm_row,
            }),
        }
        out.append(row)


async def _fetch_fullstats_rows(
    token: str,
    campaign_ids: Sequence[int],
    *,
    period_start: date,
    period_end: date,
) -> list[dict[str, Any]]:
    if not campaign_ids:
        return []
    repo = WBAdvertRepository(token=token)
    out: list[dict[str, Any]] = []
    for batch in _chunks(list(campaign_ids), 50):
        batch_data = await asyncio.to_thread(
            repo.get_fullstats,
            batch,
            begin_date=period_start.isoformat(),
            end_date=period_end.isoformat(),
        )
        if not isinstance(batch_data, list):
            continue
        for item in batch_data:
            if not isinstance(item, dict):
                continue
            leaf_before = len(out)
            _collect_fullstats_leaf_rows(
                item,
                campaign_item=item,
                period_start=period_start,
                period_end=period_end,
                out=out,
            )
            if len(out) == leaf_before:
                row = {
                    "source": "fullstats",
                    "period_start": period_start,
                    "period_end": period_end,
                    "metric_date": "",
                    "advert_id": _to_int(item.get("advertId") or item.get("advert_id") or item.get("id")) or "",
                    "campaign_title": _campaign_title(item),
                    "campaign_total_spend": round(_to_float(item.get("sum")), 2),
                    "nm_id": "",
                    "nm_title": "",
                    "views": "",
                    "clicks": "",
                    "ad_orders": "",
                    "ad_gmv": "",
                    "ad_spend": "",
                    "raw_json": _json_cell(item),
                }
                out.append(row)
    return out


def _extract_campaign_ids(rows: Sequence[dict[str, Any]]) -> list[int]:
    ids: set[int] = set()
    for row in rows:
        value = row.get("advert_id")
        try:
            if value:
                ids.add(int(value))
        except Exception:
            pass
    return sorted(ids)


def _extract_nm_ids(rows: Sequence[dict[str, Any]]) -> list[int]:
    ids: set[int] = set()
    for row in rows:
        nm_id = _to_int(row.get("nm_id"))
        if nm_id > 0:
            ids.add(nm_id)
    return sorted(ids)


async def _fetch_funnel_rows(
    tokens: Sequence[str],
    *,
    period_start: date,
    period_end: date,
    nm_ids: Sequence[int],
) -> list[dict[str, Any]]:
    if not nm_ids:
        return []
    url = f"{settings.WB_ANALYTICS_API_URL}/api/analytics/v3/sales-funnel/products"
    last_error = "No analytics token available"

    for token in tokens:
        try:
            async with httpx.AsyncClient(timeout=httpx.Timeout(60.0)) as client:
                out: list[dict[str, Any]] = []
                for batch in _chunks(list(nm_ids), 500):
                    payload = {
                        "selectedPeriod": {
                            "start": period_start.isoformat(),
                            "end": period_end.isoformat(),
                        },
                        "nmIds": batch,
                        "limit": len(batch),
                        "offset": 0,
                        "skipDeletedNm": False,
                        "orderBy": {"field": "openCard", "mode": "desc"},
                    }
                    response = await _request_with_retry(
                        client,
                        "POST",
                        url,
                        headers={
                            "Authorization": token,
                            "Content-Type": "application/json",
                            "Accept": "application/json",
                        },
                        json_body=payload,
                    )
                    if response.status_code >= 400:
                        raise RuntimeError(f"sales-funnel/products {response.status_code}: {response.text[:300]}")
                    data = response.json() or {}
                    products = ((data.get("data") or {}).get("products") or [])
                    for product_row in products:
                        if not isinstance(product_row, dict):
                            continue
                        product = product_row.get("product") or {}
                        stats = ((product_row.get("statistic") or {}).get("selected") or {})
                        out.append({
                            "source": "analytics_sales_funnel_products",
                            "period_start": period_start,
                            "period_end": period_end,
                            "nm_id": _to_int(product.get("nmId")),
                            "nm_title": str(product.get("title") or "").strip(),
                            "open_count": _to_int(stats.get("openCount")),
                            "cart_count": _to_int(stats.get("cartCount")),
                            "order_count": _to_int(stats.get("orderCount")),
                            "order_sum": round(_to_float(stats.get("orderSum")), 2),
                            "buyout_count": _to_int(stats.get("buyoutCount")),
                            "buyout_sum": round(_to_float(stats.get("buyoutSum")), 2),
                            "raw_json": _json_cell(product_row),
                        })
                return out
        except Exception as exc:
            last_error = str(exc)
    raise RuntimeError(f"WB sales-funnel/products export failed: {last_error}")


async def _fetch_finance_rows(
    tokens: Sequence[str],
    *,
    period_start: date,
    period_end: date,
) -> list[dict[str, Any]]:
    url = f"{settings.WB_STATISTICS_API_URL}/api/v5/supplier/reportDetailByPeriod"
    last_error = "No statistics token available"
    for token in tokens:
        try:
            params = {
                "dateFrom": period_start.isoformat(),
                "dateTo": period_end.isoformat(),
                "limit": 100000,
                "rrdid": 0,
                "period": "daily",
            }
            async with httpx.AsyncClient(timeout=httpx.Timeout(90.0)) as client:
                response = await _request_with_retry(
                    client,
                    "GET",
                    url,
                    headers={"Authorization": token, "Accept": "application/json"},
                    params=params,
                    retries=5,
                )
            if response.status_code == 204:
                return []
            if response.status_code >= 400:
                raise RuntimeError(f"reportDetailByPeriod {response.status_code}: {response.text[:300]}")

            payload = response.json() or []
            out: list[dict[str, Any]] = []
            for item in payload:
                if not isinstance(item, dict):
                    continue
                quantity = _to_int(item.get("quantity"))
                doc_type = str(
                    item.get("doc_type_name") or item.get("supplier_oper_name") or item.get("docType") or ""
                ).strip()
                is_return = ("возврат" in doc_type.lower()) or ("return" in doc_type.lower()) or quantity < 0
                extra_costs = sum(
                    _to_float(item.get(field))
                    for field in (
                        "delivery_rub",
                        "acquiring_fee",
                        "penalty",
                        "storage_fee",
                        "deduction",
                        "acceptance",
                        "rebill_logistic_cost",
                        "additional_payment",
                    )
                )
                out.append({
                    "source": "finance_report_detail",
                    "period_start": period_start,
                    "period_end": period_end,
                    "metric_date": sku_economics_service._parse_generic_date(
                        item.get("date_from")
                        or item.get("sale_dt")
                        or item.get("rr_dt")
                        or item.get("create_dt"),
                    ),
                    "nm_id": _to_int(item.get("nm_id") or item.get("nmId")),
                    "doc_type": doc_type,
                    "quantity": quantity,
                    "revenue": round(_to_float(item.get("retail_price_withdisc_rub") or item.get("retail_amount")), 2),
                    "payout": round(_to_float(item.get("ppvz_for_pay")), 2),
                    "wb_extra_costs": round(_safe_positive(extra_costs), 2),
                    "is_return": is_return,
                    "raw_json": _json_cell(item),
                })
            return out
        except Exception as exc:
            last_error = str(exc)
    raise RuntimeError(f"WB reportDetailByPeriod export failed: {last_error}")


def _auto_width(ws) -> None:
    for column_cells in ws.columns:
        max_len = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[column].width = min(max(max_len + 2, 8), 80)


def _style_header(ws) -> None:
    fill = PatternFill("solid", fgColor="E7EEF9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _write_sheet(ws, rows: Sequence[dict[str, Any]], headers: Sequence[str]) -> None:
    ws.append(list(headers))
    for row in rows:
        ws.append([_coerce_cell(row.get(h, "")) for h in headers])
    _style_header(ws)
    _auto_width(ws)


def _add_readme_sheet(wb, period_start: date, period_end: date, stats: Dict[str, int]) -> None:
    ws = wb.create_sheet("00_readme")
    ws.append(["Параметр", "Значение"])
    rows = [
        ["Период", f"{period_start.isoformat()} — {period_end.isoformat()}"],
        ["Сформировано", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Common rows", stats.get("common", 0)],
        ["Content rows", stats.get("content", 0)],
        ["Promotion count rows", stats.get("promotion", 0)],
        ["Fullstats rows", stats.get("fullstats", 0)],
        ["Funnel rows", stats.get("funnel", 0)],
        ["Finance rows", stats.get("finance", 0)],
        ["", ""],
        ["Источник", "Описание"],
        ["common", "ping + seller-info"],
        ["content", "Карточки из /content/v2/get/cards/list"],
        ["promotion_count", "Список рекламных кампаний"],
        ["fullstats", "Суточная статистика рекламных кампаний"],
        ["sales_funnel_products", "Воронка товаров (открытия/корзины/заказы/выручка)"],
        ["finance_report_detail", "Финансы по периодам (reportDetailByPeriod)"],
        ["", ""],
        ["Как читать", "Скрипт экспортирует сырые поля из WB API + основные вычислимые поля для удобного анализа."],
        ["", "Формулы по сайту можно смотреть в scripts/export_wb_api_inputs.py (раздел formulas)."],
    ]
    for r in rows:
        ws.append(r)
    _style_header(ws)
    _auto_width(ws)


async def run_export(
    *,
    store_id: int | None,
    explicit_api_key: str | None,
    period_start: date,
    period_end: date,
    content_limit: int,
    content_pages: int,
    output_path: Path,
) -> Path:
    store: Store | None = None
    if store_id is not None:
        store = await _load_store(store_id)
        if store is None:
            raise RuntimeError(f"Store {store_id} not found")

    print("Загрузка common (ping + seller-info)...", flush=True)
    common_token = explicit_api_key or settings.WB_API_KEY or settings.WB_ADVERT_API_KEY
    if not common_token:
        raise RuntimeError("WB API token topilmadi (settings.WB_API_KEY yoki --api-key kerak).")
    common_rows = await _fetch_common_rows(common_token, store_id=store_id)

    print("Загрузка content cards (картачки)...", flush=True)
    content_token = (
        _candidate_tokens(store, prefer_advert=False, explicit_api_key=explicit_api_key)[0]
        if _candidate_tokens(store, prefer_advert=False, explicit_api_key=explicit_api_key)
        else explicit_api_key or common_token
    )
    try:
        content_rows = await _fetch_content_rows(content_token, limit=content_limit, max_pages=content_pages)
    except Exception as exc:
        content_rows = [{
            "source": "content",
            "nm_id": "",
            "vendor_code": "",
            "subject_id": "",
            "title": "",
            "brand": "",
            "with_photo": "",
            "stock_count": "",
            "sales_priority": "",
            "raw_json": str(exc),
        }]

    advert_tokens = _candidate_tokens(store, prefer_advert=True, explicit_api_key=explicit_api_key)
    print("Загрузка promotion/count...", flush=True)
    promotion_rows, advert_token = await _fetch_promotion_count_rows(advert_tokens)
    campaign_ids = _extract_campaign_ids(promotion_rows)

    print("Загрузка advert fullstats...", flush=True)
    fullstats_rows = await _fetch_fullstats_rows(
        advert_token,
        campaign_ids,
        period_start=period_start,
        period_end=period_end,
    )

    nm_ids = _extract_nm_ids(fullstats_rows)
    finance_tokens = _candidate_tokens(store, prefer_advert=False, explicit_api_key=explicit_api_key)

    print("Загрузка funnel products...", flush=True)
    try:
        funnel_rows = await _fetch_funnel_rows(
            finance_tokens,
            period_start=period_start,
            period_end=period_end,
            nm_ids=nm_ids,
        )
    except Exception as exc:
        funnel_rows = [{
            "source": "sales_funnel_products",
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
            "nm_id": "",
            "nm_title": "",
            "open_count": "",
            "cart_count": "",
            "order_count": "",
            "order_sum": "",
            "buyout_count": "",
            "buyout_sum": "",
            "raw_json": str(exc),
        }]

    print("Загрузка reportDetailByPeriod...", flush=True)
    try:
        finance_rows = await _fetch_finance_rows(
            finance_tokens,
            period_start=period_start,
            period_end=period_end,
        )
    except Exception as exc:
        finance_rows = [{
            "source": "finance_report_detail",
            "period_start": period_start.isoformat(),
            "period_end": period_end.isoformat(),
            "metric_date": "",
            "nm_id": "",
            "doc_type": "",
            "quantity": "",
            "revenue": "",
            "payout": "",
            "wb_extra_costs": "",
            "is_return": "",
            "raw_json": str(exc),
        }]

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    stats = {
        "common": len(common_rows),
        "content": len(content_rows),
        "promotion": len(promotion_rows),
        "fullstats": len(fullstats_rows),
        "funnel": len(funnel_rows),
        "finance": len(finance_rows),
    }
    _add_readme_sheet(wb, period_start, period_end, stats)

    ws_common = wb.create_sheet("01_common")
    _write_sheet(ws_common, common_rows, COMMON_HEADERS)

    ws_content = wb.create_sheet("02_content_cards")
    _write_sheet(ws_content, content_rows, CONTENT_HEADERS)

    ws_promo = wb.create_sheet("03_promotion_count")
    _write_sheet(ws_promo, promotion_rows, PROMOTION_HEADERS)

    ws_full = wb.create_sheet("04_ad_fullstats")
    _write_sheet(ws_full, fullstats_rows, FULLSTATS_HEADERS)

    ws_funnel = wb.create_sheet("05_funnel_products")
    _write_sheet(ws_funnel, funnel_rows, FUNNEL_HEADERS)

    ws_fin = wb.create_sheet("06_finance_report")
    _write_sheet(ws_fin, finance_rows, FINANCE_HEADERS)

    wb.save(output_path)
    return output_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="WB API ma’lumotlarini bitta Excel faylga export qiladi."
    )
    parser.add_argument("--store-id", type=int, default=None, help="DBdagi store ID (agar bo‘lsa, token olish uchun).")
    parser.add_argument("--api-key", type=str, default=None, help="WB API tokeni. Agar berilmasa store-id dan olinadi.")
    parser.add_argument("--period-start", type=str, default=None, help="YYYY-MM-DD")
    parser.add_argument("--period-end", type=str, default=None, help="YYYY-MM-DD")
    parser.add_argument("--days", type=int, default=14, help="Agar start/end berilmasa period uzunligi (default 14).")
    parser.add_argument(
        "--content-page-size",
        type=int,
        default=25,
        help="content/v2/get/cards/list da bir sahifadagi kartalar soni (default 25).",
    )
    parser.add_argument(
        "--content-pages",
        type=int,
        default=4,
        help="content cards uchun maksimal sahifa soni (default 4).",
    )
    parser.add_argument(
        "--output",
        type=str,
        default=None,
        help="Excel fayl yo‘li. Default: reports/wb_export_<start>_<end>.xlsx",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    period_start = _parse_date(args.period_start)
    period_end = _parse_date(args.period_end)
    if period_start and not period_end:
        period_end = period_start
    if period_end and not period_start:
        period_start = period_end
    if period_start is None or period_end is None:
        period_start, period_end = _default_period(args.days)
    if period_end < period_start:
        raise SystemExit("period-end period-start dan kichik bo'lishi mumkin emas.")

    default_output = REPORTS_DIR / f"wb_export_{period_start.isoformat()}_{period_end.isoformat()}.xlsx"
    output_path = Path(args.output or default_output).expanduser().resolve()

    out = asyncio.run(
        run_export(
            store_id=args.store_id,
            explicit_api_key=(args.api_key or "").strip() or None,
            period_start=period_start,
            period_end=period_end,
            content_limit=max(1, int(args.content_page_size)),
            content_pages=max(1, int(args.content_pages)),
            output_path=output_path,
        )
    )
    print(f"Таймер: {out}", flush=True)


if __name__ == "__main__":
    main()
