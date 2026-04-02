from __future__ import annotations

import argparse
import asyncio
import sys
from datetime import date, timedelta
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from app.core.database import AsyncSessionLocal
from app.models import Store
from app.services.sku_economics_service import (
    _AdvertMetrics,
    _FinanceMetrics,
    _FunnelMetrics,
    sku_economics_service,
)


REPORT_DIR = Path("reports")


def _round2(value: float) -> float:
    return round(float(value or 0.0), 2)


def _copy_advert(metric: _AdvertMetrics | None) -> _AdvertMetrics:
    metric = metric or _AdvertMetrics()
    return _AdvertMetrics(
        views=int(metric.views or 0),
        clicks=int(metric.clicks or 0),
        orders=int(metric.orders or 0),
        gmv=float(metric.gmv or 0.0),
        exact_spend=float(metric.exact_spend or 0.0),
        estimated_spend=float(metric.estimated_spend or 0.0),
        manual_spend=float(metric.manual_spend or 0.0),
        manual_views=int(metric.manual_views or 0),
        manual_clicks=int(metric.manual_clicks or 0),
        manual_orders=int(metric.manual_orders or 0),
        manual_gmv=float(metric.manual_gmv or 0.0),
    )


def _auto_width(worksheet) -> None:
    for column_cells in worksheet.columns:
        values = [str(cell.value or "") for cell in column_cells]
        width = max((len(value) for value in values), default=12) + 2
        worksheet.column_dimensions[column_cells[0].column_letter].width = min(max(width, 12), 42)


def _style_header(worksheet) -> None:
    fill = PatternFill("solid", fgColor="E8EEF9")
    bold = Font(bold=True)
    for cell in worksheet[1]:
        cell.font = bold
        cell.fill = fill


def _iter_source_rows(mapping: dict[int, object], nm_ids: Iterable[int]) -> list[int]:
    return [int(nm_id) for nm_id in sorted(set(int(nm_id) for nm_id in nm_ids)) if int(nm_id) in mapping]


async def run_audit(store_id: int, period_start: date, period_end: date, force_overview: bool) -> tuple[Path, Path]:
    REPORT_DIR.mkdir(exist_ok=True)

    async with AsyncSessionLocal() as db:
        store = (
            await db.execute(select(Store).where(Store.id == int(store_id)))
        ).scalar_one_or_none()
        if not store:
            raise SystemExit(f"Store {store_id} not found")

        print("Loading overview...", flush=True)
        overview = await sku_economics_service.build_overview(
            db,
            store,
            period_start=period_start,
            period_end=period_end,
            preset="custom",
            page=1,
            page_size=1000,
            force=force_overview,
        )

        nm_ids = [int(item.nm_id) for item in overview.items]
        print(f"Overview loaded: {len(nm_ids)} SKU", flush=True)
        print("Fetching advert API...", flush=True)
        advert_data, advert_status = await sku_economics_service._fetch_advert_metrics(store, period_start, period_end)
        print("Fetching finance API...", flush=True)
        finance_data, finance_status = await sku_economics_service._fetch_finance_metrics(store, period_start, period_end)
        print("Fetching funnel API...", flush=True)
        funnel_data, funnel_status = await sku_economics_service._fetch_funnel_metrics(store, period_start, period_end, nm_ids)
        print("Loading manual layers...", flush=True)
        manual_costs = await sku_economics_service._load_costs(db, int(store_id))
        manual_spend = await sku_economics_service._load_manual_spend(db, int(store_id), period_start, period_end)
        manual_finance = await sku_economics_service._load_manual_finance(db, int(store_id), period_start, period_end)

        items_by_nm = {int(item.nm_id): item for item in overview.items}
        verification_rows: list[dict[str, object]] = []
        failures: list[dict[str, object]] = []

        for nm_id in sorted(nm_ids):
            item = items_by_nm[nm_id]
            advert = _copy_advert(advert_data["per_nm"].get(nm_id))
            manual_topup = manual_spend.get(nm_id)
            if manual_topup:
                advert.manual_spend += float(manual_topup.spend or 0.0)
                advert.manual_views += int(manual_topup.views or 0)
                advert.manual_clicks += int(manual_topup.clicks or 0)
                advert.manual_orders += int(manual_topup.orders or 0)
                advert.manual_gmv += float(manual_topup.gmv or 0.0)

            finance = finance_data.get(nm_id, _FinanceMetrics())
            if nm_id in manual_finance:
                finance = manual_finance[nm_id]
            funnel = funnel_data.get(nm_id, _FunnelMetrics())

            total_orders = max(int(finance.orders or 0), int(funnel.order_count or 0), int(advert.total_orders or 0))
            revenue = float(finance.revenue or 0.0) or float(manual_finance.get(nm_id, _FinanceMetrics()).revenue or 0.0) or float(funnel.order_sum or 0.0) or float(advert.total_gmv or 0.0)
            wb_costs = float(finance.wb_costs or 0.0)
            unit_cost = float(manual_costs.get(nm_id) or 0.0)
            cost_price = unit_cost * max(total_orders, 0)
            gross_profit_before_ads = revenue - wb_costs - cost_price
            ad_cost = float(advert.total_spend or 0.0)
            net_profit = gross_profit_before_ads - ad_cost
            max_cpo = gross_profit_before_ads / total_orders if total_orders > 0 else gross_profit_before_ads
            actual_cpo = ad_cost / advert.total_orders if advert.total_orders > 0 else ad_cost
            profit_delta = max_cpo - actual_cpo

            diff_cost_price = _round2(cost_price - item.metrics.cost_price)
            diff_gross_profit = _round2(gross_profit_before_ads - item.metrics.gross_profit_before_ads)
            diff_net_profit = _round2(net_profit - item.metrics.net_profit)
            diff_max_cpo = _round2(max_cpo - item.metrics.max_cpo)
            diff_actual_cpo = _round2(actual_cpo - item.metrics.actual_cpo)
            diff_profit_delta = _round2(profit_delta - item.metrics.profit_delta)
            verified = all(abs(value) <= 0.01 for value in (
                diff_cost_price,
                diff_gross_profit,
                diff_net_profit,
                diff_max_cpo,
                diff_actual_cpo,
                diff_profit_delta,
            ))

            row = {
                "nm_id": nm_id,
                "vendor_code": item.vendor_code or "",
                "title": item.title or "",
                "status": item.status,
                "status_label": item.status_label,
                "diagnosis": item.diagnosis,
                "advert_views": advert.total_views,
                "advert_clicks": advert.total_clicks,
                "advert_orders": advert.total_orders,
                "advert_gmv": _round2(advert.total_gmv),
                "advert_spend": _round2(advert.total_spend),
                "finance_revenue": _round2(finance.revenue),
                "finance_wb_costs": _round2(finance.wb_costs),
                "finance_orders": int(finance.orders or 0),
                "funnel_opens": int(funnel.open_count or 0),
                "funnel_carts": int(funnel.cart_count or 0),
                "funnel_orders": int(funnel.order_count or 0),
                "funnel_order_sum": _round2(funnel.order_sum),
                "unit_cost": _round2(unit_cost),
                "total_orders_formula": total_orders,
                "cost_price_formula": _round2(cost_price),
                "gross_profit_formula": _round2(gross_profit_before_ads),
                "net_profit_formula": _round2(net_profit),
                "max_cpo_formula": _round2(max_cpo),
                "actual_cpo_formula": _round2(actual_cpo),
                "profit_delta_formula": _round2(profit_delta),
                "app_cost_price": _round2(item.metrics.cost_price),
                "app_gross_profit": _round2(item.metrics.gross_profit_before_ads),
                "app_net_profit": _round2(item.metrics.net_profit),
                "app_max_cpo": _round2(item.metrics.max_cpo),
                "app_actual_cpo": _round2(item.metrics.actual_cpo),
                "app_profit_delta": _round2(item.metrics.profit_delta),
                "diff_cost_price": diff_cost_price,
                "diff_gross_profit": diff_gross_profit,
                "diff_net_profit": diff_net_profit,
                "diff_max_cpo": diff_max_cpo,
                "diff_actual_cpo": diff_actual_cpo,
                "diff_profit_delta": diff_profit_delta,
                "verified": "OK" if verified else "FAIL",
            }
            verification_rows.append(row)
            if not verified:
                failures.append(row)

    report_base = REPORT_DIR / f"sku_economics_audit_store{store_id}_{period_start.isoformat()}_{period_end.isoformat()}"
    workbook_path = report_base.with_suffix(".xlsx")
    report_path = report_base.with_suffix(".md")

    workbook = Workbook()
    print("Saving workbook...", flush=True)
    summary_ws = workbook.active
    summary_ws.title = "summary"
    summary_rows = [
        ["Параметр", "Значение"],
        ["Store ID", store_id],
        ["Store name", getattr(store, "name", "")],
        ["Период", f"{period_start.isoformat()} — {period_end.isoformat()}"],
        ["SKU в проверке", len(verification_rows)],
        ["Проверено OK", sum(1 for row in verification_rows if row["verified"] == "OK")],
        ["Проверено FAIL", len(failures)],
        ["Источник advert", f"{advert_status.mode} | {advert_status.detail or ''}"],
        ["Источник finance", f"{finance_status.mode} | {finance_status.detail or ''}"],
        ["Источник funnel", f"{funnel_status.mode} | {funnel_status.detail or ''}"],
        ["Формула 1", "total_orders = max(finance.orders, funnel.order_count, advert.ad_orders)"],
        ["Формула 2", "cost_price = unit_cost * total_orders"],
        ["Формула 3", "gross_profit = revenue - wb_costs - cost_price"],
        ["Формула 4", "net_profit = gross_profit - ad_cost"],
        ["Формула 5", "max_cpo = gross_profit / total_orders, если total_orders > 0"],
        ["Формула 6", "actual_cpo = ad_cost / ad_orders, если ad_orders > 0"],
        ["Формула 7", "profit_delta = max_cpo - actual_cpo"],
    ]
    for row in summary_rows:
        summary_ws.append(row)
    _style_header(summary_ws)
    _auto_width(summary_ws)

    check_ws = workbook.create_sheet("formula_check")
    check_headers = list(verification_rows[0].keys()) if verification_rows else ["nm_id", "verified"]
    check_ws.append(check_headers)
    for row in verification_rows:
        check_ws.append([row.get(header) for header in check_headers])
    _style_header(check_ws)
    for row_idx in range(2, check_ws.max_row + 1):
        if check_ws.cell(row=row_idx, column=check_headers.index("verified") + 1).value == "FAIL":
            for cell in check_ws[row_idx]:
                cell.fill = PatternFill("solid", fgColor="FDECEC")
    _auto_width(check_ws)

    advert_ws = workbook.create_sheet("advert_api")
    advert_headers = [
        "nm_id", "views", "clicks", "orders", "gmv", "exact_spend", "estimated_spend",
        "manual_spend", "total_spend",
    ]
    advert_ws.append(advert_headers)
    for nm_id in _iter_source_rows(advert_data["per_nm"], nm_ids):
        metric = _copy_advert(advert_data["per_nm"].get(nm_id))
        advert_ws.append([
            nm_id,
            metric.total_views,
            metric.total_clicks,
            metric.total_orders,
            _round2(metric.total_gmv),
            _round2(metric.exact_spend),
            _round2(metric.estimated_spend),
            _round2(metric.manual_spend),
            _round2(metric.total_spend),
        ])
    _style_header(advert_ws)
    _auto_width(advert_ws)

    finance_ws = workbook.create_sheet("finance_api")
    finance_headers = ["nm_id", "revenue", "payout", "wb_costs", "orders"]
    finance_ws.append(finance_headers)
    for nm_id in _iter_source_rows(finance_data, nm_ids):
        metric = finance_data[nm_id]
        finance_ws.append([
            nm_id,
            _round2(metric.revenue),
            _round2(metric.payout),
            _round2(metric.wb_costs),
            int(metric.orders or 0),
        ])
    _style_header(finance_ws)
    _auto_width(finance_ws)

    funnel_ws = workbook.create_sheet("funnel_api")
    funnel_headers = [
        "nm_id", "open_count", "cart_count", "order_count", "order_sum",
        "buyout_count", "buyout_sum", "add_to_cart_percent", "cart_to_order_percent",
    ]
    funnel_ws.append(funnel_headers)
    for nm_id in _iter_source_rows(funnel_data, nm_ids):
        metric = funnel_data[nm_id]
        funnel_ws.append([
            nm_id,
            int(metric.open_count or 0),
            int(metric.cart_count or 0),
            int(metric.order_count or 0),
            _round2(metric.order_sum),
            int(metric.buyout_count or 0),
            _round2(metric.buyout_sum),
            _round2(metric.add_to_cart_percent),
            _round2(metric.cart_to_order_percent),
        ])
    _style_header(funnel_ws)
    _auto_width(funnel_ws)

    costs_ws = workbook.create_sheet("manual_costs")
    costs_ws.append(["nm_id", "unit_cost"])
    for nm_id in sorted(set(nm_ids) | set(manual_costs.keys())):
        if nm_id in manual_costs:
            costs_ws.append([nm_id, _round2(manual_costs[nm_id])])
    _style_header(costs_ws)
    _auto_width(costs_ws)

    workbook.save(workbook_path)

    max_abs_diff = 0.0
    for row in verification_rows:
        for key in ("diff_cost_price", "diff_gross_profit", "diff_net_profit", "diff_max_cpo", "diff_actual_cpo", "diff_profit_delta"):
            max_abs_diff = max(max_abs_diff, abs(float(row[key] or 0.0)))

    report_lines = [
        f"# SKU Economics audit for store {store_id}",
        "",
        f"- Period: `{period_start.isoformat()} — {period_end.isoformat()}`",
        f"- Store: `{getattr(store, 'name', '')}`",
        f"- SKU checked: `{len(verification_rows)}`",
        f"- Verified OK: `{sum(1 for row in verification_rows if row['verified'] == 'OK')}`",
        f"- Verified FAIL: `{len(failures)}`",
        f"- Max absolute diff across key metrics: `{_round2(max_abs_diff)}`",
        "",
        "## Sources",
        f"- WB Advert: `{advert_status.mode}` — {advert_status.detail or '—'}",
        f"- WB Finance: `{finance_status.mode}` — {finance_status.detail or '—'}",
        f"- WB Funnel: `{funnel_status.mode}` — {funnel_status.detail or '—'}",
        "",
        "## Manual formula mirror",
        "- `total_orders = max(finance.orders, funnel.order_count, advert.ad_orders)`",
        "- `cost_price = unit_cost * total_orders`",
        "- `gross_profit = revenue - wb_costs - cost_price`",
        "- `net_profit = gross_profit - ad_cost`",
        "- `max_cpo = gross_profit / total_orders` if orders > 0",
        "- `actual_cpo = ad_cost / ad_orders` if ad_orders > 0",
        "- `profit_delta = max_cpo - actual_cpo`",
        "",
        f"Workbook: `{workbook_path}`",
    ]
    if failures:
        report_lines.extend([
            "",
            "## Failed rows",
        ])
        for row in failures[:20]:
            report_lines.append(
                f"- nmID `{row['nm_id']}`: net `{row['diff_net_profit']}`, max_cpo `{row['diff_max_cpo']}`, actual_cpo `{row['diff_actual_cpo']}`, delta `{row['diff_profit_delta']}`"
            )

    report_path.write_text("\n".join(report_lines), encoding="utf-8")
    print("Done.", flush=True)
    return workbook_path, report_path


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Audit SKU economics against live WB APIs")
    parser.add_argument("--store-id", type=int, default=4)
    parser.add_argument("--period-start", type=str, default="")
    parser.add_argument("--period-end", type=str, default="")
    parser.add_argument("--force-overview", action="store_true")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    period_end = date.fromisoformat(args.period_end) if args.period_end else date.today()
    period_start = date.fromisoformat(args.period_start) if args.period_start else period_end - timedelta(days=6)
    workbook_path, report_path = asyncio.run(run_audit(args.store_id, period_start, period_end, args.force_overview))
    print(workbook_path)
    print(report_path)


if __name__ == "__main__":
    main()
