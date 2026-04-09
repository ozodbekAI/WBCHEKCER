"""
Fixed File Service
Handles Excel upload, parsing, CRUD and card comparison for FixedFileEntry records.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import List, Optional, Dict

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import select, delete, update, func
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.ext.asyncio import AsyncSession

from ..core.time import utc_now
from ..models.fixed_file import FixedFileEntry

# Columns in the Excel that are NOT characteristics (skip when parsing)
_META_COLS = {"nmid", "brand", "subjectname", "артикул", "бренд", "категория"}

# Columns that should appear first in the generated template (meta columns)
_TEMPLATE_META = ["Артикул", "Бренд", "Категория"]

# Commonly fixed characteristics (used as template column headers)
_TEMPLATE_CHARS = [
    "Пол",
    "Состав",
    "Ставка НДС",
    "Размер на модели",
    "Рост модели на фото",
    "Уход за вещами",
    "Фактура материала",
    "Тип ростовки",
    "Материал подкладки",
    "ТНВЭД",
    "Страна производства",
    "Дата регистрации сертификата/декларации",
    "Дата окончания действия сертификата/декларации",
    "Номер декларации соответствия",
    "Параметры модели на фото (ОГ-ОТ-ОБ)",
    "Вес товара с упаковкой (г)",
    "Высота упаковки",
    "Длина упаковки",
    "Ширина упаковки",
]


# ─── Parsing ─────────────────────────────────────────────────────────────────

def parse_excel(file_bytes: bytes) -> List[Dict]:
    """Parse an Excel fixed-file and return list of entry dicts.

    Each dict: {nm_id, brand, subject_name, char_name, fixed_value}
    All sheets in the workbook are parsed.
    Rows/cells where value is None or empty string are skipped.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    entries = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        # Skip sheets that don't have a recognisable ID column
        has_id_col = any(h.lower() in ("nmid", "артикул") for h in headers)
        if not has_id_col:
            continue

        for row in rows[1:]:
            nm_id = None
            brand = None
            subject_name = None
            row_data: Dict[str, str] = {}

            for col_idx, header in enumerate(headers):
                if col_idx >= len(row):
                    continue
                val = row[col_idx]
                if val is None or str(val).strip() == "":
                    continue
                h_lower = header.lower().strip()
                if h_lower in ("nmid", "артикул"):
                    try:
                        nm_id = int(val)
                    except (ValueError, TypeError):
                        pass
                elif h_lower in ("brand", "бренд"):
                    brand = str(val).strip()
                elif h_lower in ("subjectname", "категория"):
                    subject_name = str(val).strip()
                elif header and h_lower not in _META_COLS:
                    row_data[header] = str(val).strip()

            if nm_id is None:
                continue

            for char_name, fixed_value in row_data.items():
                if fixed_value:
                    entries.append({
                        "nm_id": nm_id,
                        "brand": brand,
                        "subject_name": subject_name,
                        "char_name": char_name,
                        "fixed_value": fixed_value,
                    })

    return entries


# ─── Database CRUD ────────────────────────────────────────────────────────────

async def upsert_entries(
    db: AsyncSession,
    store_id: int,
    entries: List[Dict],
    user_id: int,
) -> int:
    """Insert or update fixed file entries using PostgreSQL ON CONFLICT DO UPDATE.
    Processes in chunks to stay under PostgreSQL's 65535 parameter limit.
    Returns count of upserted rows.
    """
    if not entries:
        return 0

    now = utc_now()
    rows = [
        {
            "store_id": store_id,
            "nm_id": e["nm_id"],
            "brand": e.get("brand"),
            "subject_name": e.get("subject_name"),
            "char_name": e["char_name"],
            "fixed_value": e["fixed_value"],
            "created_at": now,
            "updated_at": now,
            "created_by_id": user_id,
        }
        for e in entries
    ]

    # 9 columns per row; keep well under the 65535 param limit
    CHUNK = 2000
    for i in range(0, len(rows), CHUNK):
        chunk = rows[i : i + CHUNK]
        stmt = pg_insert(FixedFileEntry).values(chunk)
        stmt = stmt.on_conflict_do_update(
            constraint="uq_fixed_entry",
            set_={
                "fixed_value": stmt.excluded.fixed_value,
                "brand": stmt.excluded.brand,
                "subject_name": stmt.excluded.subject_name,
                "updated_at": stmt.excluded.updated_at,
            },
        )
        await db.execute(stmt)

    await db.commit()
    return len(rows)


async def get_entries(
    db: AsyncSession,
    store_id: int,
    nm_id: Optional[int] = None,
    skip: int = 0,
    limit: int = 100,
) -> tuple[List[FixedFileEntry], int]:
    """Get fixed file entries for a store with optional nm_id filter."""
    base = select(FixedFileEntry).where(FixedFileEntry.store_id == store_id)
    count_q = select(func.count(FixedFileEntry.id)).where(FixedFileEntry.store_id == store_id)

    if nm_id is not None:
        base = base.where(FixedFileEntry.nm_id == nm_id)
        count_q = count_q.where(FixedFileEntry.nm_id == nm_id)

    total_r = await db.execute(count_q)
    total = total_r.scalar() or 0

    result = await db.execute(
        base.order_by(FixedFileEntry.nm_id, FixedFileEntry.char_name)
        .offset(skip).limit(limit)
    )
    return list(result.scalars().all()), total


async def get_entries_for_card(
    db: AsyncSession,
    store_id: int,
    nm_id: int,
) -> List[FixedFileEntry]:
    """Get all fixed entries for a specific card (used during analysis)."""
    result = await db.execute(
        select(FixedFileEntry).where(
            FixedFileEntry.store_id == store_id,
            FixedFileEntry.nm_id == nm_id,
        )
    )
    return list(result.scalars().all())


async def has_any_entries(db: AsyncSession, store_id: int) -> bool:
    """Return True if the store has at least one fixed file entry."""
    result = await db.execute(
        select(func.count(FixedFileEntry.id)).where(FixedFileEntry.store_id == store_id)
    )
    return (result.scalar() or 0) > 0


async def update_entry(
    db: AsyncSession,
    entry_id: int,
    store_id: int,
    fixed_value: str,
) -> Optional[FixedFileEntry]:
    """Update a single entry's fixed_value."""
    result = await db.execute(
        select(FixedFileEntry).where(
            FixedFileEntry.id == entry_id,
            FixedFileEntry.store_id == store_id,
        )
    )
    entry = result.scalar_one_or_none()
    if not entry:
        return None
    entry.fixed_value = fixed_value
    entry.updated_at = utc_now()
    await db.commit()
    await db.refresh(entry)
    return entry


async def delete_entry(
    db: AsyncSession,
    entry_id: int,
    store_id: int,
) -> bool:
    """Delete a single entry. Returns True if deleted."""
    result = await db.execute(
        select(FixedFileEntry).where(
            FixedFileEntry.id == entry_id,
            FixedFileEntry.store_id == store_id,
        )
    )
    entry = result.scalar_one_or_none()
    if not entry:
        return False
    await db.delete(entry)
    await db.commit()
    return True


async def delete_all_entries(db: AsyncSession, store_id: int) -> int:
    """Delete ALL entries for a store (e.g. before re-upload). Returns count."""
    result = await db.execute(
        select(func.count(FixedFileEntry.id)).where(FixedFileEntry.store_id == store_id)
    )
    count = result.scalar() or 0
    await db.execute(
        delete(FixedFileEntry).where(FixedFileEntry.store_id == store_id)
    )
    await db.commit()
    return count


# ─── Template generation ──────────────────────────────────────────────────────

def generate_template_excel() -> bytes:
    """Generate an Excel template with correct column headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fixed Values"

    headers = _TEMPLATE_META + _TEMPLATE_CHARS

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    meta_fill = PatternFill("solid", fgColor="D9E1F2")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        if header in _TEMPLATE_META:
            cell.fill = meta_fill
            cell.font = Font(bold=True, color="1F3864")
        else:
            cell.fill = header_fill

        # Column width
        ws.column_dimensions[cell.column_letter].width = max(18, len(header) * 1.2)

    # Example row
    example = [123456789, "Ваш бренд", "Блузки"] + [""] * len(_TEMPLATE_CHARS)
    for col_idx, val in enumerate(example, start=1):
        ws.cell(row=2, column=col_idx, value=val)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Card comparison ──────────────────────────────────────────────────────────

def compare_card_with_fixed(
    raw_data: dict,
    fixed_entries: List[FixedFileEntry],
) -> List[dict]:
    """
    Compare card characteristics against fixed file entries.
    Returns list of mismatch dicts (one per mismatched characteristic).

    Each dict: {char_name, card_value, fixed_value, field_path}
    """
    if not fixed_entries:
        return []

    # Build card char map: name.lower() → string value
    chars_raw = raw_data.get("characteristics") or []
    card_char_map: Dict[str, str] = {}
    for ch in (chars_raw if isinstance(chars_raw, list) else []):
        name = (ch.get("name") or "").strip()
        val = ch.get("value") or ch.get("values")
        if name and val is not None:
            str_val = ", ".join(str(v) for v in val) if isinstance(val, list) else str(val)
            card_char_map[name.lower()] = str_val

    mismatches = []
    for entry in fixed_entries:
        card_val = card_char_map.get(entry.char_name.lower())
        fixed_val = entry.fixed_value.strip()

        # Normalize for comparison (semicolons ↔ commas, case-insensitive, whitespace)
        def _norm(s: str) -> str:
            return s.replace(";", ",").replace("  ", " ").strip().lower()

        if card_val is None or _norm(card_val) != _norm(fixed_val):
            mismatches.append({
                "char_name": entry.char_name,
                "card_value": card_val,
                "fixed_value": fixed_val,
                "field_path": f"characteristics.{entry.char_name}",
            })

    return mismatches


# ─── Vision fallback ──────────────────────────────────────────────────────────

async def generate_characteristics_from_photo(
    db: AsyncSession,
    store_id: int,
    nm_id: int,
    card_raw_data: dict,
    user_id: Optional[int] = None,
) -> dict:
    """
    Mahsulot fotosini GPT-4o-mini ga yuborib Product DNA generatsiya qiladi,
    keyin xarakteristikalarni FixedFileEntry ga saqlaydi.

    Returns: {
        "generated": int,  # saqlangan xarakteristikalar soni
        "characteristics": {char_name: value},
        "product_dna": dict,
    }
    """
    from app.services.vision_service import vision_service

    if not vision_service.is_enabled:
        return {"generated": 0, "characteristics": {}, "product_dna": {}, "error": "OPENAI_API_KEY sozlanmagan"}

    # Kartochkaning birinchi fotosini olish
    photos = card_raw_data.get("photos", [])
    photo_url: str | None = None
    if isinstance(photos, list):
        for p in photos:
            if isinstance(p, dict):
                url = p.get("big") or p.get("c516x688") or p.get("tm")
                if url:
                    photo_url = url
                    break
            elif isinstance(p, str):
                photo_url = p
                break

    if not photo_url:
        return {"generated": 0, "characteristics": {}, "product_dna": {}, "error": "Foto topilmadi"}

    subject_name = card_raw_data.get("subjectName") or ""

    # Vision API ga yuborish
    dna = await vision_service.analyze_photo_dna(photo_url, subject_name)
    if not dna:
        return {"generated": 0, "characteristics": {}, "product_dna": {}, "error": "Vision API javob bermadi"}

    # Product DNA dan WB xarakteristikalarini chiqarish
    chars = vision_service.extract_wb_characteristics(dna)
    if not chars:
        return {"generated": 0, "characteristics": {}, "product_dna": dna}

    # brand va subject_name ni raw_data dan olish
    brand = card_raw_data.get("brand") or ""

    # FixedFileEntry larni yaratish
    entries = [
        {
            "nm_id": nm_id,
            "brand": brand,
            "subject_name": subject_name,
            "char_name": char_name,
            "fixed_value": str(value),
        }
        for char_name, value in chars.items()
        if value and str(value).strip()
    ]

    if entries:
        await upsert_entries(db, store_id, entries, user_id=user_id or 0)

    return {
        "generated": len(entries),
        "characteristics": chars,
        "product_dna": dna,
    }
