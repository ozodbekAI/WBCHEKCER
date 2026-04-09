from __future__ import annotations

import asyncio
import csv
import io
import logging
import time
import zipfile
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from uuid import uuid4

import httpx
from openpyxl import load_workbook
from sqlalchemy import delete, func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.time import utc_now
from app.models import (
    Card,
    CardIssue,
    IssueCategory,
    IssueSeverity,
    IssueStatus,
    SkuEconomicsCost,
    SkuEconomicsManualFinance,
    SkuEconomicsManualSpend,
    SkuEconomicsDailyMetric,
    SkuEconomicsOverviewCache,
    SkuEconomicsSnapshot,
    Store,
)
from app.schemas.sku_economics import (
    AdAnalysisAlertOut,
    AdAnalysisBudgetMoveOut,
    AdAnalysisCampaignOut,
    AdAnalysisIssueSummaryOut,
    AdAnalysisItemOut,
    AdAnalysisMetricsOut,
    AdAnalysisOverviewOut,
    AdAnalysisSourceLineageOut,
    AdAnalysisSourceStatusOut,
    AdAnalysisTrendOut,
    AdAnalysisUploadNeedsOut,
    AdAnalysisUploadUnresolvedRowOut,
)
from app.services.wb_advert_repository import WBAdvertRepository
from app.services.wb_token_access import get_store_feature_api_key

logger = logging.getLogger("wbai.sku_economics")


_UNRESOLVED_ISSUE_STATUSES = {
    IssueStatus.PENDING.value,
    IssueStatus.SKIPPED.value,
    IssueStatus.POSTPONED.value,
}


def _to_float(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    raw = str(value).strip()
    if not raw:
        return 0.0
    normalized = (
        raw.replace("\u00a0", "")
        .replace(" ", "")
        .replace("%", "")
        .replace(",", ".")
    )
    cleaned = "".join(ch for ch in normalized if ch.isdigit() or ch in ".-")
    if cleaned in {"", ".", "-", "-."}:
        return 0.0
    try:
        return float(cleaned)
    except Exception:
        return 0.0


def _to_int(value: Any) -> int:
    return int(round(_to_float(value)))


def _safe_positive(value: float) -> float:
    return float(value) if value > 0 else 0.0


def _normalize_header(value: str) -> str:
    raw = str(value or "").strip().lower().replace("ё", "е")
    return "".join(ch if ch.isalnum() else " " for ch in raw).strip()


def _format_money(value: float) -> str:
    return f"{int(round(value)):,}".replace(",", " ") + " ₽"


def _normalize_vendor_code(value: Any) -> str:
    raw = str(value or "").strip().lower()
    if not raw:
        return ""
    return "".join(ch for ch in raw if ch.isalnum())


def _today_period(days: int) -> Tuple[date, date]:
    safe_days = max(int(days or 14), 1)
    period_end = date.today()
    period_start = period_end - timedelta(days=safe_days - 1)
    return period_start, period_end


def _chunks(values: Sequence[int], size: int) -> Iterable[List[int]]:
    for idx in range(0, len(values), size):
        yield list(values[idx : idx + size])


def _retry_delay_from_response(resp: httpx.Response, attempt: int) -> float:
    retry_after = str(resp.headers.get("Retry-After") or "").strip()
    if retry_after:
        try:
            return min(max(float(retry_after), 1.0), 20.0)
        except Exception:
            pass
    return min(2.0 * (attempt + 1), 12.0)


def _match_columns(headers: Sequence[str], aliases: Dict[str, List[str]]) -> Dict[str, str]:
    normalized = {header: _normalize_header(header) for header in headers}
    out: Dict[str, str] = {}
    used: set[str] = set()

    for field, field_aliases in aliases.items():
        for alias in field_aliases:
            alias_norm = _normalize_header(alias)
            for header, header_norm in normalized.items():
                if header in used:
                    continue
                if header_norm == alias_norm:
                    out[field] = header
                    used.add(header)
                    break
            if field in out:
                break

    for field, field_aliases in aliases.items():
        if field in out:
            continue
        for alias in field_aliases:
            alias_norm = _normalize_header(alias)
            for header, header_norm in normalized.items():
                if header in used:
                    continue
                if alias_norm and (alias_norm in header_norm or header_norm in alias_norm):
                    out[field] = header
                    used.add(header)
                    break
            if field in out:
                break
    return out


def _decode_csv(content: bytes) -> io.StringIO:
    for encoding in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            return io.StringIO(content.decode(encoding))
        except UnicodeDecodeError:
            continue
    return io.StringIO(content.decode("utf-8", errors="ignore"))


def _read_tabular_rows(file_name: str, content: bytes) -> Tuple[List[str], List[Dict[str, Any]]]:
    suffix = Path(file_name or "").suffix.lower()
    if suffix == ".csv":
        reader = csv.DictReader(_decode_csv(content))
        rows = [dict(row) for row in reader]
        return list(reader.fieldnames or []), rows

    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError("Поддерживаются только CSV и XLSX файлы")

    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    iterator = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(iterator)
    except StopIteration:
        return [], []

    headers = [str(cell).strip() if cell is not None else "" for cell in raw_headers]
    rows: List[Dict[str, Any]] = []
    for row_values in iterator:
        row: Dict[str, Any] = {}
        has_value = False
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = row_values[idx] if idx < len(row_values) else None
            if value not in (None, ""):
                has_value = True
            row[header] = value
        if has_value:
            rows.append(row)
    return headers, rows


@dataclass
class _IssueSnapshot:
    total: int = 0
    critical: int = 0
    warnings: int = 0
    photos: int = 0
    price: int = 0
    text: int = 0
    docs: int = 0
    top_titles: List[str] = field(default_factory=list)


@dataclass
class _AdvertMetrics:
    views: int = 0
    clicks: int = 0
    orders: int = 0
    gmv: float = 0.0
    exact_spend: float = 0.0
    estimated_spend: float = 0.0
    manual_spend: float = 0.0
    manual_views: int = 0
    manual_clicks: int = 0
    manual_orders: int = 0
    manual_gmv: float = 0.0

    @property
    def total_spend(self) -> float:
        return self.exact_spend + self.estimated_spend + self.manual_spend

    @property
    def total_views(self) -> int:
        return int(self.views + self.manual_views)

    @property
    def total_clicks(self) -> int:
        return int(self.clicks + self.manual_clicks)

    @property
    def total_orders(self) -> int:
        return int(self.orders + self.manual_orders)

    @property
    def total_gmv(self) -> float:
        return self.gmv + self.manual_gmv


@dataclass
class _FunnelMetrics:
    open_count: int = 0
    cart_count: int = 0
    order_count: int = 0
    order_sum: float = 0.0
    buyout_count: int = 0
    buyout_sum: float = 0.0
    add_to_cart_percent: float = 0.0
    cart_to_order_percent: float = 0.0


@dataclass
class _FinanceMetrics:
    revenue: float = 0.0
    payout: float = 0.0
    wb_costs: float = 0.0
    orders: int = 0


@dataclass
class _DailyMetricBucket:
    title: str = ""
    vendor_code: str = ""
    advert_views: int = 0
    advert_clicks: int = 0
    advert_orders: int = 0
    advert_gmv: float = 0.0
    advert_exact_spend: float = 0.0
    advert_estimated_spend: float = 0.0
    finance_revenue: float = 0.0
    finance_payout: float = 0.0
    finance_wb_costs: float = 0.0
    finance_orders: int = 0
    funnel_open_count: int = 0
    funnel_cart_count: int = 0
    funnel_order_count: int = 0
    funnel_order_sum: float = 0.0
    funnel_buyout_count: int = 0
    funnel_buyout_sum: float = 0.0
    has_advert: bool = False
    has_finance: bool = False
    has_funnel: bool = False


@dataclass
class _UploadResolution:
    nm_id: int = 0
    vendor_code: str = ""


@dataclass
class _UploadResult:
    imported: int = 0
    updated: int = 0
    notes: List[str] = field(default_factory=list)
    detected_headers: List[str] = field(default_factory=list)
    matched_fields: Dict[str, str] = field(default_factory=dict)
    resolved_by_vendor_code: int = 0
    unresolved_count: int = 0
    unresolved_preview: List[AdAnalysisUploadUnresolvedRowOut] = field(default_factory=list)


class _FinanceFetchError(Exception):
    def __init__(
        self,
        message: str,
        *,
        retry_count: int = 0,
        status_code: Optional[int] = None,
        partial_rows: Optional[List[Dict[str, Any]]] = None,
    ) -> None:
        super().__init__(message)
        self.message = str(message or "")
        self.retry_count = int(retry_count or 0)
        self.status_code = int(status_code) if status_code is not None else None
        self.partial_rows: List[Dict[str, Any]] = list(partial_rows or [])

    def __str__(self) -> str:
        return self.message


class SkuEconomicsService:
    CACHE_TTL_SEC = 300
    HISTORY_LOOKBACK_DAYS = 365
    MAX_PAGE_SIZE = 100
    FINANCE_PAGE_LIMIT = 100000
    FINANCE_MAX_PAGES_PER_RANGE = 500
    FINANCE_MIN_REQUEST_INTERVAL_SEC = 0.25
    FINANCE_RETRY_MAX_ATTEMPTS = 5
    FINANCE_RETRY_BASE_DELAY_SEC = 1.0
    FINANCE_RETRY_MAX_DELAY_SEC = 20.0
    FINANCE_RATE_LIMIT_STATUS = 429
    FINANCE_SERVER_ERROR_MIN_STATUS = 500
    FINANCE_SERVER_ERROR_MAX_STATUS = 599
    ADVERT_RESIDUAL_EPSILON = 0.01
    ADVERT_CAMPAIGN_SPEND_KEYS = ("sum", "spend", "spent", "total", "cost")
    ADVERT_NM_ID_KEYS = ("nmId", "nm_id", "nmid", "id")
    ADVERT_NM_TITLE_KEYS = ("name", "title", "nmName", "nm_name")
    ADVERT_NM_SPEND_KEYS = ("sum", "spend", "spent", "cost")
    ADVERT_NM_VIEWS_KEYS = ("views", "shows", "impressions")
    ADVERT_NM_CLICKS_KEYS = ("clicks", "click", "clickCount", "click_count")
    ADVERT_NM_ORDERS_KEYS = ("orders", "orderCount", "orders_count", "order_count")
    ADVERT_NM_GMV_KEYS = ("sum_price", "sumPrice", "gmv", "sumPriceWithDisc")

    COST_ALIASES = {
        "nm_id": ["nm_id", "nmid", "nm id", "артикул wb", "артикул", "wb nm", "wb article"],
        "unit_cost": ["себестоимость", "себ", "cost", "unit_cost", "cost price", "закупка", "закупочная цена"],
        "title": ["title", "название", "карточка", "товар"],
        "vendor_code": ["vendor_code", "vendor code", "артикул продавца", "артикул поставщика", "seller article", "supplier article"],
    }
    SPEND_ALIASES = {
        "nm_id": ["nm_id", "nmid", "nm id", "артикул wb", "nm"],
        "vendor_code": ["vendor_code", "vendor code", "артикул продавца", "артикул поставщика", "seller article", "supplier article", "артикул"],
        "spend": ["spend", "расход", "затраты", "sum", "ad spend", "реклама"],
        "views": ["views", "показы"],
        "clicks": ["clicks", "клики"],
        "orders": ["orders", "заказы"],
        "gmv": ["gmv", "выручка", "оборот", "sum_price"],
        "title": ["title", "название", "товар"],
    }
    FINANCE_ALIASES = {
        "nm_id": ["nm_id", "nmid", "nm id", "артикул wb"],
        "vendor_code": ["vendor_code", "vendor code", "артикул продавца", "артикул поставщика", "seller article", "supplier article", "артикул"],
        "revenue": ["revenue", "выручка", "retail_price_withdisc_rub", "retail_amount"],
        "wb_costs": ["wb_costs", "wb costs", "расходы wb", "комиссия wb"],
        "payout": ["payout", "ppvz_for_pay", "к перечислению"],
        "orders": ["orders", "заказы", "quantity"],
        "title": ["title", "название", "товар"],
    }

    def __init__(self) -> None:
        self._cache: Dict[Tuple[int, str, str], Tuple[float, AdAnalysisOverviewOut]] = {}

    def invalidate_cache(self, store_id: int) -> None:
        for key in list(self._cache.keys()):
            if key[0] == int(store_id):
                self._cache.pop(key, None)

    async def invalidate_saved_overviews(self, db: AsyncSession, store_id: int) -> None:
        await db.execute(
            delete(SkuEconomicsOverviewCache).where(SkuEconomicsOverviewCache.store_id == int(store_id))
        )
        self.invalidate_cache(store_id)

    async def _request_with_retry(
        self,
        client: httpx.AsyncClient,
        method: str,
        url: str,
        *,
        headers: Dict[str, str],
        params: Optional[Dict[str, Any]] = None,
        json: Optional[Dict[str, Any]] = None,
        retries: int = 4,
    ) -> httpx.Response:
        last_response: Optional[httpx.Response] = None
        for attempt in range(max(int(retries), 1)):
            response = await client.request(
                method,
                url,
                headers=headers,
                params=params,
                json=json,
            )
            last_response = response
            if response.status_code not in {429, 502, 503, 504} or attempt >= retries - 1:
                return response
            await asyncio.sleep(_retry_delay_from_response(response, attempt))
        if last_response is None:
            raise RuntimeError("WB request failed before response")
        return last_response

    async def _throttle_finance_request(self, last_request_ts: float) -> None:
        if last_request_ts <= 0:
            return
        wait_for = self.FINANCE_MIN_REQUEST_INTERVAL_SEC - (time.monotonic() - float(last_request_ts))
        if wait_for > 0:
            await asyncio.sleep(wait_for)

    def _finance_backoff_delay(self, response: Optional[httpx.Response], attempt: int) -> float:
        safe_attempt = max(int(attempt), 0)
        if response is not None and int(response.status_code) == self.FINANCE_RATE_LIMIT_STATUS:
            return _retry_delay_from_response(response, safe_attempt)
        return min(
            self.FINANCE_RETRY_BASE_DELAY_SEC * (2 ** safe_attempt),
            self.FINANCE_RETRY_MAX_DELAY_SEC,
        )

    async def _request_finance_page_with_retry(
        self,
        *,
        client: httpx.AsyncClient,
        url: str,
        headers: Dict[str, str],
        params: Dict[str, Any],
        date_from: date,
        date_to: date,
        rrdid: int,
        last_request_ts: float,
    ) -> Tuple[httpx.Response, float]:
        max_attempts = max(int(self.FINANCE_RETRY_MAX_ATTEMPTS), 1)

        for attempt in range(max_attempts):
            await self._throttle_finance_request(last_request_ts)
            try:
                response = await client.request(
                    "GET",
                    url,
                    headers=headers,
                    params=params,
                )
                last_request_ts = time.monotonic()
            except httpx.RequestError as exc:
                retry_count = attempt + 1
                if retry_count >= max_attempts:
                    logger.error(
                        "[sku-finance] request failed after retries type=request_error retries=%s "
                        "period=%s..%s rrdid=%s error=%s",
                        retry_count,
                        date_from.isoformat(),
                        date_to.isoformat(),
                        int(rrdid),
                        str(exc),
                    )
                    raise _FinanceFetchError(
                        f"WB finance request error after {retry_count} retries: {exc}",
                        retry_count=retry_count,
                    ) from exc

                delay = self._finance_backoff_delay(None, attempt)
                logger.warning(
                    "[sku-finance] retrying request type=request_error retry=%s/%s backoff=%.2fs "
                    "period=%s..%s rrdid=%s error=%s",
                    retry_count,
                    max_attempts,
                    delay,
                    date_from.isoformat(),
                    date_to.isoformat(),
                    int(rrdid),
                    str(exc),
                )
                await asyncio.sleep(delay)
                continue

            if response.status_code == 204 or response.status_code < 400:
                if attempt > 0:
                    logger.info(
                        "[sku-finance] request recovered after retry retries=%s status=%s "
                        "period=%s..%s rrdid=%s",
                        attempt,
                        int(response.status_code),
                        date_from.isoformat(),
                        date_to.isoformat(),
                        int(rrdid),
                    )
                return response, last_request_ts

            retry_count = attempt + 1
            retryable = (
                int(response.status_code) == self.FINANCE_RATE_LIMIT_STATUS
                or self.FINANCE_SERVER_ERROR_MIN_STATUS
                <= int(response.status_code)
                <= self.FINANCE_SERVER_ERROR_MAX_STATUS
            )
            if retryable and retry_count < max_attempts:
                delay = self._finance_backoff_delay(response, attempt)
                logger.warning(
                    "[sku-finance] retrying request status=%s retry=%s/%s backoff=%.2fs "
                    "period=%s..%s rrdid=%s",
                    int(response.status_code),
                    retry_count,
                    max_attempts,
                    delay,
                    date_from.isoformat(),
                    date_to.isoformat(),
                    int(rrdid),
                )
                await asyncio.sleep(delay)
                continue

            reason = f"WB finance error {response.status_code}: {response.text[:300]}"
            logger.error(
                "[sku-finance] request failed status=%s retries=%s period=%s..%s rrdid=%s reason=%s",
                int(response.status_code),
                retry_count,
                date_from.isoformat(),
                date_to.isoformat(),
                int(rrdid),
                response.text[:200],
            )
            raise _FinanceFetchError(
                reason,
                retry_count=retry_count,
                status_code=int(response.status_code),
            )

        raise _FinanceFetchError(
            "WB finance request failed: retry loop ended unexpectedly",
            retry_count=max_attempts,
        )

    async def _fetch_finance_rows_paginated(
        self,
        *,
        client: httpx.AsyncClient,
        url: str,
        token: str,
        date_from: date,
        date_to: date,
        period: str = "daily",
    ) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        rrdid = 0
        pages = 0
        seen_rrd_ids: set[int] = set()
        last_request_ts = 0.0

        while True:
            params = {
                "dateFrom": date_from.isoformat(),
                "dateTo": date_to.isoformat(),
                "limit": self.FINANCE_PAGE_LIMIT,
                "rrdid": int(rrdid),
                "period": period,
            }
            try:
                resp, last_request_ts = await self._request_finance_page_with_retry(
                    client=client,
                    url=url,
                    headers={"Authorization": token, "Accept": "application/json"},
                    params=params,
                    date_from=date_from,
                    date_to=date_to,
                    rrdid=int(rrdid),
                    last_request_ts=last_request_ts,
                )
            except _FinanceFetchError as exc:
                if rows:
                    raise _FinanceFetchError(
                        f"{exc} (partial rows collected: {len(rows)})",
                        retry_count=exc.retry_count,
                        status_code=exc.status_code,
                        partial_rows=rows,
                    ) from exc
                raise

            if resp.status_code == 204:
                break
            if resp.status_code >= 400:
                raise ValueError(f"WB finance error {resp.status_code}: {resp.text[:300]}")

            payload = resp.json() or []
            if isinstance(payload, dict):
                payload = payload.get("data") or payload.get("rows") or []
            if not isinstance(payload, list) or not payload:
                break

            page_rows = [item for item in payload if isinstance(item, dict)]
            if not page_rows:
                break

            rows.extend(page_rows)
            pages += 1
            if pages >= self.FINANCE_MAX_PAGES_PER_RANGE:
                raise ValueError(
                    f"WB finance pagination safeguard reached ({self.FINANCE_MAX_PAGES_PER_RANGE} pages) "
                    f"for {date_from.isoformat()}..{date_to.isoformat()}"
                )

            last_row = page_rows[-1]
            last_rrd_id = _to_int((last_row or {}).get("rrd_id") or (last_row or {}).get("rrdId"))
            if last_rrd_id <= 0:
                break
            if last_rrd_id <= int(rrdid) or last_rrd_id in seen_rrd_ids:
                break

            seen_rrd_ids.add(last_rrd_id)
            rrdid = last_rrd_id

        return rows

    async def build_overview(
        self,
        db: AsyncSession,
        store: Store,
        *,
        days: int = 14,
        period_start: Optional[date] = None,
        period_end: Optional[date] = None,
        preset: Optional[str] = None,
        page: int = 1,
        page_size: int = 50,
        status_filter: Optional[str] = None,
        search: Optional[str] = None,
        force: bool = False,
    ) -> AdAnalysisOverviewOut:
        available_start, available_end = await self._get_available_period(db, store.id)
        requested_start, requested_end, selected_preset = self._resolve_requested_period(
            days=days,
            period_start=period_start,
            period_end=period_end,
            preset=preset,
            available_start=available_start,
            available_end=available_end,
        )

        if force:
            await self._refresh_history_data(
                db,
                store,
                requested_period_start=requested_start,
                requested_period_end=requested_end,
            )
            available_start, available_end = await self._get_available_period(db, store.id)
            requested_start, requested_end, selected_preset = self._resolve_requested_period(
                days=days,
                period_start=period_start,
                period_end=period_end,
                preset=preset,
                available_start=available_start,
                available_end=available_end,
            )

        if available_start is None or available_end is None:
            previous_period_end = requested_start - timedelta(days=1)
            previous_period_start = previous_period_end - timedelta(days=(requested_end - requested_start).days)
            overview = await self._build_empty_overview(
                db,
                store_id=store.id,
                period_start=requested_start,
                period_end=requested_end,
                previous_period_start=previous_period_start,
                previous_period_end=previous_period_end,
            )
            overview.available_period_start = available_start
            overview.available_period_end = available_end
            overview.selected_preset = selected_preset
            overview.page = 1
            overview.page_size = page_size
            overview.total_items = 0
            overview.total_pages = 0
            return overview

        effective_start = max(requested_start, available_start)
        effective_end = min(requested_end, available_end)
        if effective_end < effective_start:
            effective_start, effective_end = available_start, available_end
            selected_preset = "all"

        should_use_saved_overview = not force and not status_filter and not search and int(page or 1) <= 1
        if should_use_saved_overview:
            saved_overview = await self._load_persisted_overview(
                db,
                store.id,
                effective_start,
                effective_end,
            )
            if saved_overview is not None:
                saved_overview.available_period_start = available_start
                saved_overview.available_period_end = available_end
                saved_overview.selected_preset = selected_preset
                self._cache[(int(store.id), effective_start.isoformat(), effective_end.isoformat())] = (time.time(), saved_overview)
                return saved_overview

        previous_period_end = effective_start - timedelta(days=1)
        previous_period_start = previous_period_end - timedelta(days=(effective_end - effective_start).days)

        current_data = await self._load_history_aggregate(
            db,
            store_id=store.id,
            period_start=effective_start,
            period_end=effective_end,
        )
        previous_data = await self._load_history_aggregate(
            db,
            store_id=store.id,
            period_start=previous_period_start,
            period_end=previous_period_end,
        )

        overview = await self._build_overview_from_history(
            db,
            store,
            period_start=effective_start,
            period_end=effective_end,
            previous_period_start=previous_period_start,
            previous_period_end=previous_period_end,
            current_data=current_data,
            previous_data=previous_data,
            selected_preset=selected_preset,
            page=page,
            page_size=page_size,
            status_filter=status_filter,
            search=search,
            available_period_start=available_start,
            available_period_end=available_end,
        )
        if overview.snapshot_ready and not status_filter and not search and page <= 1:
            await self._persist_overview(db, store.id, overview)
        return overview

    async def _get_available_period(
        self,
        db: AsyncSession,
        store_id: int,
    ) -> Tuple[Optional[date], Optional[date]]:
        result = await db.execute(
            select(
                func.min(SkuEconomicsDailyMetric.metric_date),
                func.max(SkuEconomicsDailyMetric.metric_date),
            ).where(
                SkuEconomicsDailyMetric.store_id == int(store_id),
                SkuEconomicsDailyMetric.nm_id > 0,
            )
        )
        min_date, max_date = result.one()
        return min_date, max_date

    def _resolve_requested_period(
        self,
        *,
        days: int,
        period_start: Optional[date],
        period_end: Optional[date],
        preset: Optional[str],
        available_start: Optional[date],
        available_end: Optional[date],
    ) -> Tuple[date, date, str]:
        normalized_preset = self._normalize_preset(preset, days)
        reference_end = available_end or date.today()

        if period_start or period_end:
            resolved_end = period_end or reference_end
            resolved_start = period_start or (
                available_start
                or resolved_end - timedelta(days=max(int(days or 14), 1) - 1)
            )
            if resolved_end < resolved_start:
                resolved_start, resolved_end = resolved_end, resolved_start
            return resolved_start, resolved_end, "custom"

        if normalized_preset == "all":
            if available_start and available_end:
                return available_start, available_end, "all"
            fallback_start, fallback_end = _today_period(days)
            return fallback_start, fallback_end, "all"

        preset_days = {
            "7d": 7,
            "14d": 14,
            "30d": 30,
            "90d": 90,
        }.get(normalized_preset, max(int(days or 14), 1))
        resolved_end = reference_end
        resolved_start = resolved_end - timedelta(days=preset_days - 1)
        return resolved_start, resolved_end, normalized_preset

    def _normalize_preset(self, preset: Optional[str], days: int) -> str:
        raw = str(preset or "").strip().lower()
        if raw in {"all", "custom", "7d", "14d", "30d", "90d"}:
            return raw
        if raw in {"7", "14", "30", "90"}:
            return f"{raw}d"
        safe_days = max(int(days or 14), 1)
        return f"{safe_days}d" if safe_days in {7, 14, 30, 90} else "custom"

    async def _refresh_history_data(
        self,
        db: AsyncSession,
        store: Store,
        *,
        requested_period_start: date,
        requested_period_end: date,
    ) -> None:
        today = date.today()
        history_floor = today - timedelta(days=self.HISTORY_LOOKBACK_DAYS - 1)
        refresh_start = max(requested_period_start, history_floor)
        refresh_end = max(min(requested_period_end, today), refresh_start)

        existing_result = await db.execute(
            select(SkuEconomicsDailyMetric).where(
                SkuEconomicsDailyMetric.store_id == int(store.id),
                SkuEconomicsDailyMetric.metric_date >= refresh_start,
                SkuEconomicsDailyMetric.metric_date <= refresh_end,
            )
        )
        existing_rows = list(existing_result.scalars().all())

        cards_result = await db.execute(
            select(Card.nm_id).where(
                Card.store_id == int(store.id),
                Card.nm_id > 0,
            )
        )
        funnel_nm_ids = sorted({int(nm_id) for nm_id in cards_result.scalars().all() if int(nm_id or 0) > 0})

        funnel_refresh_start = max(refresh_start, refresh_end - timedelta(days=6))
        advert_task = self._fetch_advert_daily_history(store, refresh_start, refresh_end)
        finance_task = self._fetch_finance_daily_history(store, refresh_start, refresh_end)
        funnel_task = self._fetch_funnel_daily_history(
            store,
            funnel_refresh_start,
            refresh_end,
            funnel_nm_ids,
        )
        (advert_rows, advert_status), (finance_rows, finance_status), (funnel_rows, funnel_status) = await asyncio.gather(
            advert_task,
            finance_task,
            funnel_task,
        )

        advert_refresh_ok = advert_status.mode != "error"
        finance_refresh_ok = finance_status.mode not in {"error", "manual_required"}
        funnel_refresh_ok = funnel_status.mode != "error"

        if not advert_refresh_ok and not finance_refresh_ok and not funnel_refresh_ok:
            return

        buckets: Dict[Tuple[date, int], _DailyMetricBucket] = {}

        for row in existing_rows:
            key = (row.metric_date, int(row.nm_id))
            bucket = buckets.setdefault(
                key,
                _DailyMetricBucket(
                    title=str(row.title or ""),
                    vendor_code=str(row.vendor_code or ""),
                ),
            )
            if not advert_refresh_ok:
                bucket.advert_views = int(row.advert_views or 0)
                bucket.advert_clicks = int(row.advert_clicks or 0)
                bucket.advert_orders = int(row.advert_orders or 0)
                bucket.advert_gmv = float(row.advert_gmv or 0.0)
                bucket.advert_exact_spend = float(row.advert_exact_spend or 0.0)
                bucket.advert_estimated_spend = float(row.advert_estimated_spend or 0.0)
                bucket.has_advert = bool(row.has_advert)
            if not finance_refresh_ok:
                bucket.finance_revenue = float(row.finance_revenue or 0.0)
                bucket.finance_payout = float(row.finance_payout or 0.0)
                bucket.finance_wb_costs = float(row.finance_wb_costs or 0.0)
                bucket.finance_orders = int(row.finance_orders or 0)
                bucket.has_finance = bool(row.has_finance)
            if (not funnel_refresh_ok) or row.metric_date < funnel_refresh_start:
                bucket.funnel_open_count = int(row.funnel_open_count or 0)
                bucket.funnel_cart_count = int(row.funnel_cart_count or 0)
                bucket.funnel_order_count = int(row.funnel_order_count or 0)
                bucket.funnel_order_sum = float(row.funnel_order_sum or 0.0)
                bucket.funnel_buyout_count = int(row.funnel_buyout_count or 0)
                bucket.funnel_buyout_sum = float(row.funnel_buyout_sum or 0.0)
                bucket.has_funnel = bool(row.has_funnel)

        for key, fresh in advert_rows.items():
            bucket = buckets.setdefault(key, _DailyMetricBucket())
            self._merge_daily_bucket(bucket, fresh, replace=("advert",))

        for key, fresh in finance_rows.items():
            bucket = buckets.setdefault(key, _DailyMetricBucket())
            self._merge_daily_bucket(bucket, fresh, replace=("finance",))

        if funnel_refresh_ok:
            for key, fresh in funnel_rows.items():
                bucket = buckets.setdefault(key, _DailyMetricBucket())
                self._merge_daily_bucket(bucket, fresh, replace=("funnel",))

        nm_ids = sorted({nm_id for (_, nm_id) in buckets.keys() if int(nm_id) > 0})
        card_meta: Dict[int, Tuple[str, str]] = {}
        if nm_ids:
            cards_result = await db.execute(
                select(Card.nm_id, Card.title, Card.vendor_code).where(
                    Card.store_id == int(store.id),
                    Card.nm_id.in_(nm_ids),
                )
            )
            for nm_id, title, vendor_code in cards_result.all():
                card_meta[int(nm_id)] = (str(title or ""), str(vendor_code or ""))

        await db.execute(
            delete(SkuEconomicsDailyMetric).where(
                SkuEconomicsDailyMetric.store_id == int(store.id),
                SkuEconomicsDailyMetric.metric_date >= refresh_start,
                SkuEconomicsDailyMetric.metric_date <= refresh_end,
            )
        )
        await db.execute(
            delete(SkuEconomicsOverviewCache).where(
                SkuEconomicsOverviewCache.store_id == int(store.id)
            )
        )
        self.invalidate_cache(store.id)

        synced_at = utc_now()
        for (metric_date, nm_id), bucket in sorted(buckets.items(), key=lambda item: (item[0][0], item[0][1])):
            if int(nm_id) > 0 and int(nm_id) in card_meta:
                title, vendor_code = card_meta[int(nm_id)]
                bucket.title = bucket.title or title
                bucket.vendor_code = bucket.vendor_code or vendor_code

            has_payload = any(
                (
                    bucket.advert_views,
                    bucket.advert_clicks,
                    bucket.advert_orders,
                    bucket.advert_gmv,
                    bucket.advert_exact_spend,
                    bucket.advert_estimated_spend,
                    bucket.finance_revenue,
                    bucket.finance_payout,
                    bucket.finance_wb_costs,
                    bucket.finance_orders,
                    bucket.funnel_open_count,
                    bucket.funnel_cart_count,
                    bucket.funnel_order_count,
                    bucket.funnel_order_sum,
                    bucket.funnel_buyout_count,
                    bucket.funnel_buyout_sum,
                )
            ) or bucket.has_advert or bucket.has_finance or bucket.has_funnel
            if not has_payload:
                continue

            db.add(
                SkuEconomicsDailyMetric(
                    store_id=int(store.id),
                    nm_id=int(nm_id),
                    metric_date=metric_date,
                    title=bucket.title or None,
                    vendor_code=bucket.vendor_code or None,
                    advert_views=int(bucket.advert_views or 0),
                    advert_clicks=int(bucket.advert_clicks or 0),
                    advert_orders=int(bucket.advert_orders or 0),
                    advert_gmv=float(bucket.advert_gmv or 0.0),
                    advert_exact_spend=float(bucket.advert_exact_spend or 0.0),
                    advert_estimated_spend=float(bucket.advert_estimated_spend or 0.0),
                    finance_revenue=float(bucket.finance_revenue or 0.0),
                    finance_payout=float(bucket.finance_payout or 0.0),
                    finance_wb_costs=float(bucket.finance_wb_costs or 0.0),
                    finance_orders=int(bucket.finance_orders or 0),
                    funnel_open_count=int(bucket.funnel_open_count or 0),
                    funnel_cart_count=int(bucket.funnel_cart_count or 0),
                    funnel_order_count=int(bucket.funnel_order_count or 0),
                    funnel_order_sum=float(bucket.funnel_order_sum or 0.0),
                    funnel_buyout_count=int(bucket.funnel_buyout_count or 0),
                    funnel_buyout_sum=float(bucket.funnel_buyout_sum or 0.0),
                    has_advert=bool(bucket.has_advert),
                    has_finance=bool(bucket.has_finance),
                    has_funnel=bool(bucket.has_funnel),
                    synced_at=synced_at,
                )
            )

        await db.commit()

    def _merge_daily_bucket(
        self,
        target: _DailyMetricBucket,
        source: _DailyMetricBucket,
        *,
        replace: Sequence[str] = (),
    ) -> None:
        if source.title and not target.title:
            target.title = source.title
        if source.vendor_code and not target.vendor_code:
            target.vendor_code = source.vendor_code

        if "advert" in replace:
            target.advert_views = int(source.advert_views or 0)
            target.advert_clicks = int(source.advert_clicks or 0)
            target.advert_orders = int(source.advert_orders or 0)
            target.advert_gmv = float(source.advert_gmv or 0.0)
            target.advert_exact_spend = float(source.advert_exact_spend or 0.0)
            target.advert_estimated_spend = float(source.advert_estimated_spend or 0.0)
            target.has_advert = bool(source.has_advert)
        if "finance" in replace:
            target.finance_revenue = float(source.finance_revenue or 0.0)
            target.finance_payout = float(source.finance_payout or 0.0)
            target.finance_wb_costs = float(source.finance_wb_costs or 0.0)
            target.finance_orders = int(source.finance_orders or 0)
            target.has_finance = bool(source.has_finance)
        if "funnel" in replace:
            target.funnel_open_count = int(source.funnel_open_count or 0)
            target.funnel_cart_count = int(source.funnel_cart_count or 0)
            target.funnel_order_count = int(source.funnel_order_count or 0)
            target.funnel_order_sum = float(source.funnel_order_sum or 0.0)
            target.funnel_buyout_count = int(source.funnel_buyout_count or 0)
            target.funnel_buyout_sum = float(source.funnel_buyout_sum or 0.0)
            target.has_funnel = bool(source.has_funnel)

    def _extract_fullstats_items(self, payload: Any) -> List[dict]:
        if isinstance(payload, list):
            return [item for item in payload if isinstance(item, dict)]
        if not isinstance(payload, dict):
            return []

        for key in ("data", "items", "adverts", "rows", "result"):
            value = payload.get(key)
            if isinstance(value, list):
                return [item for item in value if isinstance(item, dict)]
            if isinstance(value, dict):
                for inner_key in ("data", "items", "adverts", "rows"):
                    inner_value = value.get(inner_key)
                    if isinstance(inner_value, list):
                        return [item for item in inner_value if isinstance(item, dict)]

        if any(key in payload for key in ("days", "apps", "nms", "advertId", "advert_id", "id", "sum")):
            return [payload]
        return []

    def _extract_value_by_keys(self, payload: Optional[Dict[str, Any]], keys: Sequence[str]) -> Any:
        if not isinstance(payload, dict):
            return None
        for key in keys:
            if key in payload:
                value = payload.get(key)
                if value is not None:
                    return value
        return None

    def _extract_campaign_total(self, node: Dict[str, Any]) -> float:
        raw = self._extract_value_by_keys(node, self.ADVERT_CAMPAIGN_SPEND_KEYS)
        if raw is not None:
            return max(_to_float(raw), 0.0)

        for key in ("stat", "stats", "statistic", "selected"):
            nested = node.get(key)
            raw_nested = self._extract_value_by_keys(nested if isinstance(nested, dict) else None, self.ADVERT_CAMPAIGN_SPEND_KEYS)
            if raw_nested is not None:
                return max(_to_float(raw_nested), 0.0)
        return 0.0

    def _extract_nm_id(self, nm_row: Dict[str, Any]) -> int:
        direct_nm_id = _to_int(self._extract_value_by_keys(nm_row, self.ADVERT_NM_ID_KEYS))
        if direct_nm_id > 0:
            return direct_nm_id

        for key in ("nm", "product", "item"):
            nested = nm_row.get(key)
            if isinstance(nested, dict):
                nested_nm_id = _to_int(self._extract_value_by_keys(nested, self.ADVERT_NM_ID_KEYS))
                if nested_nm_id > 0:
                    return nested_nm_id
            else:
                nested_nm_id = _to_int(nested)
                if nested_nm_id > 0:
                    return nested_nm_id
        return 0

    def _extract_nm_metric_float(self, nm_row: Dict[str, Any], keys: Sequence[str]) -> float:
        raw = self._extract_value_by_keys(nm_row, keys)
        if raw is not None:
            return max(_to_float(raw), 0.0)

        for key in ("stats", "stat", "statistic", "selected"):
            nested = nm_row.get(key)
            if isinstance(nested, dict):
                raw_nested = self._extract_value_by_keys(nested, keys)
                if raw_nested is not None:
                    return max(_to_float(raw_nested), 0.0)
        return 0.0

    def _extract_nm_title(self, nm_row: Dict[str, Any]) -> str:
        raw = self._extract_value_by_keys(nm_row, self.ADVERT_NM_TITLE_KEYS)
        if raw is not None:
            return str(raw or "").strip()
        for key in ("nm", "product", "item"):
            nested = nm_row.get(key)
            if isinstance(nested, dict):
                nested_title = self._extract_value_by_keys(nested, self.ADVERT_NM_TITLE_KEYS)
                if nested_title is not None:
                    return str(nested_title or "").strip()
        return ""

    def _build_residual_weights(self, nm_rows: Sequence[Tuple[int, Dict[str, Any]]]) -> List[Tuple[int, float]]:
        if not nm_rows:
            return []

        order_weights: List[Tuple[int, float]] = [
            (nm_id, self._extract_nm_metric_float(row, self.ADVERT_NM_ORDERS_KEYS))
            for nm_id, row in nm_rows
        ]
        if sum(weight for _, weight in order_weights) > self.ADVERT_RESIDUAL_EPSILON:
            return order_weights

        click_weights: List[Tuple[int, float]] = [
            (nm_id, self._extract_nm_metric_float(row, self.ADVERT_NM_CLICKS_KEYS))
            for nm_id, row in nm_rows
        ]
        if sum(weight for _, weight in click_weights) > self.ADVERT_RESIDUAL_EPSILON:
            return click_weights

        view_weights: List[Tuple[int, float]] = [
            (nm_id, self._extract_nm_metric_float(row, self.ADVERT_NM_VIEWS_KEYS))
            for nm_id, row in nm_rows
        ]
        if sum(weight for _, weight in view_weights) > self.ADVERT_RESIDUAL_EPSILON:
            return view_weights

        return [(nm_id, 1.0) for nm_id, _ in nm_rows]

    async def _fetch_advert_daily_history(
        self,
        store: Store,
        period_start: date,
        period_end: date,
    ) -> Tuple[Dict[Tuple[date, int], _DailyMetricBucket], AdAnalysisSourceStatusOut]:
        tokens = self._candidate_tokens(store, prefer_advert=True)
        last_error = "Не найден токен для рекламного API"
        for token in tokens:
            try:
                repo = WBAdvertRepository(token=token)
                campaign_ids = await repo.get_campaign_ids()
                if not campaign_ids:
                    return (
                        {},
                        AdAnalysisSourceStatusOut(
                            id="advert",
                            label="Реклама WB",
                            mode="empty",
                            detail="Кампании за выбранный период не найдены",
                            records=0,
                            automatic=True,
                        ),
                    )

                raw_items: List[dict] = []
                chunk_start = period_start
                while chunk_start <= period_end:
                    chunk_end = min(chunk_start + timedelta(days=30), period_end)
                    for batch in _chunks(campaign_ids, 50):
                        batch_data = await asyncio.to_thread(
                            repo.get_fullstats,
                            batch,
                            begin_date=chunk_start.isoformat(),
                            end_date=chunk_end.isoformat(),
                        )
                        raw_items.extend(self._extract_fullstats_items(batch_data))
                    chunk_start = chunk_end + timedelta(days=1)

                parsed = self._parse_fullstats_daily(raw_items, fallback_end=period_end)
                mode = "partial" if parsed["unallocated_spend"] > 0.01 else "ok"
                return (
                    parsed["rows"],
                    AdAnalysisSourceStatusOut(
                        id="advert",
                        label="Реклама WB",
                        mode=mode,
                        detail=(
                            "История расходов сохранена, но часть spend осталась без точной привязки к nmID."
                            if mode == "partial"
                            else "История расходов, кликов и заказов сохранена из WB Advert."
                        ),
                        records=len(parsed["rows"]),
                        automatic=True,
                    ),
                )
            except Exception as exc:
                last_error = str(exc)

        return (
            {},
            AdAnalysisSourceStatusOut(
                id="advert",
                label="Реклама WB",
                mode="error",
                detail=last_error,
                records=0,
                automatic=True,
            ),
        )

    def _parse_fullstats_daily(
        self,
        items: Sequence[dict],
        *,
        fallback_end: date,
    ) -> Dict[str, Any]:
        rows: Dict[Tuple[date, int], _DailyMetricBucket] = {}
        unallocated_spend = 0.0
        skipped_items = 0
        invalid_nm_rows = 0
        zero_total_drift_rows = 0

        for item in items:
            if not isinstance(item, dict):
                skipped_items += 1
                continue
            day_nodes = item.get("days")
            if isinstance(day_nodes, list) and day_nodes:
                candidates = [node for node in day_nodes if isinstance(node, dict)]
            else:
                candidates = [item]
            if not candidates:
                skipped_items += 1
                continue

            for node in candidates:
                metric_date = self._parse_generic_date(
                    node.get("date")
                    or node.get("day")
                    or node.get("dt")
                    or node.get("begin")
                    or node.get("beginDate")
                ) or fallback_end

                leaf_nms: List[dict] = []
                self._collect_leaf_nms(node, leaf_nms)
                campaign_total = self._extract_campaign_total(node)
                campaign_spend_from_nms = 0.0
                valid_nm_rows: List[Tuple[int, Dict[str, Any]]] = []

                for nm_row in leaf_nms:
                    if not isinstance(nm_row, dict):
                        invalid_nm_rows += 1
                        continue
                    nm_id = self._extract_nm_id(nm_row)
                    if nm_id <= 0:
                        invalid_nm_rows += 1
                        continue
                    valid_nm_rows.append((nm_id, nm_row))
                    key = (metric_date, nm_id)
                    bucket = rows.setdefault(key, _DailyMetricBucket())
                    nm_spend = self._extract_nm_metric_float(nm_row, self.ADVERT_NM_SPEND_KEYS)
                    bucket.advert_views += _to_int(self._extract_nm_metric_float(nm_row, self.ADVERT_NM_VIEWS_KEYS))
                    bucket.advert_clicks += _to_int(self._extract_nm_metric_float(nm_row, self.ADVERT_NM_CLICKS_KEYS))
                    bucket.advert_orders += _to_int(self._extract_nm_metric_float(nm_row, self.ADVERT_NM_ORDERS_KEYS))
                    bucket.advert_gmv += self._extract_nm_metric_float(nm_row, self.ADVERT_NM_GMV_KEYS)
                    bucket.advert_exact_spend += nm_spend
                    bucket.has_advert = True
                    if not bucket.title:
                        bucket.title = self._extract_nm_title(nm_row)
                    campaign_spend_from_nms += nm_spend

                if campaign_total <= self.ADVERT_RESIDUAL_EPSILON and campaign_spend_from_nms > self.ADVERT_RESIDUAL_EPSILON:
                    zero_total_drift_rows += 1
                effective_campaign_total = max(campaign_total, campaign_spend_from_nms)
                residual = max(effective_campaign_total - campaign_spend_from_nms, 0.0)
                if residual <= self.ADVERT_RESIDUAL_EPSILON:
                    continue

                weights = self._build_residual_weights(valid_nm_rows)
                if weights:
                    total_weight = sum(weight for _, weight in weights) or float(len(weights) or 1)
                    for nm_id, weight in weights:
                        key = (metric_date, nm_id)
                        bucket = rows.setdefault(key, _DailyMetricBucket())
                        bucket.advert_estimated_spend += residual * (weight / total_weight)
                        bucket.has_advert = True
                else:
                    key = (metric_date, 0)
                    bucket = rows.setdefault(key, _DailyMetricBucket())
                    bucket.advert_estimated_spend += residual
                    bucket.has_advert = True
                    unallocated_spend += residual

        if skipped_items > 0 or invalid_nm_rows > 0 or zero_total_drift_rows > 0:
            logger.warning(
                "[sku-advert] fullstats daily parser normalized anomalies skipped_items=%s invalid_nm_rows=%s "
                "zero_total_drift_rows=%s",
                skipped_items,
                invalid_nm_rows,
                zero_total_drift_rows,
            )

        return {
            "rows": rows,
            "unallocated_spend": round(unallocated_spend, 2),
        }

    async def _fetch_finance_daily_history(
        self,
        store: Store,
        period_start: date,
        period_end: date,
    ) -> Tuple[Dict[Tuple[date, int], _DailyMetricBucket], AdAnalysisSourceStatusOut]:
        tokens = self._candidate_tokens(store, prefer_advert=False)
        last_error = "Не найден токен статистики"
        url = f"{settings.WB_STATISTICS_API_URL}/api/v5/supplier/reportDetailByPeriod"
        for token in tokens:
            try:
                rows_by_day: Dict[Tuple[date, int], _DailyMetricBucket] = {}
                total_rows = 0
                chunk_start = period_start
                partial_reason: Optional[str] = None
                async with httpx.AsyncClient(timeout=httpx.Timeout(90.0)) as client:
                    while chunk_start <= period_end:
                        chunk_end = min(chunk_start + timedelta(days=30), period_end)
                        try:
                            rows = await self._fetch_finance_rows_paginated(
                                client=client,
                                url=url,
                                token=token,
                                date_from=chunk_start,
                                date_to=chunk_end,
                                period="daily",
                            )
                        except _FinanceFetchError as finance_exc:
                            rows = list(finance_exc.partial_rows or [])
                            if not rows:
                                raise
                            partial_reason = str(finance_exc)
                            logger.warning(
                                "[sku-finance] partial daily history fetched period=%s..%s rows=%s reason=%s",
                                chunk_start.isoformat(),
                                chunk_end.isoformat(),
                                len(rows),
                                partial_reason,
                            )
                        total_rows += len(rows)
                        for row in rows:
                            nm_id = _to_int(row.get("nm_id"))
                            if nm_id <= 0:
                                continue
                            metric_date = self._parse_generic_date(
                                row.get("date_from")
                                or row.get("sale_dt")
                                or row.get("rr_dt")
                                or row.get("create_dt")
                            )
                            if metric_date is None:
                                metric_date = chunk_end

                            quantity = _to_int(row.get("quantity"))
                            doc_type = str(row.get("doc_type_name") or row.get("supplier_oper_name") or "").lower()
                            negative = "возврат" in doc_type or "return" in doc_type or quantity < 0
                            sign = -1.0 if negative else 1.0
                            revenue = _to_float(row.get("retail_price_withdisc_rub") or row.get("retail_amount"))
                            payout = _to_float(row.get("ppvz_for_pay"))
                            extra = sum(
                                _to_float(row.get(field))
                                for field in (
                                    "delivery_rub",
                                    "acquiring_fee",
                                    "penalty",
                                    "storage_fee",
                                    "deduction",
                                    "acceptance",
                                    "rebill_logistic_cost",
                                    "additional_payment",
                                )
                            )

                            key = (metric_date, nm_id)
                            bucket = rows_by_day.setdefault(key, _DailyMetricBucket())
                            bucket.finance_revenue += revenue * sign if revenue >= 0 else revenue
                            bucket.finance_payout += payout * sign if payout >= 0 else payout
                            # Keep only raw WB extra costs here.
                            # Final revenue/payout normalization and payout delta are applied
                            # after period aggregation so cached history matches live finance fetch.
                            bucket.finance_wb_costs += _safe_positive(extra)
                            bucket.finance_orders += max(quantity * int(sign), 0)
                            bucket.has_finance = True
                            if not bucket.title:
                                bucket.title = str(row.get("sa_name") or row.get("subject_name") or row.get("brand_name") or "").strip()

                        if partial_reason:
                            break
                        chunk_start = chunk_end + timedelta(days=1)

                return (
                    rows_by_day,
                    AdAnalysisSourceStatusOut(
                        id="finance",
                        label="Финансы WB",
                        mode="partial" if partial_reason else ("ok" if total_rows > 0 else "empty"),
                        detail=(
                            f"История финансового слоя сохранена частично: {partial_reason}"
                            if partial_reason
                            else (
                                "История финансового слоя сохранена из WB Statistics API."
                                if total_rows > 0
                                else "За выбранный диапазон финансовых строк нет."
                            )
                        ),
                        records=total_rows,
                        automatic=True,
                    ),
                )
            except Exception as exc:
                last_error = str(exc)

        return (
            {},
            AdAnalysisSourceStatusOut(
                id="finance",
                label="Финансы WB",
                mode="manual_required",
                detail=last_error,
                records=0,
                automatic=True,
            ),
        )

    async def _fetch_funnel_daily_history(
        self,
        store: Store,
        period_start: date,
        period_end: date,
        nm_ids: Sequence[int],
    ) -> Tuple[Dict[Tuple[date, int], _DailyMetricBucket], AdAnalysisSourceStatusOut]:
        if not nm_ids:
            return (
                {},
                AdAnalysisSourceStatusOut(
                    id="funnel",
                    label="Воронка продаж",
                    mode="empty",
                    detail="Нет карточек для запроса воронки.",
                    records=0,
                    automatic=True,
                ),
            )

        recent_floor = date.today() - timedelta(days=6)
        effective_start = max(period_start, recent_floor)
        effective_end = min(period_end, date.today())
        if effective_end < effective_start:
            return (
                {},
                AdAnalysisSourceStatusOut(
                    id="funnel",
                    label="Воронка продаж",
                    mode="partial",
                    detail="WB Analytics отдает дневную воронку только за последние 7 дней. Для более старого периода автоматической детализации нет.",
                    records=0,
                    automatic=True,
                ),
            )

        tokens = self._candidate_tokens(store, prefer_advert=False)
        last_error = "Не найден токен аналитики"
        url = f"{settings.WB_ANALYTICS_API_URL}/api/analytics/v3/sales-funnel/products/history"
        for token in tokens:
            try:
                rows_by_day: Dict[Tuple[date, int], _DailyMetricBucket] = {}
                history_rows = 0
                async with httpx.AsyncClient(timeout=httpx.Timeout(60.0)) as client:
                    for batch in _chunks(list(nm_ids), 20):
                        payload = {
                            "selectedPeriod": {
                                "start": effective_start.isoformat(),
                                "end": effective_end.isoformat(),
                            },
                            "nmIds": batch,
                            "skipDeletedNm": False,
                            "aggregationLevel": "day",
                        }
                        resp = await self._request_with_retry(
                            client,
                            "POST",
                            url,
                            headers={
                                "Authorization": token,
                                "Content-Type": "application/json",
                                "Accept": "application/json",
                            },
                            json=payload,
                        )
                        if resp.status_code >= 400:
                            raise ValueError(f"WB funnel history error {resp.status_code}: {resp.text[:300]}")
                        data = resp.json() or []
                        if isinstance(data, dict):
                            data = data.get("data") or data.get("products") or []
                        if not isinstance(data, list):
                            data = []
                        for row in data:
                            product = row.get("product") or {}
                            nm_id = _to_int(product.get("nmId"))
                            if nm_id <= 0:
                                continue
                            title = str(product.get("title") or "").strip()
                            vendor_code = str(product.get("vendorCode") or "").strip()
                            history = row.get("history") or []
                            if not isinstance(history, list):
                                continue
                            for history_row in history:
                                metric_date = self._parse_generic_date(history_row.get("date")) or effective_end
                                key = (metric_date, nm_id)
                                bucket = rows_by_day.setdefault(key, _DailyMetricBucket())
                                if title and not bucket.title:
                                    bucket.title = title
                                if vendor_code and not bucket.vendor_code:
                                    bucket.vendor_code = vendor_code
                                bucket.funnel_open_count += _to_int(history_row.get("openCount"))
                                bucket.funnel_cart_count += _to_int(history_row.get("cartCount"))
                                bucket.funnel_order_count += _to_int(history_row.get("orderCount"))
                                bucket.funnel_order_sum += _to_float(history_row.get("orderSum"))
                                bucket.funnel_buyout_count += _to_int(history_row.get("buyoutCount"))
                                bucket.funnel_buyout_sum += _to_float(history_row.get("buyoutSum"))
                                bucket.has_funnel = True
                                history_rows += 1

                return (
                    rows_by_day,
                    AdAnalysisSourceStatusOut(
                        id="funnel",
                        label="Воронка продаж",
                        mode="ok" if history_rows > 0 else "empty",
                        detail=(
                            f"История воронки из WB Analytics сохранена за {effective_start.isoformat()} — {effective_end.isoformat()}."
                            if history_rows > 0
                            else "За последние 7 дней WB Analytics не вернул строки воронки."
                        ),
                        records=history_rows,
                        automatic=True,
                    ),
                )
            except Exception as exc:
                last_error = str(exc)

        return (
            {},
            AdAnalysisSourceStatusOut(
                id="funnel",
                label="Воронка продаж",
                mode="error",
                detail=last_error,
                records=0,
                automatic=True,
            ),
        )

    def _parse_generic_date(self, value: Any) -> Optional[date]:
        if value is None:
            return None
        if isinstance(value, date) and not isinstance(value, datetime):
            return value
        if isinstance(value, datetime):
            return value.date()
        raw = str(value).strip()
        if not raw:
            return None
        candidate = raw[:10]
        try:
            return date.fromisoformat(candidate)
        except Exception:
            return None

    async def _load_history_aggregate(
        self,
        db: AsyncSession,
        *,
        store_id: int,
        period_start: date,
        period_end: date,
    ) -> Dict[str, Any]:
        if period_end < period_start:
            return {
                "per_nm": {},
                "advert_nm_ids": [],
                "exact_spend": 0.0,
                "estimated_spend": 0.0,
                "unallocated_spend": 0.0,
                "advert_records": 0,
                "finance_records": 0,
                "funnel_records": 0,
            }

        result = await db.execute(
            select(SkuEconomicsDailyMetric).where(
                SkuEconomicsDailyMetric.store_id == int(store_id),
                SkuEconomicsDailyMetric.metric_date >= period_start,
                SkuEconomicsDailyMetric.metric_date <= period_end,
            )
        )
        rows = list(result.scalars().all())
        per_nm: Dict[int, Dict[str, Any]] = {}
        exact_spend = 0.0
        estimated_spend = 0.0
        unallocated_spend = 0.0
        advert_records = 0
        finance_records = 0
        funnel_records = 0
        advert_nm_ids: set[int] = set()
        generated_at: Optional[datetime] = None

        for row in rows:
            if row.synced_at and (generated_at is None or row.synced_at > generated_at):
                generated_at = row.synced_at
            nm_id = int(row.nm_id or 0)
            if nm_id <= 0:
                unallocated_spend += float(row.advert_exact_spend or 0.0) + float(row.advert_estimated_spend or 0.0)
                continue

            record = per_nm.setdefault(
                nm_id,
                {
                    "title": str(row.title or ""),
                    "vendor_code": str(row.vendor_code or ""),
                    "advert": _AdvertMetrics(),
                    "finance": _FinanceMetrics(),
                    "funnel": _FunnelMetrics(),
                    "has_advert": False,
                    "has_finance": False,
                    "has_funnel": False,
                },
            )
            if row.title and not record["title"]:
                record["title"] = str(row.title)
            if row.vendor_code and not record["vendor_code"]:
                record["vendor_code"] = str(row.vendor_code)

            advert = record["advert"]
            advert.views += int(row.advert_views or 0)
            advert.clicks += int(row.advert_clicks or 0)
            advert.orders += int(row.advert_orders or 0)
            advert.gmv += float(row.advert_gmv or 0.0)
            advert.exact_spend += float(row.advert_exact_spend or 0.0)
            advert.estimated_spend += float(row.advert_estimated_spend or 0.0)

            finance = record["finance"]
            finance.revenue += float(row.finance_revenue or 0.0)
            finance.payout += float(row.finance_payout or 0.0)
            finance.wb_costs += float(row.finance_wb_costs or 0.0)
            finance.orders += int(row.finance_orders or 0)

            funnel = record["funnel"]
            funnel.open_count += int(row.funnel_open_count or 0)
            funnel.cart_count += int(row.funnel_cart_count or 0)
            funnel.order_count += int(row.funnel_order_count or 0)
            funnel.order_sum += float(row.funnel_order_sum or 0.0)
            funnel.buyout_count += int(row.funnel_buyout_count or 0)
            funnel.buyout_sum += float(row.funnel_buyout_sum or 0.0)

            if row.has_advert:
                record["has_advert"] = True
                advert_records += 1
                advert_nm_ids.add(nm_id)
            if row.has_finance:
                record["has_finance"] = True
                finance_records += 1
            if row.has_funnel:
                record["has_funnel"] = True
                funnel_records += 1

            exact_spend += float(row.advert_exact_spend or 0.0)
            estimated_spend += float(row.advert_estimated_spend or 0.0)

        for record in per_nm.values():
            finance = record["finance"]
            signed_revenue = float(finance.revenue or 0.0)
            signed_payout = float(finance.payout or 0.0)
            finance.wb_costs = round(max(float(finance.wb_costs or 0.0), 0.0) + max(signed_revenue - signed_payout, 0.0), 2)
            finance.revenue = round(max(signed_revenue, 0.0), 2)
            finance.payout = round(signed_payout, 2)
            finance.orders = max(int(finance.orders or 0), 0)

            funnel = record["funnel"]
            funnel.add_to_cart_percent = (
                round((funnel.cart_count / funnel.open_count) * 100.0, 2)
                if funnel.open_count > 0
                else 0.0
            )
            funnel.cart_to_order_percent = (
                round((funnel.order_count / funnel.cart_count) * 100.0, 2)
                if funnel.cart_count > 0
                else 0.0
            )

        return {
            "per_nm": per_nm,
            "advert_nm_ids": sorted(advert_nm_ids),
            "exact_spend": round(exact_spend, 2),
            "estimated_spend": round(estimated_spend, 2),
            "unallocated_spend": round(unallocated_spend, 2),
            "advert_records": advert_records,
            "finance_records": finance_records,
            "funnel_records": funnel_records,
            "generated_at": generated_at,
        }

    async def _load_manual_spend_overlap(
        self,
        db: AsyncSession,
        store_id: int,
        period_start: date,
        period_end: date,
    ) -> Dict[int, _AdvertMetrics]:
        result = await db.execute(
            select(SkuEconomicsManualSpend).where(
                SkuEconomicsManualSpend.store_id == int(store_id),
                SkuEconomicsManualSpend.period_start <= period_end,
                SkuEconomicsManualSpend.period_end >= period_start,
            )
        )
        out: Dict[int, _AdvertMetrics] = defaultdict(_AdvertMetrics)
        for row in result.scalars().all():
            overlap_start = max(period_start, row.period_start)
            overlap_end = min(period_end, row.period_end)
            if overlap_end < overlap_start:
                continue
            total_days = max((row.period_end - row.period_start).days + 1, 1)
            overlap_days = max((overlap_end - overlap_start).days + 1, 1)
            ratio = overlap_days / total_days
            metric = out[int(row.nm_id)]
            metric.manual_spend += float(row.spend or 0.0) * ratio
            metric.manual_views += int(round(float(row.views or 0) * ratio))
            metric.manual_clicks += int(round(float(row.clicks or 0) * ratio))
            metric.manual_orders += int(round(float(row.orders or 0) * ratio))
            metric.manual_gmv += float(row.gmv or 0.0) * ratio
        return out

    async def _load_manual_finance_overlap(
        self,
        db: AsyncSession,
        store_id: int,
        period_start: date,
        period_end: date,
    ) -> Dict[int, _FinanceMetrics]:
        result = await db.execute(
            select(SkuEconomicsManualFinance).where(
                SkuEconomicsManualFinance.store_id == int(store_id),
                SkuEconomicsManualFinance.period_start <= period_end,
                SkuEconomicsManualFinance.period_end >= period_start,
            )
        )
        out: Dict[int, _FinanceMetrics] = defaultdict(_FinanceMetrics)
        for row in result.scalars().all():
            overlap_start = max(period_start, row.period_start)
            overlap_end = min(period_end, row.period_end)
            if overlap_end < overlap_start:
                continue
            total_days = max((row.period_end - row.period_start).days + 1, 1)
            overlap_days = max((overlap_end - overlap_start).days + 1, 1)
            ratio = overlap_days / total_days
            metric = out[int(row.nm_id)]
            metric.revenue += float(row.revenue or 0.0) * ratio
            metric.wb_costs += float(row.wb_costs or 0.0) * ratio
            metric.payout += float(row.payout or 0.0) * ratio
            metric.orders += int(round(float(row.orders or 0) * ratio))
        return out

    async def _build_overview_from_history(
        self,
        db: AsyncSession,
        store: Store,
        *,
        period_start: date,
        period_end: date,
        previous_period_start: date,
        previous_period_end: date,
        current_data: Dict[str, Any],
        previous_data: Dict[str, Any],
        selected_preset: str,
        page: int,
        page_size: int,
        status_filter: Optional[str],
        search: Optional[str],
        available_period_start: Optional[date],
        available_period_end: Optional[date],
    ) -> AdAnalysisOverviewOut:
        safe_page = max(int(page or 1), 1)
        safe_page_size = max(1, min(int(page_size or 50), self.MAX_PAGE_SIZE))

        manual_costs = await self._load_costs(db, store.id)
        manual_spend = await self._load_manual_spend_overlap(db, store.id, period_start, period_end)
        manual_finance = await self._load_manual_finance_overlap(db, store.id, period_start, period_end)
        previous_manual_spend = await self._load_manual_spend_overlap(db, store.id, previous_period_start, previous_period_end)
        previous_manual_finance = await self._load_manual_finance_overlap(db, store.id, previous_period_start, previous_period_end)
        funnel_coverage = await self._get_source_coverage(db, store.id, "funnel")

        current_nm_ids = sorted(
            set(current_data.get("advert_nm_ids", []))
            | set(manual_spend.keys())
        )
        previous_nm_ids = sorted(
            set(previous_data.get("advert_nm_ids", []))
            | set(previous_manual_spend.keys())
        )

        cards_map, issue_map = await self._load_cards_and_issues(db, store.id, current_nm_ids)

        def build_items(
            aggregate: Dict[str, Any],
            spend_topups: Dict[int, _AdvertMetrics],
            finance_topups: Dict[int, _FinanceMetrics],
            *,
            include_cards: bool,
        ) -> Tuple[List[AdAnalysisItemOut], List[int]]:
            nm_ids = sorted(set(aggregate.get("advert_nm_ids", [])) | set(spend_topups.keys()))
            built: List[AdAnalysisItemOut] = []
            missing_cost_nm_ids: List[int] = []

            for nm_id in nm_ids:
                base = aggregate["per_nm"].get(
                    nm_id,
                    {
                        "title": "",
                        "vendor_code": "",
                        "advert": _AdvertMetrics(),
                        "finance": _FinanceMetrics(),
                        "funnel": _FunnelMetrics(),
                        "has_advert": False,
                        "has_finance": False,
                        "has_funnel": False,
                    },
                )
                advert = _AdvertMetrics(
                    views=base["advert"].views,
                    clicks=base["advert"].clicks,
                    orders=base["advert"].orders,
                    gmv=base["advert"].gmv,
                    exact_spend=base["advert"].exact_spend,
                    estimated_spend=base["advert"].estimated_spend,
                )
                if nm_id in spend_topups:
                    topup = spend_topups[nm_id]
                    advert.manual_spend += topup.manual_spend
                    advert.manual_views += topup.manual_views
                    advert.manual_clicks += topup.manual_clicks
                    advert.manual_orders += topup.manual_orders
                    advert.manual_gmv += topup.manual_gmv

                finance = _FinanceMetrics(
                    revenue=base["finance"].revenue,
                    payout=base["finance"].payout,
                    wb_costs=base["finance"].wb_costs,
                    orders=base["finance"].orders,
                )
                if nm_id in finance_topups:
                    finance = finance_topups[nm_id]

                funnel_metric = _FunnelMetrics(
                    open_count=base["funnel"].open_count,
                    cart_count=base["funnel"].cart_count,
                    order_count=base["funnel"].order_count,
                    order_sum=base["funnel"].order_sum,
                    buyout_count=base["funnel"].buyout_count,
                    buyout_sum=base["funnel"].buyout_sum,
                    add_to_cart_percent=base["funnel"].add_to_cart_percent,
                    cart_to_order_percent=base["funnel"].cart_to_order_percent,
                )

                card = cards_map.get(nm_id) if include_cards else None
                issues = issue_map.get(card.id if card else -1, []) if include_cards and card else []
                issue_summary = self._summarize_issues(issues)
                total_orders = max(finance.orders, funnel_metric.order_count, advert.total_orders)
                revenue = finance.revenue or funnel_metric.order_sum or advert.total_gmv
                wb_costs = finance.wb_costs
                unit_cost = manual_costs.get(nm_id)
                missing_cost = unit_cost is None
                if unit_cost is None:
                    unit_cost = 0.0
                    if advert.total_spend > 0.01 or revenue > 0.01:
                        missing_cost_nm_ids.append(nm_id)

                cost_price = unit_cost * max(total_orders, 0)
                gross_profit_before_ads = revenue - wb_costs - cost_price
                ad_cost = advert.total_spend
                net_profit = gross_profit_before_ads - ad_cost

                order_denominator = max(total_orders, 1)
                ad_orders_denominator = max(advert.total_orders, 1)
                max_cpo = gross_profit_before_ads / order_denominator if total_orders > 0 else gross_profit_before_ads
                actual_cpo = ad_cost / ad_orders_denominator if advert.total_orders > 0 else ad_cost
                profit_delta = max_cpo - actual_cpo
                ctr = (advert.total_clicks / advert.total_views * 100.0) if advert.total_views > 0 else 0.0
                cr = (advert.total_orders / advert.total_clicks * 100.0) if advert.total_clicks > 0 else 0.0
                cpc = (ad_cost / advert.total_clicks) if advert.total_clicks > 0 else 0.0
                drr = (ad_cost / revenue * 100.0) if revenue > 0 else 0.0
                precision, precision_label = self._resolve_precision(advert, aggregate["unallocated_spend"])
                ad_cost_confidence = self._resolve_ad_cost_confidence(advert, aggregate["unallocated_spend"])

                metrics = AdAnalysisMetricsOut(
                    revenue=round(revenue, 2),
                    wb_costs=round(wb_costs, 2),
                    cost_price=round(cost_price, 2),
                    gross_profit_before_ads=round(gross_profit_before_ads, 2),
                    ad_cost=round(ad_cost, 2),
                    ad_cost_total=round(ad_cost, 2),
                    ad_cost_exact=round(advert.exact_spend, 2),
                    ad_cost_estimated=round(advert.estimated_spend, 2),
                    ad_cost_manual=round(advert.manual_spend, 2),
                    ad_cost_source_mode=precision,
                    ad_cost_confidence=ad_cost_confidence,
                    net_profit=round(net_profit, 2),
                    profit_per_order=round(net_profit / order_denominator, 2) if total_orders > 0 else round(net_profit, 2),
                    max_cpo=round(max_cpo, 2),
                    actual_cpo=round(actual_cpo, 2),
                    profit_delta=round(profit_delta, 2),
                    views=advert.total_views,
                    clicks=advert.total_clicks,
                    ad_orders=advert.total_orders,
                    ad_gmv=round(advert.total_gmv, 2),
                    ctr=round(ctr, 2),
                    cr=round(cr, 2),
                    open_count=funnel_metric.open_count,
                    cart_count=funnel_metric.cart_count,
                    order_count=funnel_metric.order_count,
                    buyout_count=funnel_metric.buyout_count,
                    add_to_cart_percent=round(funnel_metric.add_to_cart_percent, 2),
                    cart_to_order_percent=round(funnel_metric.cart_to_order_percent, 2),
                    cpc=round(cpc, 2),
                    drr=round(drr, 2),
                )

                finance_ready = aggregate["finance_records"] > 0 or nm_id in finance_topups
                diagnosis, diagnosis_label, reasons, hints = self._resolve_diagnosis(
                    metrics=metrics,
                    issue_summary=issue_summary,
                    missing_cost=missing_cost,
                    finance_ready=finance_ready,
                )
                status, status_label = self._resolve_status(
                    metrics=metrics,
                    diagnosis=diagnosis,
                    missing_cost=missing_cost,
                    finance_ready=finance_ready,
                )
                priority, priority_label = self._resolve_priority(
                    status=status,
                    diagnosis=diagnosis,
                    metrics=metrics,
                    trend_signal="no_history",
                )
                source_status_stub = {
                    "advert": AdAnalysisSourceStatusOut(
                        id="advert",
                        label="Реклама WB",
                        mode="partial" if aggregate["unallocated_spend"] > 0.01 else ("ok" if advert.total_spend > 0.01 else "empty"),
                        detail=None,
                        records=0,
                        automatic=True,
                    ),
                    "finance": AdAnalysisSourceStatusOut(
                        id="finance",
                        label="Финансы WB",
                        mode="ok" if finance.revenue > 0.01 or finance.wb_costs > 0.01 else "manual_required",
                        detail=None,
                        records=0,
                        automatic=True,
                    ),
                    "funnel": AdAnalysisSourceStatusOut(
                        id="funnel",
                        label="Воронка продаж",
                        mode="ok" if funnel_metric.open_count > 0 or funnel_metric.order_count > 0 else "empty",
                        detail=None,
                        records=0,
                        automatic=True,
                    ),
                }
                source_lineage = self._build_source_lineage(source_status_stub)
                action_title, action_description, steps, insights, risk_flags = self._build_actions(
                    status=status,
                    diagnosis=diagnosis,
                    issue_summary=issue_summary,
                    missing_cost=missing_cost,
                    precision=precision,
                    source_statuses=source_status_stub,
                    metrics=metrics,
                )

                title = (
                    (card.title if card and card.title else None)
                    or (base["title"] or None)
                    or f"nmID {nm_id}"
                )
                vendor_code = (card.vendor_code if card and card.vendor_code else None) or (base["vendor_code"] or None)
                photo_url = card.photos[0] if card and getattr(card, "photos", None) else None

                built.append(
                    AdAnalysisItemOut(
                        nm_id=nm_id,
                        card_id=card.id if card else None,
                        title=title,
                        vendor_code=vendor_code,
                        photo_url=photo_url,
                        wb_link=f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
                        workspace_link=f"/workspace/cards/{card.id}" if card else None,
                        price=card.price if card else None,
                        card_score=card.score if card else None,
                        status=status,
                        status_label=status_label,
                        diagnosis=diagnosis,
                        diagnosis_label=diagnosis_label,
                        status_reason=reasons,
                        status_hint=hints,
                        action_title=action_title,
                        action_description=action_description,
                        priority=priority,
                        priority_label=priority_label,
                        precision=precision,
                        precision_label=precision_label,
                        source_lineage=source_lineage,
                        trend=AdAnalysisTrendOut(),
                        issue_summary=AdAnalysisIssueSummaryOut(
                            total=issue_summary.total,
                            critical=issue_summary.critical,
                            warnings=issue_summary.warnings,
                            photos=issue_summary.photos,
                            price=issue_summary.price,
                            text=issue_summary.text,
                            docs=issue_summary.docs,
                            top_titles=issue_summary.top_titles,
                        ),
                        metrics=metrics,
                        spend_sources={
                            "exact": round(advert.exact_spend, 2),
                            "estimated": round(advert.estimated_spend, 2),
                            "manual": round(advert.manual_spend, 2),
                        },
                        insights=insights,
                        steps=steps,
                        risk_flags=risk_flags,
                    )
                )

            built.sort(key=self._item_sort_key)
            return built, missing_cost_nm_ids

        all_items, missing_cost_nm_ids = build_items(current_data, manual_spend, manual_finance, include_cards=True)
        previous_items, _ = build_items(previous_data, previous_manual_spend, previous_manual_finance, include_cards=False)

        status_counts = {status: 0 for status in ("stop", "rescue", "control", "grow", "low_data")}
        profitable_count = 0
        problematic_count = 0
        loss_count = 0
        total_revenue = 0.0
        total_ad_spend = 0.0
        total_net_profit = 0.0

        for item in all_items:
            status_counts[item.status] = status_counts.get(item.status, 0) + 1
            total_revenue += item.metrics.revenue
            total_ad_spend += item.metrics.ad_cost
            total_net_profit += item.metrics.net_profit
            if item.status in {"grow", "control"}:
                profitable_count += 1
            elif item.status == "stop":
                loss_count += 1
            else:
                problematic_count += 1

        remaining_unallocated = max(
            float(current_data["unallocated_spend"]) - sum(metric.manual_spend for metric in manual_spend.values()),
            0.0,
        )

        advert_status = AdAnalysisSourceStatusOut(
            id="advert",
            label="Реклама WB",
            mode="partial" if current_data["advert_records"] > 0 and remaining_unallocated > 0.01 else ("ok" if current_data["advert_records"] > 0 else "empty"),
            detail=(
                "Историческая реклама сохранена. Часть spend осталась без точной привязки к nmID."
                if current_data["advert_records"] > 0 and remaining_unallocated > 0.01
                else "Исторический слой WB Advert сохранен в backend и используется для любого периода."
                if current_data["advert_records"] > 0
                else "История WB Advert еще не сохранена. Нажмите «Обновить»."
            ),
            records=int(current_data["advert_records"]),
            automatic=True,
        )
        finance_status = AdAnalysisSourceStatusOut(
            id="finance",
            label="Финансы WB",
            mode=(
                "partial"
                if current_data["finance_records"] > 0 and manual_finance
                else "ok"
                if current_data["finance_records"] > 0
                else "manual"
                if manual_finance
                else "manual_required"
            ),
            detail=(
                "Финансовый слой собран из истории WB и ручного файла."
                if current_data["finance_records"] > 0 and manual_finance
                else "Исторический финансовый слой сохранен в backend и фильтруется по периоду."
                if current_data["finance_records"] > 0
                else "Используется ручной финансовый файл за выбранный период."
                if manual_finance
                else "Исторического финансового слоя нет. Нажмите «Обновить» или загрузите файл."
            ),
            records=max(int(current_data["finance_records"]), len(manual_finance)),
            automatic=not bool(manual_finance),
        )
        funnel_status = AdAnalysisSourceStatusOut(
            id="funnel",
            label="Воронка продаж",
            mode=(
                "empty"
                if funnel_coverage["records"] <= 0
                else "ok"
                if funnel_coverage["start"] and funnel_coverage["end"] and funnel_coverage["start"] <= period_start and funnel_coverage["end"] >= period_end
                else "partial"
            ),
            detail=(
                "Исторический слой воронки сохранен и фильтруется по периоду."
                if funnel_coverage["records"] > 0
                and funnel_coverage["start"]
                and funnel_coverage["end"]
                and funnel_coverage["start"] <= period_start
                and funnel_coverage["end"] >= period_end
                else (
                    f"Воронка сохранена только за {funnel_coverage['start']} — {funnel_coverage['end']}. "
                    "WB Analytics отдает дневную детализацию только за последние 7 дней, поэтому старый период покрыт не полностью."
                )
                if funnel_coverage["records"] > 0 and funnel_coverage["start"] and funnel_coverage["end"]
                else "История воронки еще не загружалась. Нажмите «Обновить данные»."
            ),
            records=int(current_data["funnel_records"] or funnel_coverage["records"]),
            automatic=True,
        )
        costs_status = AdAnalysisSourceStatusOut(
            id="costs",
            label="Себестоимость",
            mode=(
                "partial"
                if manual_costs and missing_cost_nm_ids
                else "manual"
                if manual_costs
                else "manual_required"
            ),
            detail=(
                "Часть SKU покрыта файлом себестоимости, но не все позиции."
                if manual_costs and missing_cost_nm_ids
                else "Используется загруженный файл себестоимости."
                if manual_costs
                else "WB не отдает себестоимость. Для точной экономики нужен файл."
            ),
            records=len(manual_costs),
            automatic=False,
        )
        manual_spend_status = AdAnalysisSourceStatusOut(
            id="manual_spend",
            label="Ручное распределение рекламы",
            mode="manual" if manual_spend else "empty",
            detail=(
                "Ручной файл использован для выбранного периода."
                if manual_spend
                else "Нужен только если у WB останутся нераспределенные расходы."
            ),
            records=len(manual_spend),
            automatic=False,
        )

        alerts = self._build_alerts(
            remaining_unallocated=remaining_unallocated,
            missing_cost_nm_ids=missing_cost_nm_ids,
            finance_status=finance_status,
            advert_status=advert_status,
            items=all_items,
        )
        budget_moves = self._build_budget_moves(all_items)

        previous_overview = AdAnalysisOverviewOut(
            store_id=int(store.id),
            generated_at=previous_data.get("generated_at") or utc_now(),
            snapshot_ready=True,
            period_start=previous_period_start,
            period_end=previous_period_end,
            upload_needs=AdAnalysisUploadNeedsOut(
                period_start=previous_period_start,
                period_end=previous_period_end,
            ),
            items=previous_items,
        )

        overview = AdAnalysisOverviewOut(
            store_id=int(store.id),
            generated_at=current_data.get("generated_at") or utc_now(),
            snapshot_ready=bool(all_items),
            period_start=period_start,
            period_end=period_end,
            available_period_start=available_period_start,
            available_period_end=available_period_end,
            previous_period_start=previous_period_start,
            previous_period_end=previous_period_end,
            selected_preset=selected_preset,
            total_skus=len(all_items),
            total_revenue=round(total_revenue, 2),
            total_ad_spend=round(total_ad_spend, 2),
            total_net_profit=round(total_net_profit, 2),
            exact_spend=round(float(current_data["exact_spend"]), 2),
            estimated_spend=round(float(current_data["estimated_spend"]), 2),
            manual_spend=round(sum(metric.manual_spend for metric in manual_spend.values()), 2),
            unallocated_spend=round(remaining_unallocated, 2),
            profitable_count=profitable_count,
            problematic_count=problematic_count,
            loss_count=loss_count,
            status_counts=status_counts,
            source_statuses=[advert_status, funnel_status, finance_status, costs_status, manual_spend_status],
            source_lineage=self._build_source_lineage([advert_status, funnel_status, finance_status]),
            alerts=alerts,
            budget_moves=budget_moves,
            campaigns=[],
            upload_needs=AdAnalysisUploadNeedsOut(
                period_start=period_start,
                period_end=period_end,
                missing_costs_count=len(sorted(set(missing_cost_nm_ids))),
                missing_cost_nm_ids=sorted(set(missing_cost_nm_ids))[:20],
                needs_manual_spend=remaining_unallocated > 0.01,
                needs_manual_finance=finance_status.mode in {"error", "manual_required"},
                can_upload_costs=True,
                can_upload_manual_spend=True,
                can_upload_manual_finance=True,
            ),
            items=list(all_items),
        )
        self._apply_trends(overview, previous_overview)

        query = str(search or "").strip().lower()
        filtered_items = [
            item
            for item in overview.items
            if (not status_filter or status_filter == "all" or item.status == status_filter)
            and (
                not query
                or any(
                    query in part.lower()
                    for part in (
                        item.title or "",
                        item.vendor_code or "",
                        str(item.nm_id),
                        item.diagnosis_label or "",
                    )
                )
            )
        ]
        total_items = len(filtered_items)
        total_pages = (total_items + safe_page_size - 1) // safe_page_size if total_items else 0
        if total_pages and safe_page > total_pages:
            safe_page = total_pages
        if total_pages == 0:
            safe_page = 1
        offset = max((safe_page - 1) * safe_page_size, 0)

        saved_overview = await self._load_persisted_overview(db, store.id, period_start, period_end)
        if saved_overview and saved_overview.campaigns:
            overview.campaigns = saved_overview.campaigns

        overview.critical_preview = [item for item in overview.items if item.status in {"stop", "rescue"}][:4]
        overview.growth_preview = [item for item in overview.items if item.status == "grow"][:4]
        overview.page = safe_page
        overview.page_size = safe_page_size
        overview.total_items = total_items
        overview.total_pages = total_pages
        overview.items = filtered_items[offset:offset + safe_page_size]
        return overview

    async def _get_source_coverage(
        self,
        db: AsyncSession,
        store_id: int,
        source_id: str,
    ) -> Dict[str, Any]:
        source_column = {
            "advert": SkuEconomicsDailyMetric.has_advert,
            "finance": SkuEconomicsDailyMetric.has_finance,
            "funnel": SkuEconomicsDailyMetric.has_funnel,
        }.get(str(source_id))
        if source_column is None:
            return {"start": None, "end": None, "records": 0}
        result = await db.execute(
            select(
                func.min(SkuEconomicsDailyMetric.metric_date),
                func.max(SkuEconomicsDailyMetric.metric_date),
                func.count(),
            ).where(
                SkuEconomicsDailyMetric.store_id == int(store_id),
                SkuEconomicsDailyMetric.nm_id > 0,
                source_column.is_(True),
            )
        )
        start, end, records = result.one()
        return {
            "start": start,
            "end": end,
            "records": int(records or 0),
        }

    async def _load_persisted_overview(
        self,
        db: AsyncSession,
        store_id: int,
        period_start: date,
        period_end: date,
    ) -> Optional[AdAnalysisOverviewOut]:
        result = await db.execute(
            select(SkuEconomicsOverviewCache).where(
                SkuEconomicsOverviewCache.store_id == int(store_id),
                SkuEconomicsOverviewCache.period_start == period_start,
                SkuEconomicsOverviewCache.period_end == period_end,
            )
        )
        row = result.scalar_one_or_none()
        if row is None or not isinstance(row.payload, dict):
            return None
        payload = dict(row.payload)
        payload.setdefault("snapshot_ready", True)
        return AdAnalysisOverviewOut.model_validate(payload)

    async def _persist_overview(
        self,
        db: AsyncSession,
        store_id: int,
        overview: AdAnalysisOverviewOut,
    ) -> None:
        result = await db.execute(
            select(SkuEconomicsOverviewCache).where(
                SkuEconomicsOverviewCache.store_id == int(store_id),
                SkuEconomicsOverviewCache.period_start == overview.period_start,
                SkuEconomicsOverviewCache.period_end == overview.period_end,
            )
        )
        row = result.scalar_one_or_none()
        if row is None:
            row = SkuEconomicsOverviewCache(
                store_id=int(store_id),
                period_start=overview.period_start,
                period_end=overview.period_end,
            )
            db.add(row)
        row.payload = overview.model_dump(mode="json")
        row.generated_at = overview.generated_at
        await db.commit()

    async def _build_empty_overview(
        self,
        db: AsyncSession,
        *,
        store_id: int,
        period_start: date,
        period_end: date,
        previous_period_start: date,
        previous_period_end: date,
    ) -> AdAnalysisOverviewOut:
        manual_costs = await self._load_costs(db, store_id)
        manual_spend = await self._load_manual_spend(db, store_id, period_start, period_end)
        manual_finance = await self._load_manual_finance(db, store_id, period_start, period_end)

        source_statuses = [
            AdAnalysisSourceStatusOut(
                id="advert",
                label="Реклама WB",
                mode="empty",
                detail="История рекламы еще не загружалась. Нажмите «Обновить данные», чтобы сохранить слой WB Advert в backend.",
                records=0,
                automatic=True,
            ),
            AdAnalysisSourceStatusOut(
                id="funnel",
                label="Воронка продаж",
                mode="empty",
                detail="История воронки еще не загружалась. Нажмите «Обновить данные», чтобы сохранить слой WB Analytics в backend.",
                records=0,
                automatic=True,
            ),
            AdAnalysisSourceStatusOut(
                id="finance",
                label="Финансы WB",
                mode="manual" if manual_finance else "empty",
                detail=(
                    "Используется ранее загруженный финансовый файл. Для полного снимка нажмите «Обновить»."
                    if manual_finance
                    else "История финансов еще не загружалась. Нажмите «Обновить данные», чтобы сохранить слой WB в backend."
                ),
                records=len(manual_finance),
                automatic=not bool(manual_finance),
            ),
            AdAnalysisSourceStatusOut(
                id="costs",
                label="Себестоимость",
                mode="manual" if manual_costs else "manual_required",
                detail=(
                    "Используется загруженный файл себестоимости."
                    if manual_costs
                    else "WB не отдает себестоимость. Загрузите файл и после этого обновите snapshot."
                ),
                records=len(manual_costs),
                automatic=False,
            ),
            AdAnalysisSourceStatusOut(
                id="manual_spend",
                label="Ручное распределение рекламы",
                mode="manual" if manual_spend else "empty",
                detail=(
                    "Есть ручное распределение расходов за выбранный период."
                    if manual_spend
                    else "Понадобится только если после обновления останутся нераспределенные расходы."
                ),
                records=len(manual_spend),
                automatic=False,
            ),
        ]

        alerts = [
            AdAnalysisAlertOut(
                level="info",
                title="Снимок еще не собран",
                description="При открытии страницы показывается сохраненный backend snapshot. Новый расчет и запросы в WB запускаются только по кнопке «Обновить».",
                action="Обновить",
            )
        ]
        if not manual_costs:
            alerts.append(
                AdAnalysisAlertOut(
                    level="warning",
                    title="Себестоимость пока не загружена",
                    description="Без файла себестоимости чистая прибыль и Max CPO считаются неточно.",
                    action="Загрузить себестоимость",
                )
            )

        return AdAnalysisOverviewOut(
            store_id=int(store_id),
            generated_at=utc_now(),
            snapshot_ready=False,
            period_start=period_start,
            period_end=period_end,
            previous_period_start=previous_period_start,
            previous_period_end=previous_period_end,
            source_statuses=source_statuses,
            source_lineage=self._build_source_lineage(source_statuses),
            alerts=alerts,
            upload_needs=AdAnalysisUploadNeedsOut(
                period_start=period_start,
                period_end=period_end,
                missing_costs_count=0,
                missing_cost_nm_ids=[],
                needs_manual_spend=False,
                needs_manual_finance=False,
                can_upload_costs=True,
                can_upload_manual_spend=True,
                can_upload_manual_finance=True,
            ),
            items=[],
        )

    async def _build_period_overview(
        self,
        db: AsyncSession,
        store: Store,
        *,
        period_start: date,
        period_end: date,
        force: bool = False,
    ) -> AdAnalysisOverviewOut:
        cache_key = (int(store.id), period_start.isoformat(), period_end.isoformat())
        cached = self._cache.get(cache_key)
        if cached and not force and (time.time() - cached[0]) <= self.CACHE_TTL_SEC:
            return cached[1]

        source_map = {
            "advert": AdAnalysisSourceStatusOut(id="advert", label="Реклама WB", mode="empty", detail="Данные еще не загружены", records=0, automatic=True),
            "funnel": AdAnalysisSourceStatusOut(id="funnel", label="Воронка продаж", mode="empty", detail="Данные еще не загружены", records=0, automatic=True),
            "finance": AdAnalysisSourceStatusOut(id="finance", label="Финансы WB", mode="empty", detail="Данные еще не загружены", records=0, automatic=True),
            "costs": AdAnalysisSourceStatusOut(id="costs", label="Себестоимость", mode="manual_required", detail="Нужен файл с себестоимостью", records=0, automatic=False),
            "manual_spend": AdAnalysisSourceStatusOut(id="manual_spend", label="Ручное распределение рекламы", mode="empty", detail="Понадобится только если WB не отдал nmID", records=0, automatic=False),
        }

        manual_costs = await self._load_costs(db, store.id)
        manual_spend = await self._load_manual_spend(db, store.id, period_start, period_end)
        manual_finance = await self._load_manual_finance(db, store.id, period_start, period_end)

        if manual_costs:
            source_map["costs"] = AdAnalysisSourceStatusOut(
                id="costs",
                label="Себестоимость",
                mode="manual",
                detail="Используется ваш загруженный файл себестоимости",
                records=len(manual_costs),
                automatic=False,
            )

        if manual_spend:
            source_map["manual_spend"] = AdAnalysisSourceStatusOut(
                id="manual_spend",
                label="Ручное распределение рекламы",
                mode="manual",
                detail="Часть расходов распределена вручную по nmID",
                records=len(manual_spend),
                automatic=False,
            )

        advert_data, source_map["advert"] = await self._fetch_advert_metrics(store, period_start, period_end)
        base_nm_ids = set(advert_data["per_nm"].keys()) | set(manual_spend.keys()) | set(manual_finance.keys())

        finance_data: Dict[int, _FinanceMetrics] = {}
        if base_nm_ids:
            finance_data, source_map["finance"] = await self._fetch_finance_metrics(store, period_start, period_end)
            funnel_data, source_map["funnel"] = await self._fetch_funnel_metrics(store, period_start, period_end, sorted(base_nm_ids))
        else:
            funnel_data = {}
            source_map["funnel"] = AdAnalysisSourceStatusOut(
                id="funnel",
                label="Воронка продаж",
                mode="empty",
                detail="Нет SKU с рекламными расходами за выбранный период",
                records=0,
                automatic=True,
            )
            source_map["finance"] = AdAnalysisSourceStatusOut(
                id="finance",
                label="Финансы WB",
                mode="empty",
                detail="Нет SKU с рекламными расходами за выбранный период",
                records=0,
                automatic=True,
            )

        if manual_finance:
            if source_map["finance"].mode in {"error", "manual_required", "empty"}:
                source_map["finance"] = AdAnalysisSourceStatusOut(
                    id="finance",
                    label="Финансы WB",
                    mode="manual",
                    detail="Используется загруженный вручную финансовый файл",
                    records=len(manual_finance),
                    automatic=False,
                )
            else:
                source_map["finance"] = AdAnalysisSourceStatusOut(
                    id="finance",
                    label="Финансы WB",
                    mode="partial",
                    detail="WB + ручной файл финансов объединены в один расчет",
                    records=max(source_map["finance"].records, len(manual_finance)),
                    automatic=True,
                )

        nm_ids = sorted(base_nm_ids | set(finance_data.keys()))
        cards_map, issue_map = await self._load_cards_and_issues(db, store.id, nm_ids)

        items: List[AdAnalysisItemOut] = []
        missing_cost_nm_ids: List[int] = []

        total_exact = advert_data["exact_spend"]
        total_estimated = advert_data["estimated_spend"]
        total_manual_spend = float(sum(_to_float(row.spend) for row in manual_spend.values()))
        raw_unallocated = advert_data["unallocated_spend"]

        for nm_id in nm_ids:
            advert = advert_data["per_nm"].get(nm_id, _AdvertMetrics())
            manual_topup = manual_spend.get(nm_id)
            if manual_topup:
                advert.manual_spend += manual_topup.spend
                advert.manual_views += manual_topup.views or 0
                advert.manual_clicks += manual_topup.clicks or 0
                advert.manual_orders += manual_topup.orders or 0
                advert.manual_gmv += manual_topup.gmv or 0.0

            finance = finance_data.get(nm_id, _FinanceMetrics())
            if manual_finance.get(nm_id):
                finance = manual_finance[nm_id]

            card = cards_map.get(nm_id)
            issues = issue_map.get(card.id if card else -1, [])
            issue_summary = self._summarize_issues(issues)

            funnel_metric = funnel_data.get(nm_id, _FunnelMetrics())
            total_orders = max(finance.orders, funnel_metric.order_count, advert.total_orders)
            revenue = finance.revenue or manual_finance.get(nm_id, _FinanceMetrics()).revenue or funnel_metric.order_sum or advert.total_gmv
            wb_costs = finance.wb_costs
            unit_cost = manual_costs.get(nm_id)
            missing_cost = unit_cost is None
            if unit_cost is None:
                unit_cost = 0.0
                if advert.total_spend > 0:
                    missing_cost_nm_ids.append(nm_id)

            cost_price = unit_cost * max(total_orders, 0)
            gross_profit_before_ads = revenue - wb_costs - cost_price
            ad_cost = advert.total_spend
            net_profit = gross_profit_before_ads - ad_cost

            order_denominator = max(total_orders, 1)
            ad_orders_denominator = max(advert.total_orders, 1)
            max_cpo = gross_profit_before_ads / order_denominator if total_orders > 0 else gross_profit_before_ads
            actual_cpo = ad_cost / ad_orders_denominator if advert.total_orders > 0 else ad_cost
            profit_delta = max_cpo - actual_cpo

            ctr = (advert.total_clicks / advert.total_views * 100.0) if advert.total_views > 0 else 0.0
            cr = (advert.total_orders / advert.total_clicks * 100.0) if advert.total_clicks > 0 else 0.0
            cpc = (ad_cost / advert.total_clicks) if advert.total_clicks > 0 else 0.0
            drr = (ad_cost / revenue * 100.0) if revenue > 0 else 0.0
            precision, precision_label = self._resolve_precision(advert, raw_unallocated)
            ad_cost_confidence = self._resolve_ad_cost_confidence(advert, raw_unallocated)
            source_lineage = self._build_source_lineage(source_map)

            metrics = AdAnalysisMetricsOut(
                revenue=round(revenue, 2),
                wb_costs=round(wb_costs, 2),
                cost_price=round(cost_price, 2),
                gross_profit_before_ads=round(gross_profit_before_ads, 2),
                ad_cost=round(ad_cost, 2),
                ad_cost_total=round(ad_cost, 2),
                ad_cost_exact=round(advert.exact_spend, 2),
                ad_cost_estimated=round(advert.estimated_spend, 2),
                ad_cost_manual=round(advert.manual_spend, 2),
                ad_cost_source_mode=precision,
                ad_cost_confidence=ad_cost_confidence,
                net_profit=round(net_profit, 2),
                profit_per_order=round(net_profit / order_denominator, 2) if total_orders > 0 else round(net_profit, 2),
                max_cpo=round(max_cpo, 2),
                actual_cpo=round(actual_cpo, 2),
                profit_delta=round(profit_delta, 2),
                views=advert.total_views,
                clicks=advert.total_clicks,
                ad_orders=advert.total_orders,
                ad_gmv=round(advert.total_gmv, 2),
                ctr=round(ctr, 2),
                cr=round(cr, 2),
                open_count=funnel_metric.open_count,
                cart_count=funnel_metric.cart_count,
                order_count=funnel_metric.order_count,
                buyout_count=funnel_metric.buyout_count,
                add_to_cart_percent=round(funnel_metric.add_to_cart_percent, 2),
                cart_to_order_percent=round(funnel_metric.cart_to_order_percent, 2),
                cpc=round(cpc, 2),
                drr=round(drr, 2),
            )

            diagnosis, diagnosis_label, reasons, hints = self._resolve_diagnosis(
                metrics=metrics,
                issue_summary=issue_summary,
                missing_cost=missing_cost,
                finance_ready=(source_map["finance"].mode not in {"error", "manual_required", "empty"} or bool(manual_finance)),
            )
            status, status_label = self._resolve_status(
                metrics=metrics,
                diagnosis=diagnosis,
                missing_cost=missing_cost,
                finance_ready=(source_map["finance"].mode not in {"error", "manual_required", "empty"} or bool(manual_finance)),
            )
            priority, priority_label = self._resolve_priority(
                status=status,
                diagnosis=diagnosis,
                metrics=metrics,
                trend_signal="no_history",
            )
            action_title, action_description, steps, insights, risk_flags = self._build_actions(
                status=status,
                diagnosis=diagnosis,
                issue_summary=issue_summary,
                missing_cost=missing_cost,
                precision=precision,
                source_statuses=source_map,
                metrics=metrics,
            )

            title = None
            if card and card.title:
                title = card.title
            elif manual_topup and manual_topup.title:
                title = manual_topup.title
            elif advert_data["titles"].get(nm_id):
                title = advert_data["titles"].get(nm_id)
            photo_url = card.photos[0] if card and getattr(card, "photos", None) else None

            items.append(
                AdAnalysisItemOut(
                    nm_id=nm_id,
                    card_id=card.id if card else None,
                    title=title,
                    vendor_code=card.vendor_code if card else None,
                    photo_url=photo_url,
                    wb_link=f"https://www.wildberries.ru/catalog/{nm_id}/detail.aspx",
                    workspace_link=f"/workspace/cards/{card.id}" if card else None,
                    price=card.price if card else None,
                    card_score=card.score if card else None,
                    status=status,
                    status_label=status_label,
                    diagnosis=diagnosis,
                    diagnosis_label=diagnosis_label,
                    status_reason=reasons,
                    status_hint=hints,
                    action_title=action_title,
                    action_description=action_description,
                    priority=priority,
                    priority_label=priority_label,
                    precision=precision,
                    precision_label=precision_label,
                    source_lineage=source_lineage,
                    trend=AdAnalysisTrendOut(),
                    issue_summary=AdAnalysisIssueSummaryOut(
                        total=issue_summary.total,
                        critical=issue_summary.critical,
                        warnings=issue_summary.warnings,
                        photos=issue_summary.photos,
                        price=issue_summary.price,
                        text=issue_summary.text,
                        docs=issue_summary.docs,
                        top_titles=issue_summary.top_titles,
                    ),
                    metrics=metrics,
                    spend_sources={
                        "exact": round(advert.exact_spend, 2),
                        "estimated": round(advert.estimated_spend, 2),
                        "manual": round(advert.manual_spend, 2),
                    },
                    insights=insights,
                    steps=steps,
                    risk_flags=risk_flags,
                )
            )

        if items:
            if manual_costs and missing_cost_nm_ids:
                source_map["costs"] = AdAnalysisSourceStatusOut(
                    id="costs",
                    label="Себестоимость",
                    mode="partial",
                    detail="Часть SKU покрыта файлом себестоимости, но не все позиции",
                    records=len(manual_costs),
                    automatic=False,
                )
            elif not manual_costs:
                source_map["costs"] = AdAnalysisSourceStatusOut(
                    id="costs",
                    label="Себестоимость",
                    mode="manual_required",
                    detail="Для точной экономики нужен файл с себестоимостью",
                    records=0,
                    automatic=False,
                )

        items.sort(key=self._item_sort_key)

        remaining_unallocated = max(raw_unallocated - total_manual_spend, 0.0)
        alerts = self._build_alerts(
            remaining_unallocated=remaining_unallocated,
            missing_cost_nm_ids=missing_cost_nm_ids,
            finance_status=source_map["finance"],
            advert_status=source_map["advert"],
            items=items,
        )
        budget_moves = self._build_budget_moves(items)

        status_counts = {status: 0 for status in ("stop", "rescue", "control", "grow", "low_data")}
        profitable_count = 0
        problematic_count = 0
        loss_count = 0
        total_revenue = 0.0
        total_ad_spend = 0.0
        total_net_profit = 0.0

        for item in items:
            status_counts[item.status] = status_counts.get(item.status, 0) + 1
            total_revenue += item.metrics.revenue
            total_ad_spend += item.metrics.ad_cost
            total_net_profit += item.metrics.net_profit
            if item.status in {"grow", "control"}:
                profitable_count += 1
            elif item.status == "stop":
                loss_count += 1
            else:
                problematic_count += 1

        upload_needs = AdAnalysisUploadNeedsOut(
            period_start=period_start,
            period_end=period_end,
            missing_costs_count=len(sorted(set(missing_cost_nm_ids))),
            missing_cost_nm_ids=sorted(set(missing_cost_nm_ids))[:20],
            needs_manual_spend=remaining_unallocated > 0.01 or source_map["advert"].mode in {"error", "manual_required"},
            needs_manual_finance=source_map["finance"].mode in {"error", "manual_required"},
            can_upload_costs=True,
            can_upload_manual_spend=True,
            can_upload_manual_finance=True,
        )

        overview = AdAnalysisOverviewOut(
            store_id=int(store.id),
            generated_at=utc_now(),
            snapshot_ready=True,
            period_start=period_start,
            period_end=period_end,
            total_skus=len(items),
            total_revenue=round(total_revenue, 2),
            total_ad_spend=round(total_ad_spend, 2),
            total_net_profit=round(total_net_profit, 2),
            exact_spend=round(total_exact, 2),
            estimated_spend=round(total_estimated, 2),
            manual_spend=round(total_manual_spend, 2),
            unallocated_spend=round(remaining_unallocated, 2),
            profitable_count=profitable_count,
            problematic_count=problematic_count,
            loss_count=loss_count,
            status_counts=status_counts,
            source_statuses=list(source_map.values()),
            source_lineage=self._build_source_lineage(source_map),
            alerts=alerts,
            budget_moves=budget_moves,
            campaigns=list(advert_data.get("campaigns", []))[:12],
            upload_needs=upload_needs,
            items=items,
        )

        self._cache[cache_key] = (time.time(), overview)
        return overview

    async def upload_costs(
        self,
        db: AsyncSession,
        *,
        store_id: int,
        file_name: str,
        content: bytes,
    ) -> _UploadResult:
        headers, rows = _read_tabular_rows(file_name, content)
        mapping = _match_columns(headers, self.COST_ALIASES)
        if (not mapping.get("nm_id") and not mapping.get("vendor_code")) or not mapping.get("unit_cost"):
            raise ValueError("В файле нужны nm_id или vendor_code, а также колонка себестоимости")

        existing = await db.execute(
            select(SkuEconomicsCost).where(SkuEconomicsCost.store_id == int(store_id))
        )
        existing_map = {int(row.nm_id): row for row in existing.scalars().all()}
        vendor_code_map = await self._load_vendor_code_map(db, store_id)
        imported = 0
        updated = 0
        notes: List[str] = []
        resolved_by_vendor_code = 0
        unresolved_rows = 0
        unresolved_preview: List[AdAnalysisUploadUnresolvedRowOut] = []

        for index, raw in enumerate(rows, start=2):
            resolved = self._resolve_upload_identifier(raw, mapping, vendor_code_map)
            nm_id = resolved.nm_id
            unit_cost = _to_float(raw.get(mapping["unit_cost"]))
            if nm_id <= 0 or unit_cost <= 0:
                if unit_cost > 0:
                    unresolved_rows += 1
                    if len(unresolved_preview) < 5:
                        unresolved_preview.append(self._build_unresolved_row(index, raw, mapping))
                continue
            if resolved.vendor_code and not _to_int(raw.get(mapping.get("nm_id") or "")):
                resolved_by_vendor_code += 1
            row = existing_map.get(nm_id)
            if row is None:
                row = SkuEconomicsCost(
                    store_id=int(store_id),
                    nm_id=nm_id,
                )
                db.add(row)
                existing_map[nm_id] = row
                imported += 1
            else:
                updated += 1
            row.unit_cost = unit_cost
            row.title = str(raw.get(mapping.get("title") or "") or "").strip() or row.title
            row.vendor_code = str(raw.get(mapping.get("vendor_code") or "") or "").strip() or row.vendor_code

        if imported == 0 and updated == 0:
            notes.append("Ни одной валидной строки не найдено")
        if resolved_by_vendor_code:
            notes.append(f"{resolved_by_vendor_code} строк сопоставлены по vendor_code")
        if unresolved_rows:
            notes.append(f"{unresolved_rows} строк не удалось сопоставить. Добавьте nm_id или проверьте vendor_code")

        if imported > 0 or updated > 0:
            await self.invalidate_saved_overviews(db, store_id)
        await db.commit()
        return _UploadResult(
            imported=imported,
            updated=updated,
            notes=notes,
            detected_headers=[header for header in headers if header],
            matched_fields=mapping,
            resolved_by_vendor_code=resolved_by_vendor_code,
            unresolved_count=unresolved_rows,
            unresolved_preview=unresolved_preview,
        )

    async def upload_manual_spend(
        self,
        db: AsyncSession,
        *,
        store_id: int,
        file_name: str,
        content: bytes,
        period_start: date,
        period_end: date,
    ) -> _UploadResult:
        headers, rows = _read_tabular_rows(file_name, content)
        mapping = _match_columns(headers, self.SPEND_ALIASES)
        if (not mapping.get("nm_id") and not mapping.get("vendor_code")) or not mapping.get("spend"):
            raise ValueError("В файле нужны nm_id или vendor_code, а также колонка расход")

        existing = await db.execute(
            select(SkuEconomicsManualSpend).where(
                SkuEconomicsManualSpend.store_id == int(store_id),
                SkuEconomicsManualSpend.period_start == period_start,
                SkuEconomicsManualSpend.period_end == period_end,
            )
        )
        existing_map = {int(row.nm_id): row for row in existing.scalars().all()}
        vendor_code_map = await self._load_vendor_code_map(db, store_id)
        imported = 0
        updated = 0
        notes: List[str] = []
        resolved_by_vendor_code = 0
        unresolved_rows = 0
        unresolved_preview: List[AdAnalysisUploadUnresolvedRowOut] = []

        for index, raw in enumerate(rows, start=2):
            resolved = self._resolve_upload_identifier(raw, mapping, vendor_code_map)
            nm_id = resolved.nm_id
            spend = _to_float(raw.get(mapping["spend"]))
            if nm_id <= 0 or spend <= 0:
                if spend > 0:
                    unresolved_rows += 1
                    if len(unresolved_preview) < 5:
                        unresolved_preview.append(self._build_unresolved_row(index, raw, mapping))
                continue
            if resolved.vendor_code and not _to_int(raw.get(mapping.get("nm_id") or "")):
                resolved_by_vendor_code += 1
            row = existing_map.get(nm_id)
            if row is None:
                row = SkuEconomicsManualSpend(
                    store_id=int(store_id),
                    nm_id=nm_id,
                    period_start=period_start,
                    period_end=period_end,
                )
                db.add(row)
                existing_map[nm_id] = row
                imported += 1
            else:
                updated += 1
            row.title = str(raw.get(mapping.get("title") or "") or "").strip() or row.title
            row.spend = spend
            row.views = _to_int(raw.get(mapping.get("views") or ""))
            row.clicks = _to_int(raw.get(mapping.get("clicks") or ""))
            row.orders = _to_int(raw.get(mapping.get("orders") or ""))
            row.gmv = _to_float(raw.get(mapping.get("gmv") or ""))
            row.source_file_name = file_name

        if imported == 0 and updated == 0:
            notes.append("Ни одной валидной строки не найдено")
        if resolved_by_vendor_code:
            notes.append(f"{resolved_by_vendor_code} строк сопоставлены по vendor_code")
        if unresolved_rows:
            notes.append(f"{unresolved_rows} строк не удалось сопоставить. Добавьте nm_id или проверьте vendor_code")

        if imported > 0 or updated > 0:
            await self.invalidate_saved_overviews(db, store_id)
        await db.commit()
        return _UploadResult(
            imported=imported,
            updated=updated,
            notes=notes,
            detected_headers=[header for header in headers if header],
            matched_fields=mapping,
            resolved_by_vendor_code=resolved_by_vendor_code,
            unresolved_count=unresolved_rows,
            unresolved_preview=unresolved_preview,
        )

    async def upload_manual_finance(
        self,
        db: AsyncSession,
        *,
        store_id: int,
        file_name: str,
        content: bytes,
        period_start: date,
        period_end: date,
    ) -> _UploadResult:
        headers, rows = _read_tabular_rows(file_name, content)
        notes: List[str] = []
        matched_fields: Dict[str, str] = {}
        resolved_by_vendor_code = 0
        unresolved_rows = 0
        unresolved_preview: List[AdAnalysisUploadUnresolvedRowOut] = []

        wb_export_mode = _match_columns(headers, {
            "nm_id": ["nm_id"],
            "doc_type_name": ["doc_type_name", "supplier_oper_name"],
            "retail_price_withdisc_rub": ["retail_price_withdisc_rub"],
            "ppvz_for_pay": ["ppvz_for_pay"],
        })

        aggregate: Dict[int, _FinanceMetrics] = defaultdict(_FinanceMetrics)
        titles: Dict[int, str] = {}

        if wb_export_mode.get("nm_id") and wb_export_mode.get("retail_price_withdisc_rub") and wb_export_mode.get("ppvz_for_pay"):
            for raw in rows:
                nm_id = _to_int(raw.get(wb_export_mode["nm_id"]))
                if nm_id <= 0:
                    continue
                doc_type = str(raw.get(wb_export_mode.get("doc_type_name") or "") or "").lower()
                negative = "возврат" in doc_type or "return" in doc_type or _to_float(raw.get("quantity")) < 0
                sign = -1.0 if negative else 1.0
                revenue = _to_float(raw.get("retail_price_withdisc_rub") or raw.get("retail_amount"))
                payout = _to_float(raw.get("ppvz_for_pay"))
                extra = sum(
                    _to_float(raw.get(field))
                    for field in (
                        "delivery_rub",
                        "acquiring_fee",
                        "penalty",
                        "storage_fee",
                        "deduction",
                        "acceptance",
                        "rebill_logistic_cost",
                        "additional_payment",
                    )
                )
                metric = aggregate[nm_id]
                metric.revenue += revenue * sign if revenue >= 0 else revenue
                metric.payout += payout * sign if payout >= 0 else payout
                metric.wb_costs += _safe_positive(extra)
                metric.orders += max(_to_int(raw.get("quantity")) * int(sign), 0)
        else:
            mapping = _match_columns(headers, self.FINANCE_ALIASES)
            matched_fields = mapping
            if (not mapping.get("nm_id") and not mapping.get("vendor_code")) or not mapping.get("revenue"):
                raise ValueError("В файле нужны nm_id или vendor_code, а также выручка. Либо загрузите raw экспорт отчета реализации WB")
            vendor_code_map = await self._load_vendor_code_map(db, store_id)
            for index, raw in enumerate(rows, start=2):
                resolved = self._resolve_upload_identifier(raw, mapping, vendor_code_map)
                nm_id = resolved.nm_id
                if nm_id <= 0:
                    if _to_float(raw.get(mapping["revenue"])) > 0:
                        unresolved_rows += 1
                        if len(unresolved_preview) < 5:
                            unresolved_preview.append(self._build_unresolved_row(index, raw, mapping))
                    continue
                if resolved.vendor_code and not _to_int(raw.get(mapping.get("nm_id") or "")):
                    resolved_by_vendor_code += 1
                metric = aggregate[nm_id]
                metric.revenue += _to_float(raw.get(mapping["revenue"]))
                metric.wb_costs += _to_float(raw.get(mapping.get("wb_costs") or ""))
                metric.payout += _to_float(raw.get(mapping.get("payout") or ""))
                metric.orders += _to_int(raw.get(mapping.get("orders") or ""))
                titles[nm_id] = str(raw.get(mapping.get("title") or "") or "").strip() or titles.get(nm_id, "")
            if resolved_by_vendor_code:
                notes.append(f"{resolved_by_vendor_code} строк сопоставлены по vendor_code")
            if unresolved_rows:
                notes.append(f"{unresolved_rows} строк не удалось сопоставить. Добавьте nm_id или проверьте vendor_code")

        existing = await db.execute(
            select(SkuEconomicsManualFinance).where(
                SkuEconomicsManualFinance.store_id == int(store_id),
                SkuEconomicsManualFinance.period_start == period_start,
                SkuEconomicsManualFinance.period_end == period_end,
            )
        )
        existing_map = {int(row.nm_id): row for row in existing.scalars().all()}
        imported = 0
        updated = 0

        for nm_id, metric in aggregate.items():
            row = existing_map.get(nm_id)
            if row is None:
                row = SkuEconomicsManualFinance(
                    store_id=int(store_id),
                    nm_id=nm_id,
                    period_start=period_start,
                    period_end=period_end,
                )
                db.add(row)
                existing_map[nm_id] = row
                imported += 1
            else:
                updated += 1
            revenue = metric.revenue
            payout = metric.payout
            base_costs = max(revenue - payout, 0.0)
            row.revenue = round(max(revenue, 0.0), 2)
            row.wb_costs = round(max(metric.wb_costs, 0.0) + base_costs, 2)
            row.payout = round(payout, 2)
            row.orders = max(int(metric.orders), 0)
            row.title = titles.get(nm_id) or row.title
            row.source_file_name = file_name

        if imported == 0 and updated == 0:
            notes.append("Ни одной валидной строки не найдено")

        if imported > 0 or updated > 0:
            await self.invalidate_saved_overviews(db, store_id)
        await db.commit()
        return _UploadResult(
            imported=imported,
            updated=updated,
            notes=notes,
            detected_headers=[header for header in headers if header],
            matched_fields=matched_fields,
            resolved_by_vendor_code=resolved_by_vendor_code,
            unresolved_count=unresolved_rows,
            unresolved_preview=unresolved_preview,
        )

    async def _load_costs(self, db: AsyncSession, store_id: int) -> Dict[int, float]:
        result = await db.execute(
            select(SkuEconomicsCost).where(SkuEconomicsCost.store_id == int(store_id))
        )
        return {int(row.nm_id): float(row.unit_cost or 0.0) for row in result.scalars().all()}

    async def _load_manual_spend(
        self,
        db: AsyncSession,
        store_id: int,
        period_start: date,
        period_end: date,
    ) -> Dict[int, SkuEconomicsManualSpend]:
        result = await db.execute(
            select(SkuEconomicsManualSpend).where(
                SkuEconomicsManualSpend.store_id == int(store_id),
                SkuEconomicsManualSpend.period_start == period_start,
                SkuEconomicsManualSpend.period_end == period_end,
            )
        )
        return {int(row.nm_id): row for row in result.scalars().all()}

    async def _load_manual_finance(
        self,
        db: AsyncSession,
        store_id: int,
        period_start: date,
        period_end: date,
    ) -> Dict[int, _FinanceMetrics]:
        result = await db.execute(
            select(SkuEconomicsManualFinance).where(
                SkuEconomicsManualFinance.store_id == int(store_id),
                SkuEconomicsManualFinance.period_start == period_start,
                SkuEconomicsManualFinance.period_end == period_end,
            )
        )
        out: Dict[int, _FinanceMetrics] = {}
        for row in result.scalars().all():
            out[int(row.nm_id)] = _FinanceMetrics(
                revenue=float(row.revenue or 0.0),
                payout=float(row.payout or 0.0),
                wb_costs=float(row.wb_costs or 0.0),
                orders=int(row.orders or 0),
            )
        return out

    async def _load_vendor_code_map(self, db: AsyncSession, store_id: int) -> Dict[str, int]:
        result = await db.execute(
            select(Card.vendor_code, Card.nm_id).where(Card.store_id == int(store_id))
        )
        out: Dict[str, int] = {}
        for vendor_code, nm_id in result.all():
            normalized = _normalize_vendor_code(vendor_code)
            if normalized and int(nm_id or 0) > 0 and normalized not in out:
                out[normalized] = int(nm_id)
        return out

    def _resolve_upload_identifier(
        self,
        raw: Dict[str, Any],
        mapping: Dict[str, str],
        vendor_code_map: Dict[str, int],
    ) -> _UploadResolution:
        nm_id = _to_int(raw.get(mapping.get("nm_id") or ""))
        vendor_code = str(raw.get(mapping.get("vendor_code") or "") or "").strip()
        if nm_id > 0:
            return _UploadResolution(nm_id=nm_id, vendor_code=vendor_code)
        normalized_vendor_code = _normalize_vendor_code(vendor_code)
        matched_nm_id = int(vendor_code_map.get(normalized_vendor_code) or 0)
        return _UploadResolution(nm_id=matched_nm_id, vendor_code=vendor_code)

    def _build_unresolved_row(
        self,
        row_number: int,
        raw: Dict[str, Any],
        mapping: Dict[str, str],
    ) -> AdAnalysisUploadUnresolvedRowOut:
        return AdAnalysisUploadUnresolvedRowOut(
            row_number=row_number,
            raw_nm_id=str(raw.get(mapping.get("nm_id") or "") or "").strip() or None,
            raw_vendor_code=str(raw.get(mapping.get("vendor_code") or "") or "").strip() or None,
            raw_title=str(raw.get(mapping.get("title") or "") or "").strip() or None,
        )

    async def _load_cards_and_issues(
        self,
        db: AsyncSession,
        store_id: int,
        nm_ids: Sequence[int],
    ) -> Tuple[Dict[int, Card], Dict[int, List[CardIssue]]]:
        if not nm_ids:
            return {}, {}
        cards_result = await db.execute(
            select(Card).where(Card.store_id == int(store_id), Card.nm_id.in_(list(nm_ids)))
        )
        cards = list(cards_result.scalars().all())
        cards_map = {int(card.nm_id): card for card in cards}
        card_ids = [card.id for card in cards]
        if not card_ids:
            return cards_map, {}

        issues_result = await db.execute(
            select(CardIssue).where(
                CardIssue.card_id.in_(card_ids),
                CardIssue.status.in_(list(_UNRESOLVED_ISSUE_STATUSES)),
            )
        )
        issue_map: Dict[int, List[CardIssue]] = defaultdict(list)
        for issue in issues_result.scalars().all():
            issue_map[int(issue.card_id)].append(issue)
        return cards_map, issue_map

    async def _fetch_advert_metrics(
        self,
        store: Store,
        period_start: date,
        period_end: date,
    ) -> Tuple[Dict[str, Any], AdAnalysisSourceStatusOut]:
        tokens = self._candidate_tokens(store, prefer_advert=True)
        last_error = "Не найден токен для рекламного API"
        for token in tokens:
            try:
                repo = WBAdvertRepository(token=token)
                campaign_ids = await repo.get_campaign_ids()
                if not campaign_ids:
                    return (
                        {"per_nm": {}, "titles": {}, "campaigns": [], "exact_spend": 0.0, "estimated_spend": 0.0, "manual_spend": 0.0, "unallocated_spend": 0.0},
                        AdAnalysisSourceStatusOut(
                            id="advert",
                            label="Реклама WB",
                            mode="empty",
                            detail="Кампании за выбранный период не найдены",
                            records=0,
                            automatic=True,
                        ),
                    )

                raw_items: List[dict] = []
                for batch in _chunks(campaign_ids, 50):
                    batch_data = await asyncio.to_thread(
                        repo.get_fullstats,
                        batch,
                        begin_date=period_start.isoformat(),
                        end_date=period_end.isoformat(),
                    )
                    raw_items.extend(self._extract_fullstats_items(batch_data))

                parsed = self._parse_fullstats(raw_items)
                mode = "partial" if parsed["unallocated_spend"] > 0.01 else "ok"
                detail = (
                    "Часть кампаний не удалось точно привязать к nmID"
                    if mode == "partial"
                    else "Расходы, клики и заказы подтянуты из WB Advert"
                )
                return (
                    parsed,
                    AdAnalysisSourceStatusOut(
                        id="advert",
                        label="Реклама WB",
                        mode=mode,
                        detail=detail,
                        records=len(raw_items),
                        automatic=True,
                    ),
                )
            except Exception as exc:
                last_error = str(exc)

        return (
            {"per_nm": {}, "titles": {}, "campaigns": [], "exact_spend": 0.0, "estimated_spend": 0.0, "manual_spend": 0.0, "unallocated_spend": 0.0},
            AdAnalysisSourceStatusOut(
                id="advert",
                label="Реклама WB",
                mode="error",
                detail=last_error,
                records=0,
                automatic=True,
            ),
        )

    def _parse_fullstats(self, items: Sequence[dict]) -> Dict[str, Any]:
        per_nm: Dict[int, _AdvertMetrics] = defaultdict(_AdvertMetrics)
        titles: Dict[int, str] = {}
        campaigns: List[AdAnalysisCampaignOut] = []
        exact_spend = 0.0
        estimated_spend = 0.0
        unallocated_spend = 0.0
        skipped_items = 0
        invalid_nm_rows = 0
        zero_total_drift_items = 0

        for item in items:
            if not isinstance(item, dict):
                skipped_items += 1
                continue
            leaf_nms: List[dict] = []
            self._collect_leaf_nms(item, leaf_nms)
            campaign_total = self._extract_campaign_total(item)
            campaign_spend_from_nms = 0.0
            campaign_orders = 0
            campaign_gmv = 0.0
            linked_skus = 0
            valid_nm_rows: List[Tuple[int, Dict[str, Any]]] = []

            for nm_row in leaf_nms:
                if not isinstance(nm_row, dict):
                    invalid_nm_rows += 1
                    continue
                nm_id = self._extract_nm_id(nm_row)
                if nm_id <= 0:
                    invalid_nm_rows += 1
                    continue
                valid_nm_rows.append((nm_id, nm_row))
                linked_skus += 1
                metric = per_nm[nm_id]
                nm_spend = self._extract_nm_metric_float(nm_row, self.ADVERT_NM_SPEND_KEYS)
                nm_orders = _to_int(self._extract_nm_metric_float(nm_row, self.ADVERT_NM_ORDERS_KEYS))
                nm_gmv = self._extract_nm_metric_float(nm_row, self.ADVERT_NM_GMV_KEYS)
                metric.views += _to_int(self._extract_nm_metric_float(nm_row, self.ADVERT_NM_VIEWS_KEYS))
                metric.clicks += _to_int(self._extract_nm_metric_float(nm_row, self.ADVERT_NM_CLICKS_KEYS))
                metric.orders += nm_orders
                metric.gmv += nm_gmv
                metric.exact_spend += nm_spend
                title = self._extract_nm_title(nm_row)
                titles[nm_id] = title or titles.get(nm_id, "")
                exact_spend += nm_spend
                campaign_spend_from_nms += nm_spend
                campaign_orders += nm_orders
                campaign_gmv += nm_gmv

            if campaign_total <= self.ADVERT_RESIDUAL_EPSILON and campaign_spend_from_nms > self.ADVERT_RESIDUAL_EPSILON:
                zero_total_drift_items += 1
            effective_campaign_total = max(campaign_total, campaign_spend_from_nms)
            residual = max(effective_campaign_total - campaign_spend_from_nms, 0.0)
            if residual <= self.ADVERT_RESIDUAL_EPSILON:
                residual = 0.0
            else:
                weights = self._build_residual_weights(valid_nm_rows)
                if weights:
                    total_weight = sum(weight for _, weight in weights) or float(len(weights) or 1)
                    for nm_id, weight in weights:
                        share = residual * (weight / total_weight)
                        per_nm[nm_id].estimated_spend += share
                        estimated_spend += share
                else:
                    unallocated_spend += residual

            advert_id = _to_int(item.get("advertId") or item.get("advert_id") or item.get("id"))
            campaign_title = (
                str(item.get("advertName") or item.get("advert_name") or item.get("name") or item.get("advertTitle") or "").strip()
                or (f"РК {advert_id}" if advert_id > 0 else "Рекламная кампания")
            )
            if residual > self.ADVERT_RESIDUAL_EPSILON and linked_skus == 0:
                precision = "unallocated"
                precision_label = "Нераспределенные расходы"
            elif residual > self.ADVERT_RESIDUAL_EPSILON:
                precision = "estimated" if campaign_spend_from_nms <= self.ADVERT_RESIDUAL_EPSILON else "mixed"
                precision_label = "Оценка" if precision == "estimated" else "Смешанный источник"
            else:
                precision = "exact"
                precision_label = "Точные данные"

            campaigns.append(
                AdAnalysisCampaignOut(
                    advert_id=advert_id or None,
                    title=campaign_title,
                    ad_cost=round(effective_campaign_total, 2),
                    ad_gmv=round(campaign_gmv, 2),
                    drr=round((effective_campaign_total / campaign_gmv * 100.0) if campaign_gmv > 0 else 0.0, 2),
                    linked_skus=linked_skus,
                    precision=precision,
                    precision_label=precision_label,
                )
            )

        if skipped_items > 0 or invalid_nm_rows > 0 or zero_total_drift_items > 0:
            logger.warning(
                "[sku-advert] fullstats parser normalized anomalies skipped_items=%s invalid_nm_rows=%s "
                "zero_total_drift_items=%s",
                skipped_items,
                invalid_nm_rows,
                zero_total_drift_items,
            )

        return {
            "per_nm": per_nm,
            "titles": titles,
            "campaigns": sorted(campaigns, key=lambda row: (-row.ad_cost, row.title)),
            "exact_spend": exact_spend,
            "estimated_spend": estimated_spend,
            "manual_spend": 0.0,
            "unallocated_spend": unallocated_spend,
        }

    def _collect_leaf_nms(self, node: Any, out: List[dict]) -> None:
        if not isinstance(node, dict):
            return
        nested = []
        for key in ("days", "apps"):
            values = node.get(key)
            if isinstance(values, list):
                nested.extend(item for item in values if isinstance(item, dict))
        if nested:
            for child in nested:
                self._collect_leaf_nms(child, out)

        for key in ("nms", "products", "items"):
            values = node.get(key)
            if isinstance(values, list):
                out.extend(item for item in values if isinstance(item, dict))

        nm_single = node.get("nm")
        if isinstance(nm_single, dict):
            out.append(nm_single)

    async def _fetch_funnel_metrics(
        self,
        store: Store,
        period_start: date,
        period_end: date,
        nm_ids: Sequence[int],
    ) -> Tuple[Dict[int, _FunnelMetrics], AdAnalysisSourceStatusOut]:
        if not nm_ids:
            return {}, AdAnalysisSourceStatusOut(
                id="funnel",
                label="Воронка продаж",
                mode="empty",
                detail="Нет SKU для анализа",
                records=0,
                automatic=True,
            )

        tokens = self._candidate_tokens(store, prefer_advert=False)
        last_error = "Не найден токен аналитики"
        url = f"{settings.WB_ANALYTICS_API_URL}/api/analytics/v3/sales-funnel/products"
        for token in tokens:
            try:
                out: Dict[int, _FunnelMetrics] = {}
                async with httpx.AsyncClient(timeout=httpx.Timeout(45.0)) as client:
                    for batch in _chunks(list(nm_ids), 500):
                        payload = {
                            "selectedPeriod": {
                                "start": period_start.isoformat(),
                                "end": period_end.isoformat(),
                            },
                            "nmIds": batch,
                            "limit": len(batch),
                            "offset": 0,
                            "skipDeletedNm": False,
                            "orderBy": {"field": "openCard", "mode": "desc"},
                        }
                        resp = await self._request_with_retry(
                            client,
                            "POST",
                            url,
                            headers={
                                "Authorization": token,
                                "Content-Type": "application/json",
                                "Accept": "application/json",
                            },
                            json=payload,
                        )
                        if resp.status_code >= 400:
                            raise ValueError(f"WB funnel error {resp.status_code}: {resp.text[:300]}")
                        data = resp.json() or {}
                        products = ((data.get("data") or {}).get("products") or [])
                        for row in products:
                            product = row.get("product") or {}
                            stats = ((row.get("statistic") or {}).get("selected") or {})
                            nm_id = _to_int(product.get("nmId"))
                            if nm_id <= 0:
                                continue
                            conversions = stats.get("conversions") or {}
                            out[nm_id] = _FunnelMetrics(
                                open_count=_to_int(stats.get("openCount")),
                                cart_count=_to_int(stats.get("cartCount")),
                                order_count=_to_int(stats.get("orderCount")),
                                order_sum=_to_float(stats.get("orderSum")),
                                buyout_count=_to_int(stats.get("buyoutCount")),
                                buyout_sum=_to_float(stats.get("buyoutSum")),
                                add_to_cart_percent=_to_float(conversions.get("addToCartPercent")),
                                cart_to_order_percent=_to_float(conversions.get("cartToOrderPercent")),
                            )
                return (
                    out,
                    AdAnalysisSourceStatusOut(
                        id="funnel",
                        label="Воронка продаж",
                        mode="ok",
                        detail="Открытия карточки, корзина и заказы подтянуты из WB Analytics",
                        records=len(out),
                        automatic=True,
                    ),
                )
            except Exception as exc:
                last_error = str(exc)

        return (
            {},
            AdAnalysisSourceStatusOut(
                id="funnel",
                label="Воронка продаж",
                mode="error",
                detail=last_error,
                records=0,
                automatic=True,
            ),
        )

    async def _fetch_finance_metrics(
        self,
        store: Store,
        period_start: date,
        period_end: date,
    ) -> Tuple[Dict[int, _FinanceMetrics], AdAnalysisSourceStatusOut]:
        tokens = self._candidate_tokens(store, prefer_advert=False)
        last_error = "Не найден токен статистики"
        url = f"{settings.WB_STATISTICS_API_URL}/api/v5/supplier/reportDetailByPeriod"
        for token in tokens:
            try:
                partial_reason: Optional[str] = None
                async with httpx.AsyncClient(timeout=httpx.Timeout(90.0)) as client:
                    try:
                        rows = await self._fetch_finance_rows_paginated(
                            client=client,
                            url=url,
                            token=token,
                            date_from=period_start,
                            date_to=period_end,
                            period="daily",
                        )
                    except _FinanceFetchError as finance_exc:
                        rows = list(finance_exc.partial_rows or [])
                        if not rows:
                            raise
                        partial_reason = str(finance_exc)
                        logger.warning(
                            "[sku-finance] partial finance metrics fetched period=%s..%s rows=%s reason=%s",
                            period_start.isoformat(),
                            period_end.isoformat(),
                            len(rows),
                            partial_reason,
                        )
                if not rows:
                    return (
                        {},
                        AdAnalysisSourceStatusOut(
                            id="finance",
                            label="Финансы WB",
                            mode="empty",
                            detail="За период нет строк в отчете реализации",
                            records=0,
                            automatic=True,
                        ),
                    )
                out: Dict[int, _FinanceMetrics] = defaultdict(_FinanceMetrics)
                for row in rows:
                    nm_id = _to_int(row.get("nm_id"))
                    if nm_id <= 0:
                        continue
                    quantity = _to_int(row.get("quantity"))
                    doc_type = str(row.get("doc_type_name") or row.get("supplier_oper_name") or "").lower()
                    negative = "возврат" in doc_type or "return" in doc_type or quantity < 0
                    sign = -1.0 if negative else 1.0
                    revenue = _to_float(row.get("retail_price_withdisc_rub") or row.get("retail_amount"))
                    payout = _to_float(row.get("ppvz_for_pay"))
                    extra = sum(
                        _to_float(row.get(field))
                        for field in (
                            "delivery_rub",
                            "acquiring_fee",
                            "penalty",
                            "storage_fee",
                            "deduction",
                            "acceptance",
                            "rebill_logistic_cost",
                            "additional_payment",
                        )
                    )
                    metric = out[nm_id]
                    metric.revenue += revenue * sign if revenue >= 0 else revenue
                    metric.payout += payout * sign if payout >= 0 else payout
                    metric.wb_costs += _safe_positive(extra)
                    metric.orders += max(quantity * int(sign), 0)

                for metric in out.values():
                    metric.wb_costs = round(max(metric.wb_costs, 0.0) + max(metric.revenue - metric.payout, 0.0), 2)
                    metric.revenue = round(max(metric.revenue, 0.0), 2)
                    metric.payout = round(metric.payout, 2)
                    metric.orders = max(int(metric.orders), 0)

                return (
                    out,
                    AdAnalysisSourceStatusOut(
                        id="finance",
                        label="Финансы WB",
                        mode="partial" if partial_reason else "ok",
                        detail=(
                            f"Финансовый отчет WB подтянут частично: {partial_reason}"
                            if partial_reason
                            else "Финансовый отчет WB подтянут по API (с постраничной загрузкой)."
                        ),
                        records=len(rows),
                        automatic=True,
                    ),
                )
            except Exception as exc:
                last_error = str(exc)

        return (
            {},
            AdAnalysisSourceStatusOut(
                id="finance",
                label="Финансы WB",
                mode="manual_required",
                detail=last_error,
                records=0,
                automatic=True,
            ),
        )

    def _candidate_tokens(self, store: Store, *, prefer_advert: bool) -> List[str]:
        feature_token = str(get_store_feature_api_key(store, "ad_analysis") or "").strip()
        preferred = [
            feature_token,
            settings.WB_ADVERT_API_KEY if prefer_advert else None,
            store.api_key,
            settings.WB_API_KEY,
            settings.WB_ADVERT_API_KEY,
        ]
        out: List[str] = []
        seen: set[str] = set()
        for token in preferred:
            candidate = str(token or "").strip()
            if candidate and candidate not in seen:
                out.append(candidate)
                seen.add(candidate)
        return out

    def _summarize_issues(self, issues: Sequence[CardIssue]) -> _IssueSnapshot:
        top_titles: List[str] = []
        snapshot = _IssueSnapshot()
        sorted_issues = sorted(
            issues,
            key=lambda issue: (
                0 if str(issue.severity) == IssueSeverity.CRITICAL.value else 1,
                -(int(issue.score_impact or 0)),
                int(issue.id),
            ),
        )
        for issue in sorted_issues:
            snapshot.total += 1
            severity = getattr(issue.severity, "value", issue.severity)
            category = getattr(issue.category, "value", issue.category)
            field_path = str(issue.field_path or "").lower()
            title = str(issue.title or "").strip()
            if severity == IssueSeverity.CRITICAL.value:
                snapshot.critical += 1
            elif severity == IssueSeverity.WARNING.value:
                snapshot.warnings += 1

            if category in {IssueCategory.PHOTOS.value, IssueCategory.VIDEO.value}:
                snapshot.photos += 1
            if category == IssueCategory.PRICE.value or "цена" in title.lower():
                snapshot.price += 1
            if category in {
                IssueCategory.TITLE.value,
                IssueCategory.DESCRIPTION.value,
                IssueCategory.SEO.value,
                IssueCategory.CHARACTERISTICS.value,
            }:
                snapshot.text += 1
            if "сертифик" in field_path or "декларац" in field_path or "тнвэд" in field_path or "ндс" in field_path:
                snapshot.docs += 1
            if title and title not in top_titles and len(top_titles) < 3:
                top_titles.append(title)
        snapshot.top_titles = top_titles
        return snapshot

    def _resolve_precision(
        self,
        advert: _AdvertMetrics,
        raw_unallocated: float,
    ) -> Tuple[str, str]:
        exact = advert.exact_spend > 0.01
        estimated = advert.estimated_spend > 0.01
        manual = advert.manual_spend > 0.01
        if exact and not estimated and not manual:
            return "exact", "Точные данные"
        if estimated and not exact and not manual:
            return "estimated", "Оценка"
        if manual and not exact and not estimated:
            return "manual", "Загружено вручную"
        if exact or estimated or manual:
            return "mixed", "Смешанный источник"
        if raw_unallocated > 0.01:
            return "unallocated", "Нераспределенные расходы"
        return "exact", "Точные данные"

    def _resolve_ad_cost_confidence(self, advert: _AdvertMetrics, raw_unallocated: float) -> str:
        if advert.manual_spend > 0.01:
            return "low"
        if raw_unallocated > self.ADVERT_RESIDUAL_EPSILON:
            return "low"
        if advert.estimated_spend > 0.01:
            return "medium"
        return "high"

    def _lineage_mode_from_source_status(self, source_status: Optional[AdAnalysisSourceStatusOut]) -> str:
        if source_status is None:
            return "failed"
        mode = str(source_status.mode or "")
        if mode in {"error", "manual_required"}:
            return "failed"
        if mode == "partial":
            return "partial"
        if mode == "manual" or not bool(source_status.automatic):
            return "manual"
        return "automatic"

    def _build_source_lineage(
        self,
        source_statuses: Dict[str, AdAnalysisSourceStatusOut] | Sequence[AdAnalysisSourceStatusOut],
    ) -> AdAnalysisSourceLineageOut:
        by_id: Dict[str, AdAnalysisSourceStatusOut] = {}
        if isinstance(source_statuses, dict):
            by_id = {
                str(key): value
                for key, value in source_statuses.items()
                if isinstance(value, AdAnalysisSourceStatusOut)
            }
        else:
            for source in source_statuses:
                if isinstance(source, AdAnalysisSourceStatusOut):
                    by_id[str(source.id)] = source

        return AdAnalysisSourceLineageOut(
            advert=self._lineage_mode_from_source_status(by_id.get("advert")),
            finance=self._lineage_mode_from_source_status(by_id.get("finance")),
            funnel=self._lineage_mode_from_source_status(by_id.get("funnel")),
        )

    def _resolve_diagnosis(
        self,
        *,
        metrics: AdAnalysisMetricsOut,
        issue_summary: _IssueSnapshot,
        missing_cost: bool,
        finance_ready: bool,
    ) -> Tuple[str, str, str, str]:
        if missing_cost or not finance_ready:
            return (
                "data",
                "Проблема в данных",
                "Недостаточно данных для точной экономики. Не хватает себестоимости или финансового слоя, поэтому вывод по прибыли пока приблизительный.",
                "Сначала дозагрузите недостающие данные, потом принимайте решение по бюджету.",
            )

        orders_total = max(metrics.order_count, metrics.ad_orders)
        cpo_ratio = (metrics.actual_cpo / metrics.max_cpo) if metrics.max_cpo > 0 else float("inf")
        if metrics.gross_profit_before_ads <= 0:
            return (
                "economics",
                "Проблема в экономике",
                f"Маржа товара слишком слабая даже до рекламы. Gross Profit до рекламы {metrics.gross_profit_before_ads:.0f} ₽, значит реклама только ускоряет убыток.",
                "Проверьте цену, себестоимость, скидки и комиссии WB прежде чем продолжать открутку.",
            )

        card_conversion_low = (
            (metrics.add_to_cart_percent > 0 and metrics.add_to_cart_percent < 7.0)
            or (metrics.cr > 0 and metrics.cr < 2.0)
        )
        traffic_low = (
            (metrics.ctr > 0 and metrics.ctr < 1.3)
            or (metrics.clicks >= 20 and metrics.cpc > 0 and metrics.max_cpo > 0 and metrics.cpc >= metrics.max_cpo * 0.25)
        )
        card_has_issues = issue_summary.photos > 0 or issue_summary.price > 0 or issue_summary.text > 1 or issue_summary.critical > 0

        if (metrics.clicks >= 25 and orders_total <= 1 and not traffic_low) or (card_conversion_low and card_has_issues):
            return (
                "card",
                "Проблема в карточке",
                f"Люди доходят до карточки, но почти не покупают. CTR {metrics.ctr:.1f}% выглядит рабочим, а CR всего {metrics.cr:.1f}%, поэтому проблема вероятнее в карточке или цене.",
                "Проверьте фото, оффер, цену, СПП и первые экраны карточки.",
            )

        if traffic_low:
            return (
                "traffic",
                "Проблема в трафике",
                f"Трафик слишком слабый или дорогой. CTR {metrics.ctr:.1f}% и CPC {metrics.cpc:.0f} ₽ не дают качественного входа в карточку.",
                "Проверьте ставки, семантику, поисковые фразы и креативы.",
            )

        if cpo_ratio >= 0.85:
            return (
                "economics",
                "Проблема в экономике",
                f"Реклама еще работает, но уперлась в предел маржи. Actual CPO {metrics.actual_cpo:.0f} ₽ уже почти равен Max CPO {metrics.max_cpo:.0f} ₽.",
                "Снижайте стоимость заказа или не увеличивайте бюджет, пока не появится запас по марже.",
            )

        return (
            "economics",
            "Проблема в экономике",
            f"Экономика держится на пределе. Profit Delta {metrics.profit_delta:.0f} ₽ показывает, сколько еще запаса осталось между Max CPO и фактической стоимостью заказа.",
            "Сравните допустимый CPO с текущим и решите: снижать расходы или пока держать под контролем.",
        )

    def _resolve_status(
        self,
        *,
        metrics: AdAnalysisMetricsOut,
        diagnosis: str,
        missing_cost: bool,
        finance_ready: bool,
    ) -> Tuple[str, str]:
        orders_total = max(metrics.order_count, metrics.ad_orders)
        cpo_ratio = (metrics.actual_cpo / metrics.max_cpo) if metrics.max_cpo > 0 else float("inf")

        if missing_cost or not finance_ready:
            return "low_data", "Мало данных"
        if metrics.gross_profit_before_ads <= 0 and orders_total >= 1:
            return "stop", "Остановить"
        if metrics.net_profit < 0 or metrics.profit_delta < 0 or (metrics.max_cpo <= 0 and metrics.ad_cost > 0):
            return "stop", "Остановить"
        if metrics.ad_cost <= 0.01:
            return "low_data", "Мало данных"
        if orders_total < 3:
            return "low_data", "Мало данных"
        if cpo_ratio >= 0.85:
            return "rescue", "Спасти"
        if cpo_ratio <= 0.70 and metrics.net_profit > 0 and orders_total >= 3:
            return "grow", "Растить"
        return "control", "Контролировать"

    def _resolve_priority(
        self,
        *,
        status: str,
        diagnosis: str,
        metrics: AdAnalysisMetricsOut,
        trend_signal: str,
    ) -> Tuple[str, str]:
        if status == "stop" or metrics.net_profit < 0 or metrics.profit_delta < 0:
            return "critical", "Критично"
        if status == "rescue" or trend_signal == "worsening":
            return "high", "Высокий приоритет"
        if status in {"control", "low_data"} or trend_signal == "volatile":
            return "medium", "Средний приоритет"
        return "low", "Низкий приоритет"

    def _build_actions(
        self,
        *,
        status: str,
        diagnosis: str,
        issue_summary: _IssueSnapshot,
        missing_cost: bool,
        precision: str,
        source_statuses: Dict[str, AdAnalysisSourceStatusOut],
        metrics: AdAnalysisMetricsOut,
    ) -> Tuple[str, str, List[str], List[str], List[str]]:
        risk_flags: List[str] = []
        insights: List[str] = []
        steps: List[str] = []

        if missing_cost:
            risk_flags.append("Нет себестоимости: прибыль и Max CPO завышены")
        if source_statuses["finance"].mode in {"error", "manual_required"}:
            risk_flags.append("Нет финансового отчета WB: revenue и wb costs рассчитаны по упрощенной схеме")
        if precision in {"estimated", "mixed", "unallocated"}:
            risk_flags.append("Расходы по рекламе привязаны не полностью или частично оценочно")
        if issue_summary.top_titles:
            insights.append("Главные проблемы карточки: " + ", ".join(issue_summary.top_titles))
        if metrics.ctr > 0:
            insights.append(f"CTR: {metrics.ctr:.1f}%")
        if metrics.cr > 0:
            insights.append(f"CR из рекламы: {metrics.cr:.1f}%")
        if metrics.add_to_cart_percent > 0:
            insights.append(f"Конверсия в корзину: {metrics.add_to_cart_percent:.1f}%")

        if status == "stop":
            steps.extend([
                "Остановите рекламу по SKU и снимите бюджет с убыточной позиции.",
                "Проверьте Gross Profit до рекламы: хватает ли у товара маржи без рекламных расходов.",
                "Если Gross Profit <= 0, исправляйте цену, скидку, себестоимость или экономику товара.",
                "Если Gross Profit > 0, перезапускайте рекламу только после снижения ставок и повторного теста.",
            ])
            return (
                "Остановить рекламу по SKU",
                "SKU уже уводит деньги в минус. Сначала прекратите потери, потом решайте, можно ли вернуть рекламу после исправлений.",
                steps,
                insights,
                risk_flags,
            )
        if status == "rescue":
            steps.extend([
                "Снизьте ставку или бюджет на 10-15%, чтобы вернуть запас по CPO.",
                "Сравните CTR и CR, чтобы понять, проблема в трафике или в карточке.",
                "Если люди кликают, но не покупают, проверьте фото, цену, оффер и первые экраны карточки.",
                "Если CTR слабый, пересоберите запросы, креатив или тип кампании.",
            ])
            return (
                "Спасти SKU до ухода в минус",
                "Запас по прибыли почти съеден. Нужно быстро снизить стоимость заказа или исправить карточку, пока SKU не стал убыточным.",
                steps,
                insights,
                risk_flags,
            )
        if status == "grow":
            steps.extend([
                "Увеличьте бюджет по SKU на 10-20% без резких скачков.",
                "Следите за Actual CPO и Profit Delta после роста бюджета.",
                "Если экономика остается стабильной, повторите увеличение постепенно.",
            ])
            return (
                "Можно аккуратно растить бюджет",
                "SKU выдерживает рекламу и сохраняет запас по прибыли. Рост лучше делать ступенчато, а не одним резким увеличением.",
                steps,
                insights,
                risk_flags,
            )
        if status == "control":
            steps.extend([
                "Не меняйте бюджет резко и держите SKU под наблюдением.",
                "Следите за Profit Delta, Actual CPO и количеством заказов.",
                "Если тренд пойдет вниз, переводите SKU в сценарий спасения до ухода в минус.",
            ])
            return (
                "Держать под контролем",
                "SKU пока в рабочей зоне, но запас по марже не настолько большой, чтобы масштабировать его без наблюдения.",
                steps,
                insights,
                risk_flags,
            )
        if missing_cost or source_statuses["finance"].mode in {"error", "manual_required"}:
            steps.extend([
                "Загрузите недостающий файл или обновите проблемный источник данных.",
                "После загрузки пересчитайте snapshot, чтобы система получила точную прибыль и Max CPO.",
                "Принимайте решение по бюджету только после пересчета.",
            ])
            return (
                "Сначала закрыть недостающие данные",
                "По этому SKU пока не хватает обязательных слоев для точной экономики. Сначала закройте данные, потом возвращайтесь к решению по бюджету.",
                steps,
                insights,
                risk_flags,
            )
        if metrics.ad_cost <= 0.01:
            steps.extend([
                "Проверьте, запускалась ли реклама по SKU в выбранном периоде.",
                "Если теста еще не было, дайте SKU первый небольшой рекламный бюджет.",
                "После появления кликов и заказов система сможет дать управленческий статус.",
            ])
            return (
                "SKU еще не прошел рекламный тест",
                "Сейчас по SKU нет достаточного рекламного расхода, поэтому система не может оценить допустимый CPO и решение по бюджету.",
                steps,
                insights,
                risk_flags,
            )
        steps.extend([
            "Не меняйте бюджет резко, пока не накопится хотя бы 3-5 заказов с рекламы.",
            "Проверьте, растут ли показы, клики и первые заказы после текущих изменений.",
            "После накопления статистики пересчитайте статус SKU еще раз.",
        ])
        return (
            "Набрать больше статистики",
            "Данные уже есть, но заказов пока слишком мало для уверенного решения. Дайте SKU накопить статистику, прежде чем усиливать или отключать рекламу.",
            steps,
            insights,
            risk_flags,
        )

    def _build_alerts(
        self,
        *,
        remaining_unallocated: float,
        missing_cost_nm_ids: Sequence[int],
        finance_status: AdAnalysisSourceStatusOut,
        advert_status: AdAnalysisSourceStatusOut,
        items: Sequence[AdAnalysisItemOut],
    ) -> List[AdAnalysisAlertOut]:
        alerts: List[AdAnalysisAlertOut] = []
        if remaining_unallocated > 0.01:
            alerts.append(
                AdAnalysisAlertOut(
                    level="warning",
                    title=f"{_format_money(remaining_unallocated)} расходов на рекламу не удалось привязать к SKU",
                    description="WB не отдал nmID по части кампаний. Загрузите ручное распределение расходов, чтобы карта экономики была полной.",
                    action="Загрузить ручное распределение",
                )
            )
        if missing_cost_nm_ids:
            alerts.append(
                AdAnalysisAlertOut(
                    level="warning",
                    title=f"Для {len(set(missing_cost_nm_ids))} SKU не хватает себестоимости",
                    description="Без себестоимости прибыль и Max CPO считаются приблизительно. Загрузите файл по себестоимости.",
                    action="Загрузить себестоимость",
                )
            )
        if finance_status.mode in {"error", "manual_required"}:
            alerts.append(
                AdAnalysisAlertOut(
                    level="warning",
                    title="Нет финансового отчета WB по API",
                    description=finance_status.detail or "Загрузите выгрузку отчета реализации вручную.",
                    action="Загрузить финансовый файл",
                )
            )
        if advert_status.mode == "error":
            alerts.append(
                AdAnalysisAlertOut(
                    level="error",
                    title="Не удалось загрузить рекламную статистику из WB",
                    description=advert_status.detail or "Загрузите распределение расходов вручную.",
                    action="Загрузить рекламный файл",
                )
            )
        if not alerts and items:
            alerts.append(
                AdAnalysisAlertOut(
                    level="success",
                    title="Данные подтянуты и расчет готов",
                    description="WB-источники и ручные файлы собраны в единый SKU economics отчет.",
                    action=None,
                )
            )
        return alerts

    def _build_budget_moves(self, items: Sequence[AdAnalysisItemOut]) -> List[AdAnalysisBudgetMoveOut]:
        losers = [item for item in items if item.status == "stop" and item.metrics.ad_cost > 0]
        growers = [item for item in items if item.status == "grow" and item.metrics.profit_delta > 0]
        moves: List[AdAnalysisBudgetMoveOut] = []
        if not losers or not growers:
            return moves

        top_losers = losers[:3]
        top_growers = growers[:3]
        for idx, grower in enumerate(top_growers):
            source = top_losers[idx % len(top_losers)]
            uplift = 10
            if grower.metrics.actual_cpo > 0:
                uplift = int(max(10, min(40, round(grower.metrics.profit_delta / grower.metrics.actual_cpo * 100))))
            moves.append(
                AdAnalysisBudgetMoveOut(
                    from_nm_id=source.nm_id,
                    from_title=source.title or f"nmID {source.nm_id}",
                    from_amount=round(source.metrics.ad_cost, 2),
                    to_nm_id=grower.nm_id,
                    to_title=grower.title or f"nmID {grower.nm_id}",
                    uplift_percent=uplift,
                )
            )
        return moves

    def _resolve_trend(
        self,
        current: AdAnalysisItemOut,
        previous: Optional[AdAnalysisItemOut],
    ) -> AdAnalysisTrendOut:
        if previous is None:
            return AdAnalysisTrendOut(
                signal="new",
                label="Новый SKU",
                summary="В предыдущем окне сравнения по этому SKU не было достаточных данных.",
            )

        current_orders = max(current.metrics.order_count, current.metrics.ad_orders)
        previous_orders = max(previous.metrics.order_count, previous.metrics.ad_orders)
        actual_cpo_change = round(current.metrics.actual_cpo - previous.metrics.actual_cpo, 2)
        net_profit_change = round(current.metrics.net_profit - previous.metrics.net_profit, 2)
        profit_delta_change = round(current.metrics.profit_delta - previous.metrics.profit_delta, 2)
        orders_change = int(current_orders - previous_orders)
        ctr_change = round(current.metrics.ctr - previous.metrics.ctr, 2)
        cr_change = round(current.metrics.cr - previous.metrics.cr, 2)

        worsening = 0
        improving = 0
        if actual_cpo_change > max(100.0, abs(previous.metrics.actual_cpo) * 0.15):
            worsening += 1
        elif actual_cpo_change < -max(100.0, abs(previous.metrics.actual_cpo) * 0.15):
            improving += 1

        if profit_delta_change < -max(100.0, abs(previous.metrics.profit_delta) * 0.15):
            worsening += 1
        elif profit_delta_change > max(100.0, abs(previous.metrics.profit_delta) * 0.15):
            improving += 1

        if net_profit_change < -max(300.0, abs(previous.metrics.net_profit) * 0.15):
            worsening += 1
        elif net_profit_change > max(300.0, abs(previous.metrics.net_profit) * 0.15):
            improving += 1

        if orders_change <= -2:
            worsening += 1
        elif orders_change >= 2:
            improving += 1

        if ctr_change < -0.3 or cr_change < -0.3:
            worsening += 1
        elif ctr_change > 0.3 or cr_change > 0.3:
            improving += 1

        if worsening >= 2 and improving == 0:
            summary = (
                f"SKU ухудшается: Actual CPO вырос на {_format_money(actual_cpo_change)}, "
                f"Profit Delta изменился на {_format_money(profit_delta_change)}."
            )
            return AdAnalysisTrendOut(
                signal="worsening",
                label="Ухудшается",
                summary=summary,
                actual_cpo_change=actual_cpo_change,
                net_profit_change=net_profit_change,
                profit_delta_change=profit_delta_change,
                orders_change=orders_change,
                ctr_change=ctr_change,
                cr_change=cr_change,
            )
        if improving >= 2 and worsening == 0:
            summary = (
                f"SKU улучшается: Net Profit изменился на {_format_money(net_profit_change)}, "
                f"а Profit Delta вырос на {_format_money(profit_delta_change)}."
            )
            return AdAnalysisTrendOut(
                signal="improving",
                label="Улучшается",
                summary=summary,
                actual_cpo_change=actual_cpo_change,
                net_profit_change=net_profit_change,
                profit_delta_change=profit_delta_change,
                orders_change=orders_change,
                ctr_change=ctr_change,
                cr_change=cr_change,
            )
        if worsening > 0 and improving > 0:
            return AdAnalysisTrendOut(
                signal="volatile",
                label="Нестабильно",
                summary="Метрики двигаются в разные стороны: резкие решения по бюджету лучше не принимать.",
                actual_cpo_change=actual_cpo_change,
                net_profit_change=net_profit_change,
                profit_delta_change=profit_delta_change,
                orders_change=orders_change,
                ctr_change=ctr_change,
                cr_change=cr_change,
            )
        return AdAnalysisTrendOut(
            signal="stable",
            label="Стабильно",
            summary="Сильных изменений по сравнению с предыдущим окном нет.",
            actual_cpo_change=actual_cpo_change,
            net_profit_change=net_profit_change,
            profit_delta_change=profit_delta_change,
            orders_change=orders_change,
            ctr_change=ctr_change,
            cr_change=cr_change,
        )

    def _apply_trends(
        self,
        current: AdAnalysisOverviewOut,
        previous: AdAnalysisOverviewOut,
    ) -> None:
        previous_map = {item.nm_id: item for item in previous.items}
        worsening_count = 0
        improving_count = 0
        window_days = max((current.period_end - current.period_start).days + 1, 1)

        for item in current.items:
            trend = self._resolve_trend(item, previous_map.get(item.nm_id))
            if trend.signal == "worsening" and item.metrics.net_profit > 0 and trend.net_profit_change < 0:
                daily_drop = abs(trend.net_profit_change) / window_days
                if daily_drop > 0:
                    days_to_negative = max(int(round(item.metrics.net_profit / daily_drop)), 1)
                    trend.summary = (
                        f"{trend.summary} Если ничего не менять, SKU может уйти в минус примерно через {days_to_negative} дн."
                    )
            item.trend = trend
            item.priority, item.priority_label = self._resolve_priority(
                status=item.status,
                diagnosis=item.diagnosis,
                metrics=item.metrics,
                trend_signal=trend.signal,
            )
            if trend.signal == "worsening":
                worsening_count += 1
            elif trend.signal == "improving":
                improving_count += 1

        current.worsening_count = worsening_count
        current.improving_count = improving_count
        current.main_takeaway = self._build_main_takeaway(current)
        current.items.sort(key=self._item_sort_key)

    def _build_main_takeaway(self, overview: AdAnalysisOverviewOut) -> str:
        stop_count = overview.status_counts.get("stop", 0)
        rescue_count = overview.status_counts.get("rescue", 0)
        grow_count = overview.status_counts.get("grow", 0)
        if stop_count > 0:
            return f"{stop_count} SKU уже теряют деньги после рекламы. Их нужно останавливать первыми и перераспределять бюджет."
        if rescue_count > 0:
            return f"{rescue_count} SKU близки к пределу по CPO. Сейчас важнее спасать карточку или ставки, чем масштабировать рекламу."
        if overview.worsening_count > 0:
            return f"{overview.worsening_count} SKU ухудшаются по сравнению с прошлым окном. Их стоит проверить до того, как они уйдут в минус."
        if grow_count > 0:
            return f"{grow_count} SKU держат запас по прибыли и готовы к аккуратному росту бюджета."
        return "Картина по SKU в целом стабильная: критичных потерь не видно, но решения лучше принимать по приоритету и тренду."

    async def _persist_snapshots(
        self,
        db: AsyncSession,
        store_id: int,
        overview: AdAnalysisOverviewOut,
    ) -> None:
        if not overview.items:
            return
        nm_ids = [item.nm_id for item in overview.items]
        existing = await db.execute(
            select(SkuEconomicsSnapshot).where(
                SkuEconomicsSnapshot.store_id == int(store_id),
                SkuEconomicsSnapshot.period_start == overview.period_start,
                SkuEconomicsSnapshot.period_end == overview.period_end,
                SkuEconomicsSnapshot.nm_id.in_(nm_ids),
            )
        )
        existing_map = {int(row.nm_id): row for row in existing.scalars().all()}

        for item in overview.items:
            row = existing_map.get(int(item.nm_id))
            if row is None:
                row = SkuEconomicsSnapshot(
                    store_id=int(store_id),
                    nm_id=int(item.nm_id),
                    period_start=overview.period_start,
                    period_end=overview.period_end,
                )
                db.add(row)
                existing_map[int(item.nm_id)] = row
            row.title = item.title
            row.vendor_code = item.vendor_code
            row.status = item.status
            row.diagnosis = item.diagnosis
            row.priority = item.priority
            row.precision = item.precision
            row.revenue = float(item.metrics.revenue or 0.0)
            row.ad_cost = float(item.metrics.ad_cost or 0.0)
            row.net_profit = float(item.metrics.net_profit or 0.0)
            row.max_cpo = float(item.metrics.max_cpo or 0.0)
            row.actual_cpo = float(item.metrics.actual_cpo or 0.0)
            row.profit_delta = float(item.metrics.profit_delta or 0.0)
            row.ctr = float(item.metrics.ctr or 0.0)
            row.cr = float(item.metrics.cr or 0.0)
            row.orders = int(max(item.metrics.order_count, item.metrics.ad_orders))
            row.ad_orders = int(item.metrics.ad_orders or 0)
            row.generated_at = overview.generated_at

        await db.commit()

    def _item_sort_key(self, item: AdAnalysisItemOut) -> Tuple[int, float]:
        priority_order = {"critical": 0, "high": 1, "medium": 2, "low": 3}
        status_order = {"stop": 0, "rescue": 1, "control": 2, "low_data": 3, "grow": 4}
        trend_order = {"worsening": 0, "volatile": 1, "stable": 2, "improving": 3, "new": 4, "no_history": 5}
        score = item.metrics.net_profit if item.status == "grow" else -item.metrics.net_profit
        return (
            priority_order.get(item.priority, 99),
            status_order.get(item.status, 99),
            trend_order.get(item.trend.signal, 99),
            score,
        )


sku_economics_service = SkuEconomicsService()
